import os
import sys
import traceback

PROJECT_PATH = os.getcwd()
sys.path.append(PROJECT_PATH)

from library.study_space import Study
from scripts.batch_map.LEC_naming import LEC_naming_format, extract_name_lec
import tkinter as tk
from tkinter import filedialog
import time
from openpyxl.utils.cell import get_column_letter
from library.maps.map_utils import disk_mask
import numpy as np
from scipy import ndimage
import imageio
from skimage import color
import skimage.measure
import matplotlib.pyplot as plt
import openpyxl as xl
from openpyxl.worksheet.dimensions import ColumnDimension
from x_io.rw.axona.batch_read import make_study
from x_io.rw.axona.read_pos import _get_position
import pandas as pd
import re

""" SETTINGS AT THE BOTTOM OF FILE """

def _check_single_format(filename, format, fxn):
    print(filename, format, fxn)
    if re.match(str(format), str(filename)) is not None:
        return fxn(filename)

def flat_disk_mask(rate_map):
    masked_rate_map = disk_mask(rate_map)
    masked_rate_map.data[masked_rate_map.mask] = np.nan
    return  masked_rate_map.data

def batch_map(study: Study, settings_dict: dict, saveDir=None, sum_sheet_count=None):
    """
    Computes rate maps across all animals, sessions, cells in a study.

    Use tasks dictionary as true/false flag with variable to compute
    e.g. {'rate_map': True, 'binary_map': False}
    """

    tasks = settings_dict['tasks']
    plotTasks = settings_dict['plotTasks']
    csv_header = settings_dict['header']

    # Root path creation
    save_dir = saveDir
    root_path = os.path.join(save_dir, 'Animal_Performance_')
    run_number = 1
    while os.path.isdir(root_path+str(run_number)):
        run_number+=1

    root_path += str(run_number)
    os.mkdir(root_path)


    # Grabs headers whose value is true
    headers = [k for k, v in csv_header.items() if v]

    headers_dict = dict()

    for i, header in enumerate(headers):
        headers_dict[header] = get_column_letter(i+1)

    file_name = os.path.join(root_path, "ap_parameters.txt")
    with open(file_name, 'w') as f:
        for ky in settings_dict:
            f.write(str(ky) + ' is: ')
            f.write(str(settings_dict[ky]) + '\n')
        f.close()

    if study.animals is None:
        study.make_animals()
        print('Animals made, batching map')

    per_animal_tracker = 1

    animal_tet_count = {}
    animal_max_tet_count = {}
    animal_workbooks = {}
    sorted_animal_ids = np.unique(np.sort(study.animal_ids))

    one_for_parent_wb = xl.Workbook()
    sum_sheet = one_for_parent_wb['Sheet']
    sum_sheet.title = 'Summary'
    # sum_sheet['A' + str(1)] = 'Session'
    visited = []
    for animalID in sorted_animal_ids:
        animal = study.get_animal_by_id(animalID)

        animal_id = animal.animal_id.split('_tet')[0]
        if animal_id not in animal_tet_count:
            animal_tet_count[animal_id] = 1
            # animal_sessions_tets_events[animal_id] = {}
            wb = xl.Workbook()
            animal_workbooks[animal_id] = wb
            animal_max_tet_count[animal_id] = len(sorted_animal_ids[list(map(lambda x: True if animal_id in x else False, sorted_animal_ids))])
            sum_sheet = wb['Sheet']
            sum_sheet.title = 'Summary'

            wb = one_for_parent_wb
            sum_sheet = wb['Summary']
            for header in headers:
                sum_sheet[headers_dict[header] + str(1)] = header
        else:
            wb = one_for_parent_wb

        k = 1

        for session_key in animal.sessions:
            session = animal.sessions[session_key]

            c = 0

            pos_obj = session.get_position_data()['position']
            spike_cluster = session.get_spike_data()["spike_cluster"]
            spike_param_dict = spike_cluster.spikeparam


            dateandtime = spike_param_dict["datetime"] # datetime object (NOT SAVED)
            trial_time = str(dateandtime.strftime("%H:%M:%S")) # STR

            pos_file = session.session_metadata.file_paths['pos']

            # root_path
            ap_stats = get_animal_performance(pos_file, root_path, settings_dict)

            tet_file = session.session_metadata.file_paths['tet']

            # Session stamp
            signature = tet_file.split("/")[-1][:-2]

            if signature not in visited:
                visited.append(signature)

                group, name = extract_name_lec(signature)
                formats = LEC_naming_format[group][name]['object']

                for format in list(formats.keys()):
                    checked = _check_single_format(signature, format, formats[format])
                    if checked is not None:
                        break
                    else:
                        continue
                
                stim, depth, name, date = checked


                current_statistics_sheet = wb['Summary']


                excel_cell_index = per_animal_tracker
                current_statistics_sheet[headers_dict['signature'] + str(excel_cell_index+1)] = signature
                current_statistics_sheet[headers_dict['tetrode'] + str(excel_cell_index+1)] = tet_file[-1]
                current_statistics_sheet[headers_dict['date'] + str(excel_cell_index+1)] = date 
                current_statistics_sheet[headers_dict['trial_time'] + str(excel_cell_index+1)] = trial_time
                current_statistics_sheet[headers_dict['name'] + str(excel_cell_index+1)] = name
                current_statistics_sheet[headers_dict['depth'] + str(excel_cell_index+1)] = depth  
                current_statistics_sheet[headers_dict['stim'] + str(excel_cell_index+1)] = stim
                current_statistics_sheet[headers_dict['total_area'] + str(excel_cell_index+1)] = ap_stats[0]
                current_statistics_sheet[headers_dict['total_filled'] + str(excel_cell_index+1)] = ap_stats[1]
                current_statistics_sheet[headers_dict['coverage'] + str(excel_cell_index+1)] = ap_stats[2]
                current_statistics_sheet[headers_dict['min_speed'] + str(excel_cell_index+1)] = np.min(ap_stats[4])
                current_statistics_sheet[headers_dict['min_speed_units'] + str(excel_cell_index+1)] = 'cm/s'
                current_statistics_sheet[headers_dict['max_speed'] + str(excel_cell_index+1)] = np.max(ap_stats[4])
                current_statistics_sheet[headers_dict['max_speed_units'] + str(excel_cell_index+1)] = 'cm/s'
                current_statistics_sheet[headers_dict['mean_speed'] + str(excel_cell_index+1)] = np.mean(ap_stats[4])
                current_statistics_sheet[headers_dict['mean_speed_units'] + str(excel_cell_index+1)] = 'cm/s'
                current_statistics_sheet[headers_dict['median_speed'] + str(excel_cell_index+1)] = np.median(ap_stats[4])
                current_statistics_sheet[headers_dict['median_speed_units'] + str(excel_cell_index+1)] = 'cm/s'
                current_statistics_sheet[headers_dict['distance'] + str(excel_cell_index+1)] = ap_stats[3]
                current_statistics_sheet[headers_dict['distance_units'] + str(excel_cell_index+1)] = 'cm'
                       
                ColumnDimension(current_statistics_sheet, bestFit=True)
                per_animal_tracker += 1

            k += 1

    _save_wb(wb, root_path, sum_sheet_count=sum_sheet_count)


def centerBox(posx, posy):
    # must remove Nans first because the np.amin will return nan if there is a nan
    posx = posx[~np.isnan(posx)]  # removes NaNs
    posy = posy[~np.isnan(posy)]  # remove Nans

    NE = np.array([np.amax(posx), np.amax(posy)])
    NW = np.array([np.amin(posx), np.amax(posy)])
    SW = np.array([np.amin(posx), np.amin(posy)])
    SE = np.array([np.amax(posx), np.amin(posy)])

    return findCenter(NE, NW, SW, SE)


def findCenter(NE, NW, SW, SE):
    """Finds the center point (x,y) of the position boundaries"""

    x = np.mean([np.amax([NE[0], SE[0]]), np.amin([NW[0], SW[0]])])
    y = np.mean([np.amax([NW[1], NE[1]]), np.amin([SW[1], SE[1]])])
    return np.array([x, y])


def bits2uV(data, data_fpath, set_fpath=''):
    '''

    :param data:
    :param data_fpath: example: 'C:\example\filepath.whatever'
    :param set_fpath:
    :return:
    '''
    path = os.path.split(data_fpath)[0]

    if set_fpath == '':
        set_fpath = os.path.join(path, ''.join([os.path.splitext(os.path.basename(data_fpath))[0],'.set']))

    ext = os.path.splitext(data_fpath)[1]

    if not os.path.exists(set_fpath):
        error_message = 'The following setpath does not exist, cannot convert to uV: %s' % (set_fpath)
        # raise TintException(error_message)
        raise FileNotFoundError(error_message)

    # create a tetrode map that has rows of channels that correspond to the same tetrode
    tet_map = np.asarray([np.arange(start,start+4) for start in np.arange(0, 32)*4])

    chan_gains = np.array([])
    saved_eeg = np.array([])
    eeg_chan_map = np.array([])

    with open(set_fpath, 'r') as f:
        for line in f:

            if 'ADC_fullscale_mv' in line:
                ADC_fullscale_mv = int(line.split(" ")[1])
            elif 'gain_ch_' in line:
                # create an array of channel gains [channel_number, channels_gain]
                if len(chan_gains) == 0:
                    chan_gains = np.array([int(line[len('gain_ch_'):line.find(" ")]), int(line.split(" ")[1])], ndmin=2)
                else:
                    chan_gains = np.append(chan_gains, np.array([int(line[len('gain_ch_'):line.find(" ")]), int(line.split(" ")[1])], ndmin=2), axis=0)
            elif 'saveEEG_ch_' in line:
                # create an array of EEG channels that are saved
                if int(line.split(" ")[1]) == 1:
                    if len(chan_gains) == 0:
                        saved_eeg = np.array([int(line[len('saveEEG_ch_'):line.find(" ")])])
                    else:
                        saved_eeg = np.append(saved_eeg, np.array([int(line[len('saveEEG_ch_'):line.find(" ")])]))
            elif 'EEG_ch_' in line and 'BPF' not in line:
                if len(eeg_chan_map) == 0:
                    eeg_chan_map = np.array([int(line[len('EEG_ch_'):line.find(" ")]), int(line.split(" ")[1])], ndmin=2)
                else:
                    eeg_chan_map = np.append(eeg_chan_map, np.array([int(line[len('EEG_ch_'):line.find(" ")]), int(line.split(" ")[1])], ndmin=2), axis=0)

    if '.eeg' in ext:
        if len(ext) == len('.eeg'):
            chan_num = 1
        else:
            chan_num = int(ext[len('.eeg'):])

        for index, value in enumerate(eeg_chan_map[:]):
            if value[0] == chan_num:
                eeg_chan = value[1] - 1
                break

        for index, value in enumerate(chan_gains):
            if value[0] == eeg_chan:
                gain = value[1]

        scalar = ADC_fullscale_mv*1000/(gain*128)
        if len(data) == 0:
            data_uV = []
        else:
            data_uV = np.multiply(data, scalar)
            #print(data_uV)

    elif '.egf' in ext:
        if len(ext) == len('.egf'):
            chan_num = 1
        else:
            chan_num = int(ext[len('.egf'):])

        for index, value in enumerate(eeg_chan_map[:]):
            if value[0] == chan_num:
                eeg_chan = value[1] - 1
                break

        for index, value in enumerate(chan_gains):
            if value[0] == eeg_chan:
                gain = value[1]
                break

        scalar = ADC_fullscale_mv*1000/(gain*32768)

        if len(data) == 0:
            data_uV = []
        else:
            data_uV = np.multiply(data, scalar)

    else:
        tetrode_num = int(ext[1:])

        tet_chans = tet_map[tetrode_num-1]

        gain = np.asarray([[gains[1] for gains in chan_gains if gains[0] == chan] for chan in tet_chans])

        scalar = (ADC_fullscale_mv*1000/(gain*128).reshape((1, len(gain))))[0]

        if len(data) == 0:
            data_uV = []
        else:
            data_uV = np.multiply(data, scalar)

    return data_uV, scalar


def remBadTrack(x, y, t, threshold):
    """function [x,y,t] = remBadTrack(x,y,t,treshold)

    % Indexes to position samples that are to be removed
   """

    remInd = []
    diffx = np.diff(x, axis=0)
    diffy = np.diff(y, axis=0)
    diffR = np.sqrt(diffx ** 2 + diffy ** 2)

    # the MATLAB works fine without NaNs, if there are Nan's just set them to threshold they will be removed later
    diffR[np.isnan(diffR)] = threshold # setting the nan values to threshold
    ind = np.where((diffR > threshold))[0]

    if len(ind) == 0:  # no bad samples to remove
        return x, y, t

    if ind[-1] == len(x):
        offset = 2
    else:
        offset = 1

    for index in range(len(ind) - offset):
        if ind[index + 1] == ind[index] + 1:
            # A single sample position jump, tracker jumps out one sample and
            # then jumps back to path on the next sample. Remove bad sample.
            remInd.append(ind[index] + 1)
        else:
            ''' Not a single jump. 2 possibilities:
             1. TrackerMetadata jumps out, and stay out at the same place for several
             samples and then jumps back.
             2. TrackerMetadata just has a small jump before path continues as normal,
             unknown reason for this. In latter case the samples are left
             untouched'''
            idx = np.where(x[ind[index] + 1:ind[index + 1] + 1 + 1] == x[ind[index] + 1])[0]
            if len(idx) == len(x[ind[index] + 1:ind[index + 1] + 1 + 1]):
                remInd.extend(
                    list(range(ind[index] + 1, ind[index + 1] + 1 + 1)))  # have that extra since range goes to end-1
    # keep_ind = [val for val in range(len(x)) if val not in remInd]
    keep_ind = np.setdiff1d(np.arange(len(x)), remInd)

    # avoid trying to slice with an invalid index
    if keep_ind[-1] == len(x):
        keep_ind = keep_ind[:-1]

    x = x[keep_ind]
    y = y[keep_ind]
    t = t[keep_ind]

    return x.reshape((len(x), 1)), y.reshape((len(y), 1)), t.reshape((len(t), 1))



def get_animal_performance(posfile, save_figures_directory, settings_dict):
    plot_linewidth = 5.334
    posx, posy, post, Fs_pos, _ = _get_position(posfile, ppm=settings_dict['ppm'])  # getting the mouse position
    session_name = os.path.splitext(os.path.basename(posfile))[0]
    # centering the positions
    center = centerBox(posx, posy)
    posx = posx - center[0]
    posy = posy - center[1]
    # Threshold for how far a mouse can move (100cm/s), in one sample (sampFreq = 50 Hz
    threshold = 100 / 50  # defining the threshold
    posx, posy, post = remBadTrack(posx, posy, post, threshold)  # removing bad tracks (faster than threshold)
    nonNanValues = np.where(np.isnan(posx) == False)[0]
    # removeing any NaNs
    post = post[nonNanValues]
    posx = posx[nonNanValues]
    posy = posy[nonNanValues]
    if len(posx) == 0:
        print('There are no valid positions (all NaNs)')
        # return
    # box car smoothing, closest we could get to replicating Tint's speeds
    B = np.ones((int(np.ceil(0.4 * Fs_pos)), 1)) / np.ceil(0.4 * Fs_pos)
    # posx = scipy.ndimage.correlate(posx, B, mode='nearest')
    posx = ndimage.convolve(posx, B, mode='nearest')
    # posy = scipy.ndimage.correlate(posy, B, mode='nearest')
    posy = ndimage.convolve(posy, B, mode='nearest')
    print('calculating the total distance for the .pos file: %s ' % posfile)
    diffX = np.diff(posx)
    diffY = np.diff(posy)
    dist_sample = np.sqrt(diffX ** 2 + diffY ** 2)
    total_distance = np.sum(dist_sample)
    print('Calculating coverage!')
    points = np.hstack((posx.reshape((len(posx), 1)),
                        posy.reshape((len(posy), 1))))
    dimensions = np.array([np.amin(posx), np.amax(posx),
                           np.amin(posy), np.amax(posy)])
    arena_fig = plt.figure()
    ax = arena_fig.add_subplot(111)
    mng = plt.get_current_fig_manager()
    # mng.resize(*mng.window.maxsize())
    ax.plot(posx, posy, 'r-', lw=plot_linewidth)
    # check if it is a square or a circle arena by looking at the corners
    bin_x = np.linspace(np.amin(posx), np.amax(posx), 11)
    bin_y = np.linspace(np.amin(posx), np.amax(posx), 11)
    corners = np.array([[bin_x[-2], bin_y[-2]], [bin_x[1], bin_y[-2]],
                        [bin_x[1], bin_y[1]], [bin_x[-2], bin_y[1]]])
    plot_corners = False
    if plot_corners:
        # plots to see the test if it is a square or cirlce. I essentially
        # determine if there is data at the corners of the arena
        fig_corners = plt.figure()
        ax_corners = fig_corners.add_subplot(111)
        ax_corners.plot(posx, posy, 'r-')
        ax_corners.plot(corners[0:2, 0], corners[0:2, 1], 'g')
        ax_corners.plot(corners[1:3, 0], corners[1:3, 1], 'g')
        ax_corners.plot(corners[[0, 3], 0], corners[[0, 3], 1], 'g')
        ax_corners.plot(corners[[2, 3], 0], corners[[2, 3], 1], 'g')
    circle_bool = []
    for corner in range(4):
        # for each corner, see if there is data in the corners
        if corner == 0:  # NE corner
            bool_val = (points[:, 0] >= 0) * (points[:, 1] >= 0)
            current_points = points[bool_val, :]
            circle_bool.append(np.sum((current_points[:, 0] >= corners[corner, 0]) *
                                      (current_points[:, 1] >= corners[corner, 1])))
        elif corner == 1:  # NW Corner
            bool_val = (points[:, 0] < 0) * (points[:, 1] >= 0)
            current_points = points[bool_val, :]
            circle_bool.append(np.sum((current_points[:, 0] < corners[corner, 0]) *
                                      (current_points[:, 1] >= corners[corner, 1])))
        elif corner == 2:  # SW Corner
            bool_val = (points[:, 0] < 0) * (points[:, 1] < 0)
            current_points = points[bool_val, :]
            circle_bool.append(np.sum((current_points[:, 0] <= corners[corner, 0]) *
                                      (current_points[:, 1] < corners[corner, 1])))
        else:  # SE corner
            bool_val = (points[:, 0] > 0) * (points[:, 1] < 0)
            current_points = points[bool_val, :]
            circle_bool.append(np.sum((current_points[:, 0] > corners[corner, 0]) *
                                      (current_points[:, 1] < corners[corner, 1])))
    coverage_figure = plt.figure()
    ax_coverage = coverage_figure.add_subplot(111)
    mng = plt.get_current_fig_manager()
    # mng.resize(*mng.window.maxsize())
    if sum(circle_bool) >= 1:
        # if there is data in the corners it is a square
        print('Arena detected as being square')
        # find the average heigh tand average width
        bins = np.linspace(np.amin(posx), np.amax(posx), 20)
        bin_edges = np.hstack((bins[0:-1].reshape((len(bins[0:-1]), 1)),
                               bins[1:].reshape((len(bins[1:]), 1))))
        # graphing rectangle representing the arena border
        rectangle_points = np.array([[np.amin(posx), np.amax(posy)],  # NW
                                     [np.amax(posx), np.amax(posy)],  # NE
                                     [np.amax(posx), np.amin(posy)],  # SE
                                     [np.amin(posx), np.amin(posy)],  # SW
                                     [np.amin(posx), np.amax(posy)]  # NW
                                     ])
        border = ax_coverage.plot(rectangle_points[:, 0], rectangle_points[:, 1], 'b', lw=plot_linewidth/10)
        ax_coverage.plot(posx, posy, 'r-', lw=plot_linewidth)
        ax_coverage.set_xlim([min([dimensions[0], np.amin(rectangle_points[:, 0])]) - 0.5,
                              max([dimensions[1], np.amax(rectangle_points[:, 0])]) + 0.5])
        ax_coverage.set_ylim([min([dimensions[2], np.amin(rectangle_points[:, 1])]) - 0.5,
                              max([dimensions[3], np.amax(rectangle_points[:, 1])]) + 0.5])
        total_area_cm2 = (np.abs(np.amin(posx)) +
                          np.amax(posx)) * (np.abs(np.amin(posy) +
                                                   np.amax(posy)))  # Length * Width
    else:
        # there were no position values in the corner, must be a circle
        print('Arena detected as being circular')
        bins = np.linspace(np.amin(posx), np.amax(posx), 50)
        bin_edges = np.hstack((bins[0:-1].reshape((len(bins[0:-1]), 1)),
                               bins[1:].reshape((len(bins[1:]), 1))))
        radii = np.array([np.abs(np.amin(posx)), np.amax(posx),
                          np.abs(np.amin(posy)), np.amax(posy)])
        for bin_value in range(len(bin_edges)):
            bin_bool = (posx >= bin_edges[bin_value, 0]) * (posx < bin_edges[bin_value, 1])
            if sum(bin_bool) == 0:
                # no points in this bin
                continue
            posx_bin = posx[bin_bool]
            posy_bin = posy[bin_bool]
            max_val = np.amax(posy_bin)
            max_i = np.where(posy_bin == max_val)[0][0]
            min_val = np.amin(posy_bin)
            min_i = np.where(posy_bin == min_val)[0][0]
            append_radii = np.array([np.sqrt(max_val ** 2 + posx_bin[max_i] ** 2),
                                     np.sqrt(min_val ** 2 + posx_bin[min_i] ** 2)])
            radii = np.concatenate((radii, append_radii))
        # equivalent to ang = 0:0.01:4*pi in matlab
        # we chose 4 pi as the end because 2 pi wasn't enough for edge detection in python
        step = 0.001
        ang = np.arange(np.round((4 * np.pi + step) / step)) / (1 / step)
        xp, yp = circle_vals(0, 0, 2 * np.amax(radii), ang)
        border = ax_coverage.plot(xp, yp, 'b', lw=plot_linewidth/10)
        ax_coverage.plot(posx, posy, 'r', lw=plot_linewidth)
        ax_coverage.set_xlim([min([dimensions[0], np.amin(xp)]) - 0.5,
                              max([dimensions[1], np.amax(xp)]) + 0.5])
        ax_coverage.set_ylim([min([dimensions[2], np.amin(yp)]) - 0.5,
                              max([dimensions[3], np.amax(yp)]) + 0.5])
        total_area_cm2 = np.pi * (np.amax(radii) ** 2)  # area = pr^2
    cover_png_total = os.path.join(save_figures_directory, '%s_total.png' % session_name)
    ax_coverage.axis('off')
    # coverage_figure.show()
    coverage_figure.savefig(cover_png_total, bbox_inches='tight')
    RGBA = imageio.imread(cover_png_total)
    try:
        RGB = color.rgba2rgb(RGBA)
    except ValueError:
        RGB = RGBA
    I = rgb2gray(RGB)
    I = np.round(I).astype('int32')
    # create a binary gradient mask of image
    BWs_x = ndimage.sobel(I, 0)  # horizontal derivative
    BWs_y = ndimage.sobel(I, 1)  # vertical derivative
    BWs = np.hypot(BWs_x, BWs_y)  # magnitude
    BWs *= 255.0 / np.amax(BWs)  # normalize (Q&D)
    struct1 = np.array([[False, True, False],
                        [True, True, True],
                        [False, True, False]])
    # create a dilated gradient mask
    BWsdil = ndimage.morphology.binary_dilation(BWs)
    # BWsdil = ndimage.morphology.binary_dilation(BWs, structure=struct1)
    # BWdfill = flood_fill(BWsdil)
    # BWdfill = ndimage.morphology.binary_fill_holes(BWsdil)
    BWdfill = ndimage.morphology.binary_fill_holes(BWs)
    # finding the contours so we can use contourArea to find the fill area
    #im2, contours, hierarchy = cv2.findContours(BWdfill.astype('uint8'), mode=cv2.RETR_LIST,
    #                                            method=cv2.CHAIN_APPROX_NONE)
    area = []
    label_img = skimage.measure.label(BWdfill)
    regions = skimage.measure.regionprops(label_img)
    for prop in regions:
        area.append(prop.area)
    #for cnt in contours:
    #    area.append(cv2.contourArea(cnt))
    # the area we are interested in is likely the only contour found, however,
    # if not it should most definitely be the largest one
    debug_figs = False
    if debug_figs:
        bwsdil_fig = plt.figure(figsize=(10, 10))
        plt.imshow(BWsdil, cmap=plt.cm.gray)
        # bwsdil_fig.show()
        bwsfill_fig = plt.figure(figsize=(10, 10))
        plt.imshow(BWdfill, cmap=plt.cm.gray)
        # bwsfill_fig.show()
        #filledI = np.zeros(BWdfill.shape[0:2]).astype('uint8')
        # set up the 'ConvexImage' bit of regionprops.
        #a = cv2.drawContours(filledI, contours, -1, (255, 0, 0), 3)
        #contours_fig = plt.figure(figsize=(10, 10))
        #plt.imshow(a, cmap=plt.cm.gray)
        #contours_fig.show()
    total_area = max(area)
    # self.border.remove()  # remove border
    border[0].remove()  # remove border
    cover_png = os.path.join(save_figures_directory, '%s_coverage.png' % session_name)
    coverage_figure.savefig(cover_png, bbox_inches='tight')  # save figure without border
    # reading in the positions without the arena trace
    RGBA = imageio.imread(cover_png)
    try:
        RGB = color.rgba2rgb(RGBA)
    except ValueError:
        RGB = RGBA
    I = rgb2gray(RGB)
    I = np.round(I).astype('int32')
    if np.amax(I) <= 1:
        # then the image was saved from numpy
        BWdfill = I < 1
    else:
        BWdfill = I < 255
    # finding the contours of the path so we can find the area
    '''
    im2, contours, hierarchy = cv2.findContours(BWdfill.astype('uint8'), mode=cv2.RETR_LIST,
                                                method=cv2.CHAIN_APPROX_TC89_KCOS)
    if debug_figs:
        bwsfill_fig2 = plt.figure(figsize=(10, 10))
        plt.imshow(BWdfill, cmap=plt.cm.gray)
        bwsfill_fig2.show()
        # set up the 'FilledImage' bit of regionprops.
        filledI = np.zeros(BWdfill.shape[0:2]).astype('uint8')
        # set up the 'ConvexImage' bit of regionprops.
        a = cv2.drawContours(filledI, contours, -1, (255, 0, 0), 3)
        contour_no_border = plt.figure()
        plt.imshow(a, cmap=plt.cm.gray)
        contour_no_border.show()
    '''
    coverage_fill_figure = plt.figure()
    ax_coverage_fill = coverage_fill_figure.add_subplot(111)
    mng = plt.get_current_fig_manager()
    # mng.resize(*mng.window.maxsize())
    ax_coverage_fill.imshow(BWdfill, cmap=plt.cm.gray)
    area = []
    label_img = skimage.measure.label(BWdfill)
    regions = skimage.measure.regionprops(label_img)
    for prop in regions:
        area.append(prop.area)
    #for cnt in contours:
    #    area.append(cv2.contourArea(cnt))
    # the area we are interested in is likely the only contour found, however,
    # if not it should most definitely be the largest one
    if len(regions) > 1:
        print('Too many regions!')
    total_filled = max(area)
    print('total area: %f' % total_area)
    print('total filled: %f' % total_filled)
    percent = total_filled / total_area
    print('Coverage(Percent): %f' % (100*total_filled/total_area))

    speed = speed2D(posx, posy, post)

    return total_area, total_filled, percent, total_distance, speed


def _save_wb(wb, root_path, animal_id=None, sum_sheet_count=None):
    wb._sheets = sorted(wb._sheets, key=lambda x: x.title)
    
    if animal_id is None:
        if sum_sheet_count is None:
            pth = root_path + '/animal_performance'  + '.xlsx'
        else:
            pth = root_path + '/animal_performance_'  + str(sum_sheet_count) + '.xlsx'
    else:
        pth = root_path + '/animal_performance_' + str(animal_id)  + '.xlsx'
    print(root_path)
    wb.save(pth)
    wb.close()

    # xls = pd.read_excel(pth, sheet_name=None)
    # df_sum = xls.pop('Summary')
    # dfs = [df.sort_values(['name','date','trial_time','depth','stim']) for df in xls.values()]
    # with pd.ExcelWriter(pth, engine='xlsxwriter') as writer:
    #     df_sum.to_excel(writer, sheet_name='Summary', index=False)
    #     for sheet, df in zip(xls.keys(), dfs):
    #         df.to_excel(writer, sheet_name=sheet, index=False)


    
    #     if len(dfs) > 0:
    #         df_sum = pd.concat(dfs, axis=0).sort_values(['name','date','trial_time','depth','stim'])
    #     else:
    #         df_sum = df_sum.sort_values(['name','date','trial_time','depth','stim'])
    #     df_sum.to_excel(writer, sheet_name="Summary", index=True)
    # writer.save()
    print('Saved ' + str(pth))

def rgb2gray(rgb):
    r, g, b = rgb[:, :, 0], rgb[:, :, 1], rgb[:, :, 2]
    gray = 0.2989 * r + 0.5870 * g + 0.1140 * b
    return gray

def get_hd_score_for_cluster(hd_hist):
    angles = np.linspace(-179, 180, 360)
    angles_rad = angles*np.pi/180
    dy = np.sin(angles_rad)
    dx = np.cos(angles_rad)

    totx = sum(dx * hd_hist)/sum(hd_hist)
    toty = sum(dy * hd_hist)/sum(hd_hist)
    r = np.sqrt(totx*totx + toty*toty)
    return r

def circle_vals(x, y, d, ang):
    """x and y are the coordinates of the center o the circle with radius,r"""
    r = d / 2
    xp = np.multiply(r, np.cos(ang))
    yp = np.multiply(r, np.sin(ang))
    xp = np.add(xp, x)
    yp = np.add(yp, y)
    return xp.reshape((len(xp), 1)), yp.reshape((len(yp), 1))

def speed2D(x, y, t):
    """calculates an averaged/smoothed speed"""

    N = len(x)
    v = np.zeros((N, 1))

    for index in range(1, N-1):
        v[index] = np.sqrt((x[index + 1] - x[index - 1]) ** 2 + (y[index + 1] - y[index - 1]) ** 2) / (
        t[index + 1] - t[index - 1])

    v[0] = v[1]
    v[-1] = v[-2]

    return v


if __name__ == '__main__':

    ######################################################## EDIT BELOW HERE ########################################################

    csv_header = {}
    csv_header_keys = ['signature', 'tetrode', 'name','date','trial_time','depth','stim', 'total_area', 'total_filled', 'coverage', 'min_speed',
                       'min_speed_units', 'max_speed', 'max_speed_units', 'mean_speed', 'mean_speed_units', 'median_speed',
                          'median_speed_units', 'distance', 'distance_units']
                       
    for key in csv_header_keys:
        csv_header[key] = True

    tasks = {}
    task_keys = []
    for key in task_keys:
        tasks[key] = True

    plotTasks = {}
    plot_task_keys = []
    for key in plot_task_keys:
        plotTasks[key] = True

    animal = {'animal_id': '001', 'species': 'mouse', 'sex': 'F', 'age': 1, 'weight': 1, 'genotype': 'type', 'animal_notes': 'notes'}
    devices = {'axona_led_tracker': True, 'implant': True}
    implant = {'implant_id': '001', 'implant_type': 'tetrode', 'implant_geometry': 'square', 'wire_length': 25, 'wire_length_units': 'um', 'implant_units': 'uV'}

    session_settings = {'channel_count': 4, 'animal': animal, 'devices': devices, 'implant': implant}

    """ FOR YOU TO EDIT """
    settings = {'ppm': 485, 'session':  session_settings, 'smoothing_factor': 3, 'useMatchedCut': False}
    """ FOR YOU TO EDIT """

    tasks['disk_arena'] = True # -->
    settings['tasks'] = tasks # --> change tasks array to change tasks are run
    settings['plotTasks'] = plotTasks # --> change plot tasks array to change asks taht are plotted
    settings['header'] = csv_header # --> change csv_header header to change tasks that are saved to csv

    """ FOR YOU TO EDIT """
    settings['arena_size'] = None
    settings['speed_lowerbound'] = 0
    settings['speed_upperbound'] = 99
    settings['end_cell'] = None
    settings['start_cell'] = None
    # possible saves are:
    # 1 csv per session (all tetrode and indiv): 'one_per_session' --> 5 sheets (summary of all 4 tetrodes, tet1, tet2, tet3, tet4)
    # 1 csv per animal per tetrode (all sessions): 'one_per_animal_tetrode' --> 4 sheets one per tet 
    # 1 csv per animal (all tetrodes & sessions): 'one_for_parent' --> 1 sheet
    """ FOR YOU TO EDIT """


    start_time = time.time()
    root = tk.Tk()
    root.withdraw()
    data_dir = filedialog.askdirectory(parent=root,title='Please select a data directory.')

    ########################################################################################################################

    """ OPTION 1 """
    """ RUNS EVERYTHING UNDER PARENT FOLDER (all subfolders loaded first) """
    # study = make_study(data_dir,settings_dict=settings)
    # study.make_animals()
    # batch_map(study, settings, data_dir)

    """ OPTION 2 """
    """ RUNS EACH SUBFOLDER ONE AT A TIME """
    subdirs = np.sort([ f.path for f in os.scandir(data_dir) if f.is_dir() ])
    count = 1
    for subdir in subdirs:
        try:
            study = make_study(subdir,settings_dict=settings)
            study.make_animals()
            batch_map(study, settings, subdir, sum_sheet_count=count)
            count += 1
        except Exception:
            print(traceback.format_exc())
            print('DID NOT WORK FOR DIRECTORY ' + str(subdir))



import os
import sys
import traceback

PROJECT_PATH = os.getcwd()
sys.path.append(PROJECT_PATH)

from _prototypes.unit_matcher.waveform import time_index, troughs
import warnings
from library.study_space import Session, Study, Animal
from scripts.batch_map.LEC_naming import LEC_naming_format, extract_name_lec
from _prototypes.cell_remapping.src.MEC_naming import MEC_naming_format, extract_name_mec
from _prototypes.cell_remapping.src.LC_naming import LC_naming_format, extract_name_lc
from _prototypes.cell_remapping.src.Hande_naming import Hande_naming_format, extract_name_hande


import tkinter as tk
from tkinter import filedialog
import time
from library.spatial import place_field
from library.hafting_spatial_maps import SpatialSpikeTrain2D
from library.scores import rate_map_stats, rate_map_coherence, border_score, grid_score
from library.maps.autocorrelation import autocorrelation
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from _prototypes.cell_remapping.src.utils import check_cylinder
from _prototypes.cell_remapping.src.masks import flat_disk_mask
from library.maps.map_blobs import map_blobs
from library.maps.map_utils import disk_mask
from PIL import Image
import numpy as np
from matplotlib import cm
import matplotlib.pyplot as plt
#import random
import openpyxl as xl
from openpyxl.worksheet.dimensions import ColumnDimension
from matplotlib import cm
# from opexebo.analysis import rate_map_stats, speed_score, rate_map_coherence
# Set matplotlib backend to Agg to prevent figures from popping up.
from matplotlib import use
from library.maps.firing_rate_vs_time import firing_rate_vs_time
from library.maps.filter_pos_by_speed import filter_pos_by_speed
from library.map_utils import _speed2D
from x_io.rw.axona.batch_read import make_study
import pandas as pd
import re
import random
from scipy.ndimage import gaussian_filter
from scipy.spatial import distance


""" SETTINGS AT THE BOTTOM OF FILE """

def _check_single_format(filename, format, fxn):
    print(filename, format, fxn)
    if re.match(str(format), str(filename)) is not None:
        return fxn(filename)


def batch_map(study: Study, settings_dict: dict, saveDir=None, sum_sheet_count=None):
    """
    Computes rate maps across all animals, sessions, cells in a study.

    Use tasks dictionary as true/false flag with variable to compute
    e.g. {'rate_map': True, 'binary_map': False}
    """

    tasks = settings_dict['tasks']
    plotTasks = settings_dict['plotTasks']
    csv_header = settings_dict['header']

    # Root path creation
    save_dir = saveDir
    root_path = os.path.join(save_dir, 'Spatial_Shuffle_')
    run_number = 1
    while os.path.isdir(root_path+str(run_number)):
        run_number+=1

    root_path += str(run_number)
    os.mkdir(root_path)

    # Kernel size
    kernlen = int(settings_dict['smoothing_factor']*8)
    # Standard deviation size
    std = int(0.2*kernlen)

    # Set flag if no arguments were provided for later
    all_cells = False
    if (settings_dict['start_cell'] == settings_dict['end_cell'] == None):
        all_cells = True

    # Grabs headers whose value is true
    headers = [k for k, v in csv_header.items() if v]
    if 'border_score' in headers:
        idx = headers.index('border_score') 
        # print(idx)
        headers.remove('border_score')
        headers.insert(idx, 'border_score_top')
        headers.insert(idx+1, 'border_score_bottom')
        headers.insert(idx+2, 'border_score_left')
        headers.insert(idx+3, 'border_score_right')
        # print(headers)
        # stop()

    headers_dict = dict()

    for i, header in enumerate(headers):
        headers_dict[header] = get_column_letter(i+4)

    file_name = os.path.join(root_path, "spatial_shuffle_parameters.txt")
    with open(file_name, 'w') as f:
        for ky in settings_dict:
            f.write(str(ky) + ' is: ')
            f.write(str(settings_dict[ky]) + '\n')
        f.close()

    if study.animals is None:
        study.make_animals()
        print('Animals made, batching map')

    per_animal_tracker = 1
    per_animal_tetrode_tracker = 1

    animal_tet_count = {}
    animal_max_tet_count = {}
    animal_workbooks = {}
    sorted_animal_ids = np.unique(np.sort(study.animal_ids))

    one_for_parent_wb = xl.Workbook()
    sum_sheet = one_for_parent_wb['Sheet']
    sum_sheet.title = 'Summary'
    sum_sheet['A' + str(1)] = 'Session'
    sum_sheet['B' + str(1)] = 'Tetrode'
    sum_sheet['C' + str(1)] = 'Cell ID'

    # grid_control_npy = []  
    # border_control_npy = [] 
    # information_control_npy = []

    # grid_path = r"/home/apollo/Desktop/AllGusData/GusRemapDataSes1Control/grid_control_dist.npy"
    # if os.path.exists(grid_path):
    #     grid_control_npy = np.load(grid_path)
    #     assert len(grid_control_npy) > 0
    # else:
    #     grid_control_npy = []

    # border_path = r"/home/apollo/Desktop/AllGusData/GusRemapDataSes1Control/border_control_dist.npy"
    # if os.path.exists(border_path):
    #     border_control_npy = np.load(border_path)
    #     assert len(border_control_npy) > 0
    # else:
    #     border_control_npy = []

    information_path = r"/home/apollo/Desktop/AllGusData/GusRemapDataSes1Control/information_control_dist.npy"
    # information_path_labels = r"/home/apollo/Desktop/AllGusData/GusRemapDataSes1Control/information_control_dist_labels.npy"
    if os.path.exists(information_path):
        information_control_npy = np.load(information_path)
        assert len(information_control_npy) > 0
        # information_control_npy_labels = np.load(information_path_labels)
        # assert len(information_control_npy_labels) > 0
        # assert len(information_control_npy_labels) == len(information_control_npy)
    else:
        information_control_npy = []
        # information_control_npy_labels = []


    for animalID in sorted_animal_ids:
        animal = study.get_animal_by_id(animalID)

        animal_id = animal.animal_id.split('_tet')[0]
        if animal_id not in animal_tet_count:
            animal_tet_count[animal_id] = 1
            # animal_sessions_tets_events[animal_id] = {}
            wb = xl.Workbook()
            animal_workbooks[animal_id] = wb
            animal_max_tet_count[animal_id] = len(sorted_animal_ids[list(map(lambda x: True if animal_id in x else False, sorted_animal_ids))])
            sum_sheet = wb['Sheet']
            sum_sheet.title = 'Summary'
            if settings_dict['saveMethod'] == 'one_per_session':
                # animal_tet_count[animal_id] = np.ones(len(animal.sessions))
                animal_workbooks[animal_id] = {}
                for ses_key in animal.sessions:
                    wb = xl.Workbook()
                    sum_sheet = wb['Sheet']
                    sum_sheet.title = 'Summary'
                    animal_workbooks[animal_id][ses_key] = wb
            elif settings_dict['saveMethod'] == 'one_per_animal_tetrode':
                # wb.remove_sheet('Summary') 
                pass
            elif settings_dict['saveMethod'] == 'one_for_parent':
                # Copy headers into excel file
                #current_statistics_sheet.range('A' + str(1)).value = 'Cell'
                wb = one_for_parent_wb
                sum_sheet = wb['Summary']
                # sum_sheet['A' + str(1)] = 'Session'
                # sum_sheet['B' + str(1)] = 'Tetrode'
                # sum_sheet['C' + str(1)] = 'Cell'
                for header in headers:
                    #current_statistics_sheet.range(get_column_letter(i+2) + str(1)).value = value
                    sum_sheet[headers_dict[header] + str(1)] = header
        else:
            if settings_dict['saveMethod'] != 'one_for_parent':
                animal_tet_count[animal_id] += 1
                # animal_tets[animal_id].append(animal)
                wb = animal_workbooks[animal_id]
            else:
                wb = one_for_parent_wb

        k = 1

        # col_count = 0 

        for session_key in animal.sessions:
            session = animal.sessions[session_key]

            c = 0

            pos_obj = session.get_position_data()['position']

            # excel_cell_index = 1
            per_session_tracker = 1

            if settings_dict['saveData'] == True:
                tet_file = session.session_metadata.file_paths['tet']

                # Session stamp
                signature = tet_file.split("/")[-1][:-2]

                if settings['naming_type'] == 'LEC':
                    group, name = extract_name_lec(signature)
                    formats = LEC_naming_format[group][name]['object']
                elif settings['naming_type'] == 'MEC':
                        name = extract_name_mec(signature)
                        formats = MEC_naming_format
                elif settings['naming_type'] == 'LC':
                    name = extract_name_lc(fname)
                    formats = LC_naming_format
                elif settings['naming_type'] == 'Hande':
                    name = extract_name_hande(fname)
                    formats = Hande_naming_format

                for format in list(formats.keys()):
                    checked = _check_single_format(signature, format, formats[format])
                    if checked is not None:
                        break
                    else:
                        continue
                
                stim, depth, name, date = checked

                # Create save_folder
                save_path = os.path.join(root_path, 'session_' + signature) 

                # if not os.path.isdir(save_path):
                #     os.mkdir(save_path)

                pt = os.path.join(root_path,'shuffled_dist_plots')
                if not os.path.isdir(pt):
                    os.mkdir(pt)

                directory = 'Tetrode_' + tet_file[-1]

                # # Building save directories for tetrodes
                # tetrode_directory_paths = dict()
                # # for directory in tetrode_directories:
                # path = os.path.join(save_path, directory)

                # if not os.path.isdir(path):
                #     os.mkdir(path)
                # tetrode_directory_paths[directory] = path

                # # CREATING EXCEL SHEET FOR STATISTIC SCORES
                # wb = animal_workbooks[animal_id]

                if settings_dict['saveMethod'] == 'one_per_session':
                    wb = animal_workbooks[animal_id][session_key]
                    # sum_sheet = wb['Sheet']
                    # sum_sheet.title = 'Summary'
                    current_statistics_sheet = wb.create_sheet(title=directory)
                    # Copy headers into excel file
                    #current_statistics_sheet.range('A' + str(1)).value = 'Cell'
                    current_statistics_sheet['A' + str(1)] = 'Session'
                    current_statistics_sheet['B' + str(1)] = 'Tetrode'
                    current_statistics_sheet['C' + str(1)] = 'Cell ID'
                    for header in headers:
                        #current_statistics_sheet.range(get_column_letter(i+2) + str(1)).value = value
                        current_statistics_sheet[headers_dict[header] + str(1)] = header
                elif settings_dict['saveMethod'] == 'one_per_animal_tetrode':
                    if directory not in wb.sheetnames:
                        current_statistics_sheet = wb.create_sheet(title=directory)
                        # Copy headers into excel file
                        #current_statistics_sheet.range('A' + str(1)).value = 'Cell'
                        current_statistics_sheet['A' + str(1)] = 'Session'
                        current_statistics_sheet['B' + str(1)] = 'Tetrode'
                        current_statistics_sheet['C' + str(1)] = 'Cell ID'
                        for header in headers:
                            #current_statistics_sheet.range(get_column_letter(i+2) + str(1)).value = value
                            current_statistics_sheet[headers_dict[header] + str(1)] = header
                    else:
                        current_statistics_sheet = wb[str(directory)]
                elif settings_dict['saveMethod'] == 'one_for_parent':
                    current_statistics_sheet = wb['Summary']

            for cell in session.get_cell_data()['cell_ensemble'].cells:

                if settings_dict['saveMethod'] == 'one_for_parent':
                    excel_cell_index = per_animal_tracker
                elif settings_dict['saveMethod'] == 'one_per_animal_tetrode':
                    excel_cell_index = per_animal_tetrode_tracker
                elif settings_dict['saveMethod'] == 'one_per_session':
                    excel_cell_index = per_session_tracker

                current_statistics_sheet['C' + str(excel_cell_index+1)] = cell.cluster.cluster_label
                current_statistics_sheet['B' + str(excel_cell_index+1)] = tet_file[-1]
                # animal_tet_count[animal_id]
                current_statistics_sheet['A' + str(excel_cell_index+1)] = signature

                if all_cells == True or ((all_cells==False) & (cell.cluster.cluster_label >= settings_dict['start_cell'] & cell.cluster.cluster_label <= settings_dict['end_cell'])):
                    print('animal: ' + str(animalID) + ', session: ' + str(k) + ', cell_cluster_label: ' + str(cell.cluster.cluster_label))

                    spatial_spike_train = session.make_class(SpatialSpikeTrain2D, {'cell': cell, 'position': pos_obj, 'speed_bounds': (settings_dict['speed_lowerbound'], settings_dict['speed_upperbound'])})

                    pos_x, pos_y, pos_t, arena_size = spatial_spike_train.x, spatial_spike_train.y, spatial_spike_train.t, spatial_spike_train.arena_size
                    v = _speed2D(spatial_spike_train.x, spatial_spike_train.y, spatial_spike_train.t)

                    rate_obj = spatial_spike_train.get_map('rate')
                    rate_map, rate_map_raw = rate_obj.get_rate_map(new_size=settings_dict['ratemap_dims'][0])
                    occ_obj = spatial_spike_train.get_map('occupancy')
                    occ_map, _, _ = occ_obj.get_occupancy_map(new_size=settings_dict['ratemap_dims'][0])

                    
                    path = session.session_metadata.file_paths['tet']
                    fname = path.split('/')[-1].split('.')[0]
                    cylinder = check_cylinder(fname, settings['disk_arena'])
                    if cylinder:
                        rate_map = flat_disk_mask(rate_map)
                        occ_map = flat_disk_mask(occ_map)
                        rate_map_raw = flat_disk_mask(rate_map_raw)

                    if not cylinder:
                        b_score = border_score(spatial_spike_train, smoothing_factor=settings_dict['smoothing_factor'])
                    else:
                        print('cant do border score on circular arena')
                        b_score = [np.nan, np.nan, np.nan, np.nan]

                    b_score_top = b_score[0]
                    b_score_bottom = b_score[1]
                    b_score_left = b_score[2]
                    b_score_right = b_score[3]

                    gr_score = grid_score(spatial_spike_train)

                    current_statistics_sheet[headers_dict['border_score_top'] + str(excel_cell_index+1)] = b_score_top
                    current_statistics_sheet[headers_dict['border_score_bottom'] + str(excel_cell_index+1)] = b_score_bottom
                    current_statistics_sheet[headers_dict['border_score_left'] + str(excel_cell_index+1)] = b_score_left
                    current_statistics_sheet[headers_dict['border_score_right'] + str(excel_cell_index+1)] = b_score_right
                    current_statistics_sheet[headers_dict['grid_score'] + str(excel_cell_index+1)] = gr_score

                    # if settings['normalizeRate']:
                    #     rate_map, _ = rate_obj.get_rate_map()
                    # else:
                    #     _, rate_map = rate_obj.get_rate_map()

                    # spikex, spikey, spiket = spatial_spike_train.get_spike_positions()

                    # print('Map Stats')
                    ratemap_stats_dict  = rate_map_stats(spatial_spike_train)
                    coherence = rate_map_coherence(spatial_spike_train)
                    spatial_information_content = ratemap_stats_dict['spatial_information_content']
                    current_statistics_sheet[headers_dict['information'] + str(excel_cell_index+1)] = spatial_information_content
                    current_statistics_sheet[headers_dict['sparsity'] + str(excel_cell_index+1)] = ratemap_stats_dict['sparsity']
                    current_statistics_sheet[headers_dict['selectivity'] + str(excel_cell_index+1)] = ratemap_stats_dict['selectivity']
                    current_statistics_sheet[headers_dict['coherence'] + str(excel_cell_index+1)] = coherence
                    

                    current_statistics_sheet[headers_dict['date'] + str(excel_cell_index+1)] = date 
                    current_statistics_sheet[headers_dict['name'] + str(excel_cell_index+1)] = name
                    current_statistics_sheet[headers_dict['depth'] + str(excel_cell_index+1)] = depth  
                    current_statistics_sheet[headers_dict['stim'] + str(excel_cell_index+1)] = stim

                    shuffled_rate_maps, shuffled_rate_map_raw = rate_obj.get_rate_map(new_size=rate_map.shape[0], shuffle=True, n_repeats=1000)

                    shuffled_information_content = []
                    shuffled_sparsity = []
                    shuffled_selectivity = []
                    shuffled_coherence = []
                    shuffled_border_score_top = []
                    shuffled_border_score_bottom = []
                    shuffled_border_score_left = []
                    shuffled_border_score_right = []
                    shuffled_grid_score = []
                    shuffled_border_score_agg = []
                    gr_score_nan_count = 0
                    info_score_high_count = 0
                    for i in range(1000):
                        shuffled_map = shuffled_rate_maps[i]
                        shuffled_map_raw = shuffled_rate_map_raw[i]
                        if cylinder:
                            shuffled_map = flat_disk_mask(shuffled_map)
                            shuffled_map_raw = flat_disk_mask(shuffled_map_raw)
                        # shuffled_map, shuffled_map_raw, fields_map = spatial_shuffle_with_incremental_field_placement(rate_map)
                        
                        # if i % 100 == 0:
                        #     fig = plt.figure(figsize=(4,4))
                        #     ax = plt.subplot(1,2,1)
                        #     ax.imshow(rate_map)
                        #     ax.set_title('True map')
                        #     ax = plt.subplot(1,2,2)
                        #     ax.imshow(shuffled_map)
                        #     ax.set_title('Shuffled map')
                        #     fig.suptitle(date+'_'+name+'_'+depth+'_'+tet_file[-1]+'_'+cell.cluster.cluster_label)
                        #     fig.tight_layout()
                        #     plt.show()

                        ratemap_stats_dict  = rate_map_stats(None, ratemap=shuffled_map, occmap=occ_map, override=True)
                        spatial_information_content = ratemap_stats_dict['spatial_information_content']
                        shuffled_information_content.append(spatial_information_content)
                        shuffled_sparsity.append(ratemap_stats_dict['sparsity'])
                        shuffled_selectivity.append(ratemap_stats_dict['selectivity'])
                        shuffled_coherence.append(rate_map_coherence(shuffled_map_raw, smoothing_factor=settings['smoothing_factor']))

                        if spatial_information_content >= 2:
                            info_score_high_count += 1


                        # shuffled_binmap = np.zeros(shuffled_map.shape)
                        # shuffled_binmap[  shuffled_map >= np.percentile(shuffled_map.flatten(), 75)  ] = 1
                        
                        image, n_labels, labels, centroids, field_sizes = map_blobs(shuffled_map, ratemap_size=32, cylinder=cylinder, 
                                                                                downsample=False, downsample_factor=None, smoothing_factor=settings_dict['smoothing_factor'])

                        if len(labels[labels > 1]) > 0:
                            shuffled_binmap = np.copy(labels)
                            shuffled_binmap[shuffled_binmap > 1] = 1
                        else:
                            shuffled_binmap = np.copy(shuffled_map_raw)
                            shuffled_binmap[  shuffled_map_raw >= np.percentile(shuffled_map_raw.flatten(), 60)  ] = 1
                            shuffled_binmap[  shuffled_map_raw < np.percentile(shuffled_map_raw.flatten(), 60)  ] = 0 

                        # _, shuffled_binmap = place_field(shuffled_map)
                        # shuffled_binmap[shuffled_binmap >= 1] = 1
                        # print(len(np.unique(shuffled_binmap)))
                        # assert len(np.unique(shuffled_binmap)) == 2

                        # fig = plt.figure(figsize=(12,5))
                        # ax = plt.subplot(1,3,1)
                        # ax.imshow(rate_map) 
                        # ax=plt.subplot(1,3,2)
                        # ax.imshow(shuffled_map)
                        # ax = plt.subplot(1,3,3)
                        # ax.imshow(shuffled_binmap)
                        # fig.tight_layout()
                        # plt.show()


                        b_score = border_score(None, smoothing_factor=settings_dict['smoothing_factor'], 
                        use_objects_directly=True, rate_map = shuffled_map_raw, bin_map = shuffled_binmap)

                        b_score_top = b_score[0]
                        b_score_bottom = b_score[1]
                        b_score_left = b_score[2]
                        b_score_right = b_score[3]
                        shuffled_border_score_top.append(b_score_top)
                        shuffled_border_score_bottom.append(b_score_bottom)
                        shuffled_border_score_left.append(b_score_left)
                        shuffled_border_score_right.append(b_score_right)
                        shuffled_border_score_agg.append(b_score)

                        if settings_dict['arena_size'] is not None:
                            assert spatial_spike_train.arena_size == settings_dict['arena_size']

                        shuffled_autocorr = autocorrelation(shuffled_map, use_map_directly=True, smoothing_factor=settings_dict['smoothing_factor'], arena_size=spatial_spike_train.arena_size)

                        gr_score = grid_score(None, use_autocorr_direclty=True, autocorr=shuffled_autocorr)
                        if gr_score == gr_score:
                            shuffled_grid_score.append(gr_score)
                        else:
                            gr_score_nan_count += 1


                    current_statistics_sheet[headers_dict['shuffled_information_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_information_content)
                    current_statistics_sheet[headers_dict['shuffled_information_std'] + str(excel_cell_index+1)] = np.std(shuffled_information_content)
                    current_statistics_sheet[headers_dict['shuffled_sparsity_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_sparsity)
                    current_statistics_sheet[headers_dict['shuffled_sparsity_std'] + str(excel_cell_index+1)] = np.std(shuffled_sparsity)
                    current_statistics_sheet[headers_dict['shuffled_selectivity_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_selectivity)
                    current_statistics_sheet[headers_dict['shuffled_selectivity_std'] + str(excel_cell_index+1)] = np.std(shuffled_selectivity)
                    current_statistics_sheet[headers_dict['shuffled_coherence_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_coherence)
                    current_statistics_sheet[headers_dict['shuffled_coherence_std'] + str(excel_cell_index+1)] = np.std(shuffled_coherence)
                    current_statistics_sheet[headers_dict['shuffled_border_score_top_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_border_score_top)
                    current_statistics_sheet[headers_dict['shuffled_border_score_top_std'] + str(excel_cell_index+1)] = np.std(shuffled_border_score_top)
                    current_statistics_sheet[headers_dict['shuffled_border_score_bottom_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_border_score_bottom)
                    current_statistics_sheet[headers_dict['shuffled_border_score_bottom_std'] + str(excel_cell_index+1)] = np.std(shuffled_border_score_bottom)
                    current_statistics_sheet[headers_dict['shuffled_border_score_left_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_border_score_left)
                    current_statistics_sheet[headers_dict['shuffled_border_score_left_std'] + str(excel_cell_index+1)] = np.std(shuffled_border_score_left)
                    current_statistics_sheet[headers_dict['shuffled_border_score_right_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_border_score_right)
                    current_statistics_sheet[headers_dict['shuffled_border_score_right_std'] + str(excel_cell_index+1)] = np.std(shuffled_border_score_right)
                    current_statistics_sheet[headers_dict['shuffled_grid_score_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_grid_score)
                    current_statistics_sheet[headers_dict['shuffled_grid_score_std'] + str(excel_cell_index+1)] = np.std(shuffled_grid_score)

                    p_value_information = (np.sum(shuffled_information_content < spatial_information_content)) / (len(shuffled_information_content))
                    p_value_sparsity = (np.sum(shuffled_sparsity > ratemap_stats_dict['sparsity'])) / (len(shuffled_sparsity))
                    p_value_selectivity = (np.sum(shuffled_selectivity < ratemap_stats_dict['selectivity'])) / (len(shuffled_selectivity))
                    p_value_coherence = (np.sum(shuffled_coherence < coherence)) / (len(shuffled_coherence))

                    p_value_border_score_top = (np.sum(shuffled_border_score_top > b_score_top)) / (len(shuffled_border_score_top))
                    p_value_border_score_bottom = (np.sum(shuffled_border_score_bottom > b_score_bottom)) / (len(shuffled_border_score_bottom))
                    p_value_border_score_left = (np.sum(shuffled_border_score_left > b_score_left)) / (len(shuffled_border_score_left))
                    p_value_border_score_right = (np.sum(shuffled_border_score_right > b_score_right)) / (len(shuffled_border_score_right))
                    p_value_grid_score = (np.sum(shuffled_grid_score > gr_score)) / (len(shuffled_grid_score))


                    dists = [shuffled_information_content, shuffled_sparsity, shuffled_selectivity, shuffled_coherence, shuffled_grid_score, shuffled_border_score_top, shuffled_border_score_bottom, shuffled_border_score_left, shuffled_border_score_right]
                    quants = [spatial_information_content, ratemap_stats_dict['sparsity'], ratemap_stats_dict['selectivity'], coherence, gr_score, b_score_top, b_score_bottom, b_score_left, b_score_right]
                    order = ['Information','Sparsity','Selectivity','Coherence', 'Grid Score', 'Border Score Top', 'Border Score Bottom', 'Border Score Left', 'Border Score Right']
                    file_end = str(animal_id) + '_' + str(date) + '_' + str(session_key) + str(tet_file[-1]) + str(cell.cluster.cluster_label) + '.png'

                    plot_shuffled_dist(root_path, dists, quants, order, file_end)

                    # grid_control_npy = np.hstack((grid_control_npy, shuffled_grid_score))
                    # print('SHAPE HERE')
                    # print(grid_control_npy.shape)
                    # if len(border_control_npy) == 0:
                    #     border_control_npy = np.array(shuffled_border_score_agg).reshape((len(shuffled_grid_score),4))
                    # else:
                    #     toadd = np.array(shuffled_border_score_agg).reshape((len(shuffled_grid_score),4))
                    #     border_control_npy = np.vstack((border_control_npy, toadd))
                    # print('SHAPE2 HERE')
                    # print(border_control_npy.shape)

                    curr_labels = np.array([file_end for x in range(len(shuffled_information_content))])
                    information_control_npy = np.hstack((information_control_npy, shuffled_information_content))
                    # information_control_npy_labels = np.hstack((information_control_npy_labels, curr_labels))

                    current_statistics_sheet[headers_dict['p_value_information'] + str(excel_cell_index+1)] = p_value_information
                    current_statistics_sheet[headers_dict['p_value_sparsity'] + str(excel_cell_index+1)] = p_value_sparsity
                    current_statistics_sheet[headers_dict['p_value_selectivity'] + str(excel_cell_index+1)] = p_value_selectivity
                    current_statistics_sheet[headers_dict['p_value_coherence'] + str(excel_cell_index+1)] = p_value_coherence
                    current_statistics_sheet[headers_dict['p_value_grid_score'] + str(excel_cell_index+1)] = p_value_grid_score
                    current_statistics_sheet[headers_dict['p_value_border_score_top'] + str(excel_cell_index+1)] = p_value_border_score_top
                    current_statistics_sheet[headers_dict['p_value_border_score_bottom'] + str(excel_cell_index+1)] = p_value_border_score_bottom
                    current_statistics_sheet[headers_dict['p_value_border_score_left'] + str(excel_cell_index+1)] = p_value_border_score_left
                    current_statistics_sheet[headers_dict['p_value_border_score_right'] + str(excel_cell_index+1)] = p_value_border_score_right

                    current_statistics_sheet[headers_dict['gr_score_nan_count'] + str(excel_cell_index+1)] = gr_score_nan_count
                    current_statistics_sheet[headers_dict['info_score_high_count'] + str(excel_cell_index+1)] = info_score_high_count

                    

                    sptimes = spatial_spike_train.spike_times
                    t_stop = max(sptimes)
                    t_start = min(sptimes)
                    offset_lim = 20/60
                    if offset_lim >= 0.5 * (t_stop - t_start):
                        offset_lim = int(0.5 * (t_stop - t_start))
                        if offset_lim == 0.5 * (t_stop - t_start):
                            offset_lim -= 1
                    
                    current_statistics_sheet[headers_dict['shuffled_offset'] + str(excel_cell_index+1)] = offset_lim

                    # # print('Check Disk')
                    # # if 'disk_arena' in tasks and tasks['disk_arena'] == False:
                    # fp = session.session_metadata.file_paths['cut']
                    # possible_names = ['round', 'cylinder', 'circle']
                    # isDisk = False
                    # for nm in possible_names:
                    #     if nm in fp.lower():
                    #         isDisk = True
                    #         # tasks['disk_arena'] = True



                        # Auto-resize columns to width of header text
                        #current_statistics_sheet.autofit(axis="columns")
                    ColumnDimension(current_statistics_sheet, bestFit=True)
                    print("Cell " + str(excel_cell_index) + " is complete")
                    per_session_tracker += 1
                    per_animal_tracker += 1
                    per_animal_tetrode_tracker += 1
                #     # -------------------------------------------------------- #

                c += 1

            k += 1

            if settings_dict['saveMethod'] == 'one_per_session':
                if animal_tet_count[animal_id] == animal_max_tet_count[animal_id]:
                    list(map(lambda x: _save_wb(animal_workbooks[animal_id][x], save_path), animal_workbooks[animal_id]))

        if settings_dict['saveMethod'] == 'one_per_animal_tetrode':
            if animal_tet_count[animal_id] == animal_max_tet_count[animal_id]:
                _save_wb(wb, root_path, animal_id=animal_id)

    if settings_dict['saveMethod'] == 'one_for_parent':
        _save_wb(wb, root_path, sum_sheet_count=sum_sheet_count)

    # np.save(grid_path, grid_control_npy)

    # np.save(border_path, border_control_npy)

    np.save(information_path, information_control_npy)
    # np.save(information_path_labels, information_control_npy_labels)


def plot_shuffled_dist(root_path, dists, quants, order, file_end):
    fig = plt.figure(figsize=(16,8))

    c = 0 
    for i in [1,3,5,7,9,11]:
        ax = plt.subplot(6,2,i)
        out = ax.hist(dists[c], bins=100, color='grey')
        ax.vlines(quants[c],0,np.max(out[0]), color='r')
        ax.set_xlabel(order[c])
        ax.set_ylabel('Count')
        ax.set_title('Non-log')
        

        ax_log = plt.subplot(6,2,i+1)
        try:
            out = ax_log.hist(np.log(dists[c]), bins=100, color='k')
            ax_log.vlines(np.log(quants[c]),0,np.max(out[0]), color='r')
        except:
            pass
        ax_log.set_xlabel(order[c])
        ax_log.set_ylabel('Count')
        ax_log.set_title('Post log')

        c += 1

    fig.tight_layout()

    pth = root_path + '/shuffled_dist_plots/' + str(file_end)

    fig.savefig(pth)

    plt.close()



def _save_wb(wb, root_path, animal_id=None, sum_sheet_count=None):
    wb._sheets = sorted(wb._sheets, key=lambda x: x.title)
    if animal_id is None:
        if sum_sheet_count is None:
            pth = root_path + '/shuffle_sheet'  + '.xlsx'
        else:
            pth = root_path + '/shuffle_sheet_'  + str(sum_sheet_count) + '.xlsx'
    else:
        pth = root_path + '/shuffle_sheet_' + str(animal_id)  + '.xlsx'
    print(root_path)
    wb.save(pth)
    wb.close()

    xls = pd.read_excel(pth, sheet_name=None)
    df_sum = xls.pop('Summary')
    dfs = [df.sort_values(['Session', 'Tetrode', 'Cell ID']) for df in xls.values()]
    with pd.ExcelWriter(pth, engine='xlsxwriter') as writer:
        df_sum.to_excel(writer, sheet_name='Summary', index=False)
        for sheet, df in zip(xls.keys(), dfs):
            df.to_excel(writer, sheet_name=sheet, index=False)


    
        if len(dfs) > 0:
            df_sum = pd.concat(dfs, axis=0).sort_values(['Session', 'Tetrode', 'Cell ID'])
        else:
            df_sum = df_sum.sort_values(['Session', 'Tetrode', 'Cell ID'])
        df_sum.to_excel(writer, sheet_name="Summary", index=True)
    # writer.save()
    print('Saved ' + str(pth))

def get_hd_score_for_cluster(hd_hist):
    angles = np.linspace(-179, 180, 360)
    angles_rad = angles*np.pi/180
    dy = np.sin(angles_rad)
    dx = np.cos(angles_rad)

    totx = sum(dx * hd_hist)/sum(hd_hist)
    toty = sum(dy * hd_hist)/sum(hd_hist)
    r = np.sqrt(totx*totx + toty*toty)
    return r




def spatial_shuffle_with_incremental_field_placement(ratemap):
    # Step 1: Detect fields using the place_field function
    s_image, n_s, fields_map, _, _ = map_blobs(ratemap, smoothing_factor=3)

    num_fields = fields_map.max()  # Number of fields detected

    # Start with an empty map for shuffled fields
    shuffled_ratemap = np.zeros_like(ratemap)
    occupied_positions = np.zeros_like(ratemap, dtype=bool)

    # Track the original positions of all field bins
    original_field_bins = [np.argwhere(fields_map == field_id) for field_id in range(1, num_fields + 1)]

    # Step 2: Randomize the order of fields
    field_order = list(range(num_fields))
    random.shuffle(field_order)

    # List to store non-field values from the new positions for filling holes
    values_for_filling_holes = []

    # Step 3: Incremental Placement of Bins
    for field_index in field_order:
        field_bins = original_field_bins[field_index]

        # Step 3a: Place the peak bin first
        peak_bin = field_bins[np.argmax(ratemap[tuple(field_bins.T)])]
        new_peak_position = (
            random.randint(0, ratemap.shape[0] - 1),
            random.randint(0, ratemap.shape[1] - 1)
        )

        # Collect the original value at the new peak position if it was not part of a field
        if fields_map[new_peak_position[0], new_peak_position[1]] == 0:
            values_for_filling_holes.append(ratemap[new_peak_position[0], new_peak_position[1]])

        # Place the peak bin in the new position
        shuffled_ratemap[new_peak_position[0], new_peak_position[1]] += ratemap[peak_bin[0], peak_bin[1]]
        occupied_positions[new_peak_position[0], new_peak_position[1]] = True

        # Step 3b: Place all other bins relative to the new peak position
        for bin_position in field_bins:
            if np.array_equal(bin_position, peak_bin):
                continue  # Skip the peak bin as it's already placed

            rel_pos = bin_position - peak_bin
            new_position = new_peak_position + rel_pos
            new_position = new_position % np.array(shuffled_ratemap.shape)

            # Collect the original value at the new position if it was not part of a field
            if fields_map[new_position[0], new_position[1]] == 0:
                values_for_filling_holes.append(ratemap[new_position[0], new_position[1]])

            # Place the field bin in the new position
            shuffled_ratemap[new_position[0], new_position[1]] += ratemap[bin_position[0], bin_position[1]]
            occupied_positions[new_position[0], new_position[1]] = True

    # Step 4: Fill in the holes left by the moved fields using the collected non-field values
    remaining_bins = np.argwhere(shuffled_ratemap == 0)
    if len(values_for_filling_holes) > len(remaining_bins):
        # extend the list of values to fill holes by repeating the values
        values_for_filling_holes = values_for_filling_holes * (len(values_for_filling_holes) // len(remaining_bins) + 1)

    for hole_bin in remaining_bins:
        if values_for_filling_holes:
            chosen_value = random.choice(values_for_filling_holes)
            shuffled_ratemap[tuple(hole_bin)] = chosen_value

    # Step 5: Fill in the untouched bins with non-field areas from the original map
    untouched_bins = np.logical_not(occupied_positions)
    non_field_bins = fields_map == 0
    fill_bins = untouched_bins & non_field_bins
    shuffled_ratemap[fill_bins] = ratemap[fill_bins]

    # gaussian smoothing
    shuffled_ratemap_raw = np.copy(shuffled_ratemap)
    shuffled_ratemap = shuffled_ratemap / np.sum(shuffled_ratemap)
    shuffled_ratemap = gaussian_filter(shuffled_ratemap, sigma=2)
    shuffled_ratemap = shuffled_ratemap / np.sum(shuffled_ratemap)

    return shuffled_ratemap, shuffled_ratemap_raw, fields_map


def spatial_shuffle_based_on_paper(ratemap):
    # Step 1: Segment fields using `map_blobs` (includes smoothing and segmentation)
    s_image, n_s, fields_map, _, _ = map_blobs(ratemap, smoothing_factor=3)

    num_fields = fields_map.max()  # Number of fields detected

    # Step 2: Initialize the shuffled ratemap and occupied positions map
    shuffled_ratemap = np.zeros_like(ratemap)  # Start with an empty map
    occupied_positions = np.zeros_like(ratemap, dtype=bool)  # Track occupied positions

    # Track the original positions of all field bins
    original_field_bins = [np.argwhere(fields_map == field_id) for field_id in range(1, num_fields + 1)]

    # Step 3: Identify peaks for each field and randomly place them
    field_peaks = []
    field_positions = []  # Track all positions of bins to shuffle them

    for field_id in range(1, num_fields + 1):
        # Identify bins belonging to the current field
        field_bins = np.argwhere(fields_map == field_id)
        field_positions.append(field_bins)

        # Find the peak (local maximum) within the field
        peak_bin = field_bins[np.argmax(ratemap[tuple(field_bins.T)])]
        field_peaks.append((field_id, peak_bin))

    # Randomly shuffle the order of peaks
    random.shuffle(field_peaks)

    # Randomly select positions for field peaks in the empty map
    new_peak_positions = []
    for field_id, peak_bin in field_peaks:
        found_position = False
        while not found_position:  # Loop until an unoccupied position is found
            new_position = (
                random.randint(0, ratemap.shape[0] - 1),
                random.randint(0, ratemap.shape[1] - 1)
            )
            if not occupied_positions[new_position]:
                new_peak_positions.append((field_id, new_position))
                occupied_positions[new_position] = True

                # Place the peak bin in the new position
                shuffled_ratemap[new_position] = ratemap[peak_bin[0], peak_bin[1]]
                found_position = True  # Exit the loop when a valid position is found

    # Step 4: Iteratively place bins in the shuffled ratemap while maintaining relative positions
    for (field_id, new_peak_position) in new_peak_positions:
        field_bins = field_positions[field_id - 1]
        peak_bin = field_bins[np.argmax(ratemap[tuple(field_bins.T)])]

        for bin_position in field_bins:
            if np.array_equal(bin_position, peak_bin):
                continue  # Skip the peak bin as it's already placed

            # Calculate the relative position to the peak
            rel_pos = bin_position - peak_bin
            ideal_position = new_peak_position + rel_pos

            # Find the nearest available position using city-block distance
            if not (0 <= ideal_position[0] < ratemap.shape[0] and 0 <= ideal_position[1] < ratemap.shape[1]) or occupied_positions[ideal_position[0], ideal_position[1]]:
                # Ideal position is not available, find the nearest available spot
                ideal_position = find_nearest_available_position(occupied_positions, ideal_position)

            # Place the bin in the available position
            shuffled_ratemap[ideal_position[0], ideal_position[1]] = ratemap[bin_position[0], bin_position[1]]
            occupied_positions[ideal_position[0], ideal_position[1]] = True

    # Step 5: Place non-field bins in their original positions if possible
    non_field_bins = np.argwhere(fields_map == 0)  # Bins not part of any field
    non_field_values_to_place = []  # Track non-field bins that can't be placed in their original location

    for bin_position in non_field_bins:
        if not occupied_positions[tuple(bin_position)]:
            # Place these bins in their original positions in the shuffled ratemap
            shuffled_ratemap[tuple(bin_position)] = ratemap[tuple(bin_position)]
            occupied_positions[tuple(bin_position)] = True
        else:
            # If the original position is occupied by a moved field bin, add it to the list
            non_field_values_to_place.append(ratemap[tuple(bin_position)])

    # Step 6: Randomly fill remaining empty positions with non-field values
    remaining_bins = np.argwhere(shuffled_ratemap == 0)
    for hole_bin in remaining_bins:
        if non_field_values_to_place:
            chosen_value = random.choice(non_field_values_to_place)
            shuffled_ratemap[tuple(hole_bin)] = chosen_value
            non_field_values_to_place.remove(chosen_value)  # Remove the placed value to avoid reuse

    # gaussian smoothing
    shuffled_ratemap_raw = np.copy(shuffled_ratemap)
    shuffled_ratemap = shuffled_ratemap / np.sum(shuffled_ratemap)
    shuffled_ratemap = gaussian_filter(shuffled_ratemap, sigma=2)
    shuffled_ratemap = shuffled_ratemap / np.sum(shuffled_ratemap)

    return shuffled_ratemap, shuffled_ratemap_raw, fields_map

def find_nearest_available_position(occupied_positions, ideal_position):
    """
    Finds the nearest available position in the shuffled map using city-block distance.
    """
    available_positions = np.argwhere(~occupied_positions)  # Positions that are not occupied
    if len(available_positions) == 0:  # All positions are filled, should never happen
        raise ValueError("No available positions to place field bin!")
    # Calculate city-block distances to all available positions
    distances = distance.cdist([ideal_position], available_positions, metric='cityblock')
    nearest_position_idx = np.argmin(distances)
    return tuple(available_positions[nearest_position_idx])


if __name__ == '__main__':

    ######################################################## EDIT BELOW HERE ########################################################

    # MAKE SURE "field_sizes" IS THE LAST ELEMENT IN "csv_header_keys"
    csv_header = {}
    csv_header_keys = ['name','date','depth','stim','information', 'shuffled_information_mean','shuffled_information_std',
                        'info_score_high_count', 'p_value_information',
                       'selectivity', 'shuffled_selectivity_mean', 'shuffled_selectivity_std', 'p_value_selectivity',
                       'sparsity', 'shuffled_sparsity_mean', 'shuffled_sparsity_std', 'p_value_sparsity',
                       'coherence', 'shuffled_coherence_mean', 'shuffled_coherence_std', 'p_value_coherence',
                       'grid_score', 'shuffled_grid_score_mean', 'shuffled_grid_score_std', 'gr_score_nan_count', 'p_value_grid_score',
                       'border_score_top', 'shuffled_border_score_top_mean', 'shuffled_border_score_top_std', 'p_value_border_score_top',
                        'border_score_bottom', 'shuffled_border_score_bottom_mean', 'shuffled_border_score_bottom_std', 'p_value_border_score_bottom',
                        'border_score_left', 'shuffled_border_score_left_mean', 'shuffled_border_score_left_std', 'p_value_border_score_left',
                        'border_score_right', 'shuffled_border_score_right_mean', 'shuffled_border_score_right_std', 'p_value_border_score_right',
                       'shuffled_offset']
    for key in csv_header_keys:
        csv_header[key] = True

    tasks = {}
    task_keys = ['information', 'sparsity', 'selectivity', 'coherence','grid_score', 'border_score']
    for key in task_keys:
        tasks[key] = True



    plotTasks = {}
    plot_task_keys = ['Spikes_Over_Position_Map', 'Tuning_Curve_Plots', 'Firing_Rate_vs_Speed_Plots', 'Firing_Rate_vs_Time_Plots','autocorr_map', 'binary_map','rate_map', 'occupancy_map']
    for key in plot_task_keys:
        plotTasks[key] = True

    animal = {'animal_id': '001', 'species': 'mouse', 'sex': 'F', 'age': 1, 'weight': 1, 'genotype': 'type', 'animal_notes': 'notes'}
    devices = {'axona_led_tracker': True, 'implant': True}
    implant = {'implant_id': '001', 'implant_type': 'tetrode', 'implant_geometry': 'square', 'wire_length': 25, 'wire_length_units': 'um', 'implant_units': 'uV'}

    session_settings = {'channel_count': 4, 'animal': animal, 'devices': devices, 'implant': implant}

    """ FOR YOU TO EDIT """
    settings = {'ppm': None, 'session':  session_settings, 'smoothing_factor': 2, 'useMatchedCut': False}
    """ FOR YOU TO EDIT """

    settings['disk_arena'] = False # -->
    settings['tasks'] = tasks # --> change tasks array to change tasks are run
    settings['plotTasks'] = plotTasks # --> change plot tasks array to change asks taht are plotted
    settings['header'] = csv_header # --> change csv_header header to change tasks that are saved to csv

    """ FOR YOU TO EDIT """
    settings['naming_type'] = 'MEC'
    settings['speed_lowerbound'] = 3
    settings['speed_upperbound'] = 100
    settings['arena_size'] = (50,50)
    settings['end_cell'] = None
    settings['start_cell'] = None
    settings['saveData'] = True
    settings['ratemap_dims'] = (32, 32)
    settings['saveMethod'] = 'one_for_parent'
    # possible saves are:
    # 1 csv per session (all tetrode and indiv): 'one_per_session' --> 5 sheets (summary of all 4 tetrodes, tet1, tet2, tet3, tet4)
    # 1 csv per animal per tetrode (all sessions): 'one_per_animal_tetrode' --> 4 sheets one per tet 
    # 1 csv per animal (all tetrodes & sessions): 'one_for_parent' --> 1 sheet
    """ FOR YOU TO EDIT """


    start_time = time.time()
    root = tk.Tk()
    root.withdraw()
    data_dir = filedialog.askdirectory(parent=root,title='Please select a data directory.')

    ########################################################################################################################

    """ OPTION 1 """
    """ RUNS EVERYTHING UNDER PARENT FOLDER (all subfolders loaded first) """
    # study = make_study(data_dir,settings_dict=settings)
    # study.make_animals()
    # batch_map(study, settings, data_dir)

    """ OPTION 2 """
    """ RUNS EACH SUBFOLDER ONE AT A TIME """
    subdirs = np.sort([ f.path for f in os.scandir(data_dir) if f.is_dir() ])
    count = 1
    for subdir in subdirs:
        try:
            study = make_study(subdir,settings_dict=settings)
            study.make_animals()
            batch_map(study, settings, subdir, sum_sheet_count=count)
            count += 1
        except Exception:
            print(traceback.format_exc())
            print('DID NOT WORK FOR DIRECTORY ' + str(subdir))

    # """ OPTION 3 """
    # """ RUNS EACH SUBFOLDER ONE AT A TIME """
    # subdirs = np.sort([ f.path for f in os.scandir(data_dir) if f.is_dir() ])
    # count = 1
    # for subdir in subdirs:
    #     subdir = str(subdir)
    #     subdirs2 = np.sort([ f.path for f in os.scandir(subdir) if f.is_dir() ])
    #     for subdir2 in subdirs2:
    #         try:
    #             study = make_study(subdir2,settings_dict=settings)
    #             study.make_animals()
    #             batch_map(study, settings, subdir2, sum_sheet_count=count)
    #             count += 1
    #         except Exception:
    #             print(traceback.format_exc())
    #             print('DID NOT WORK FOR DIRECTORY ' + str(subdir2))
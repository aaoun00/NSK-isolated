import os
import sys
import traceback

PROJECT_PATH = os.getcwd()
sys.path.append(PROJECT_PATH)

from _prototypes.unit_matcher.waveform import time_index, troughs
import warnings
from library.study_space import Session, Study, Animal
from library.workspace import Workspace
import tkinter as tk
from tkinter import filedialog
import time
from library.maps import autocorrelation, binary_map, spatial_tuning_curve, map_blobs
from library.hafting_spatial_maps import HaftingOccupancyMap, HaftingRateMap, HaftingSpikeMap, SpatialSpikeTrain2D
from library.scores import hd_score, grid_score, border_score
from library.scores import rate_map_stats, rate_map_coherence, speed_score
from openpyxl.utils.cell import get_column_letter, column_index_from_string
# from library.maps.map_utils import disk_mask
from _prototypes.cell_remapping.src.utils import check_cylinder
from _prototypes.cell_remapping.src.masks import flat_disk_mask
from PIL import Image
import numpy as np
from matplotlib import cm
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec

#import random
import openpyxl as xl
from openpyxl.worksheet.dimensions import ColumnDimension
from matplotlib import cm
# from opexebo.analysis import rate_map_stats, speed_score, rate_map_coherence
# Set matplotlib backend to Agg to prevent figures from popping up.
from matplotlib import use
from library.maps.firing_rate_vs_time import firing_rate_vs_time
from library.maps.filter_pos_by_speed import filter_pos_by_speed
from library.map_utils import _speed2D
from x_io.rw.axona.batch_read import make_study
import pandas as pd
from library.cluster import create_features, L_ratio, isolation_distance
from library.spike import histogram_ISI, find_burst


# class BatchMaps(Workspace):
#     def __init__(self, study: StudyWorkspace):
#         pass

# def add_to_sheet(to_add, headers_dict, header, wb, method, sheet, cell_index):
#     current_statistics_sheet[headers_dict['border_score_top'] + str(excel_cell_index+1)] = b_score[0]

#     if method == 'one_per_session':
#         sheet[headers_dict[header] + str(cell_index+1)] = to_add
#     elif method == 'one_per_animal_tetrode':
#         sheet[headers_dict[header] + str(cell_index+1)] = to_add

""" SETTINGS AT THE BOTTOM OF FILE """


def batch_map(study: Study, settings_dict: dict, saveDir=None, sum_sheet_count=None):
    """
    Computes rate maps across all animals, sessions, cells in a study.

    Use tasks dictionary as true/false flag with variable to compute
    e.g. {'rate_map': True, 'binary_map': False}
    """

    tasks = settings_dict['tasks']
    plotTasks = settings_dict['plotTasks']
    csv_header = settings_dict['header']

    # Root path creation
    save_dir = saveDir
    root_path = os.path.join(save_dir, 'All_Neurofunc_Sessions_iter_')
    run_number = 1
    while os.path.isdir(root_path+str(run_number)):
        run_number+=1

    root_path += str(run_number)
    os.mkdir(root_path)

    # Kernel size
    kernlen = int(settings_dict['smoothing_factor']*8)
    # Standard deviation size
    std = int(0.2*kernlen)

    # Set flag if no arguments were provided for later
    all_cells = False
    if (settings_dict['start_cell'] == settings_dict['end_cell'] == None):
        all_cells = True

    # Grabs headers whose value is true
    headers = [k for k, v in csv_header.items() if v]
    if 'border_score' in headers:
        idx = headers.index('border_score') 
        # print(idx)
        headers.remove('border_score')
        headers.insert(idx, 'border_score_top')
        headers.insert(idx+1, 'border_score_bottom')
        headers.insert(idx+2, 'border_score_left')
        headers.insert(idx+3, 'border_score_right')
        # print(headers)
        # stop()

    headers_dict = dict()

    for i, header in enumerate(headers):
        headers_dict[header] = get_column_letter(i+4)

    file_name = os.path.join(root_path, "neurofunc_parameters.txt")
    with open(file_name, 'w') as f:
        for ky in settings_dict:
            f.write(str(ky) + ' is: ')
            f.write(str(settings_dict[ky]) + '\n')
        # f.write("ppm is: ")
        # f.write(str(settings_dict['ppm']) + "\n")
        # f.write("Smoothing is: ")
        # f.write(str(settings_dict['smoothing_factor']) + "\n")
        # f.write("speed bounds are: ")
        # f.write(str(settings_dict['speed_lowerbound']) + "," + str(settings_dict['speed_upperbound']) + "\n")
        # f.write("save method is: ")
        # f.write(str(settings_dict['saveMethod']) + "\n")
        f.close()

    if study.animals is None:
        study.make_animals()
        print('Animals made, batching map')

    # animals = study.animals

    # analysis_directories = [k for k, v in plotTasks.items() if v]

    # animal_sessions_tets_events = {}
    # animal_sessions_tets_events_idx = {}
    # animal_tet_count = {}
    # animal_max_tet_count = {}
    # animal_sessions_tets_events_FD = {}

    per_animal_tracker = 1
    per_animal_tetrode_tracker = 1

    # for animalID in sorted_animal_ids:
    #     animal = study.get_animal_by_id(animalID)

    #     animal_id = animal.animal_id.split('_tet')[0]
    #     if animal_id not in animal_sessions_tets_events:
    #         animal_sessions_tets_events[animal_id] = {}
    #         animal_sessions_tets_events_idx[animal_id] = {}
    #         animal_sessions_tets_events_FD[animal_id] = {}
    #         animal_tet_count[animal_id] = {}
    #         animal_max_tet_count[animal_id] = len(sorted_animal_ids[list(map(lambda x: True if animal_id in x else False, sorted_animal_ids))])
        
        # for session_key in animal.sessions:
        #     session = animal.sessions[session_key]
        #     if settings_dict['saveData'] == True:
        #         clust = session.get_spike_data()['spike_cluster']
        #         good_label_ids = clust.good_label_ids
        #         cluster_labels = clust.cluster_labels

        #         def map_func(index, label):
        #             if label in good_label_ids:
        #                 return index
        #             else:
        #                 return None

        #         indices = list(filter(lambda x: x is not None, map(map_func, range(len(cluster_labels)), cluster_labels)))
        #         waveforms = [clust.waveforms[i] for i in indices]
        #         if session_key not in animal_sessions_tets_events[animal_id]:

        #             # event_times = clust.event_times
        #             animal_sessions_tets_events[animal_id][session_key] = [event_times]
        #             animal_sessions_tets_events_idx[animal_id][session_key] = [indices]
        #             animal_tet_count[animal_id] = 1
        #         else:
        #             animal_sessions_tets_events[animal_id][session_key].append(event_times)
        #             animal_sessions_tets_events_idx[animal_id][session_key].append(indices)
        #             animal_tet_count[animal_id] += 1
                
        #         if animal_tet_count[animal_id] == animal_max_tet_count[animal_id]:
        #             ch1, ch2, ch3, ch4 = animal_sessions_tets_events[animal_id][session_key]
        #             data_concat = np.vstack((ch1, ch2, ch3, ch4)).reshape((4, -1, ch1.shape[1]))
                        
        #             # cluster_batch = session.get_spike_data()['spike_cluster']
        #             FD = create_features(data_concat)
        #             animal_sessions_tets_events_FD[animal_id][session_key] = FD

    animal_tet_count = {}
    animal_max_tet_count = {}
    animal_workbooks = {}
    sorted_animal_ids = np.unique(np.sort(study.animal_ids))

    one_for_parent_wb = xl.Workbook()
    sum_sheet = one_for_parent_wb['Sheet']
    sum_sheet.title = 'Summary'
    sum_sheet['A' + str(1)] = 'Session'
    sum_sheet['B' + str(1)] = 'Tetrode'
    sum_sheet['C' + str(1)] = 'Cell ID'

    for animalID in sorted_animal_ids:
        animal = study.get_animal_by_id(animalID)

        animal_id = animal.animal_id.split('_tet')[0]
        tet_id = animal.animal_id.split('_tet')[1]
        if animal_id not in animal_tet_count:
            animal_tet_count[animal_id] = 1
            # animal_sessions_tets_events[animal_id] = {}
            wb = xl.Workbook()
            animal_workbooks[animal_id] = wb
            animal_max_tet_count[animal_id] = len(sorted_animal_ids[list(map(lambda x: True if animal_id in x else False, sorted_animal_ids))])
            sum_sheet = wb['Sheet']
            sum_sheet.title = 'Summary'
            if settings_dict['saveMethod'] == 'one_per_session':
                # animal_tet_count[animal_id] = np.ones(len(animal.sessions))
                animal_workbooks[animal_id] = {}
                for ses_key in animal.sessions:
                    wb = xl.Workbook()
                    sum_sheet = wb['Sheet']
                    sum_sheet.title = 'Summary'
                    animal_workbooks[animal_id][ses_key] = wb
            elif settings_dict['saveMethod'] == 'one_per_animal_tetrode':
                # wb.remove_sheet('Summary') 
                pass
            elif settings_dict['saveMethod'] == 'one_for_parent':
                # Copy headers into excel file
                #current_statistics_sheet.range('A' + str(1)).value = 'Cell'
                wb = one_for_parent_wb
                sum_sheet = wb['Summary']
                # sum_sheet['A' + str(1)] = 'Session'
                # sum_sheet['B' + str(1)] = 'Tetrode'
                # sum_sheet['C' + str(1)] = 'Cell'
                for header in headers:
                    #current_statistics_sheet.range(get_column_letter(i+2) + str(1)).value = value
                    sum_sheet[headers_dict[header] + str(1)] = header
        else:
            if settings_dict['saveMethod'] != 'one_for_parent':
                animal_tet_count[animal_id] += 1
                # animal_tets[animal_id].append(animal)
                wb = animal_workbooks[animal_id]
            else:
                wb = one_for_parent_wb

        # cells, waveforms = sort_cell_spike_times(animal)

        # sort_cell_spike_times(animal)

        # cluster = session.get_spike_data['spike_cluster']

        # sort_spikes_by_cell(cluster)

        k = 1

        # col_count = 0 

        for session_key in animal.sessions:
            session = animal.sessions[session_key]

            c = 0

            pos_obj = session.get_position_data()['position']

            # excel_cell_index = 1
            per_session_tracker = 1

            if settings_dict['saveData'] == True:
                tet_file = session.session_metadata.file_paths['tet']

                # Session stamp
                signature = tet_file.split("/")[-1][:-2]

                # Create save_folder
                save_path = os.path.join(root_path, 'neurofunc_session_' + signature) 
                # + '_' + str(k)

                if not os.path.isdir(save_path):
                    os.mkdir(save_path)
                # animal_tet_count[animal_id] += 1

                directory = 'Tetrode_' + tet_file[-1]
                        
                # Creating file directories for plots
                root_directory_paths = dict()
                root_directory_paths['Occupancy_Map'] = os.path.join(save_path, 'Occupancy_Map')

                # Creating occupancy map directory
                if not os.path.isdir(root_directory_paths['Occupancy_Map']):
                    os.mkdir(root_directory_paths['Occupancy_Map'])

                # Building save directories for tetrodes
                tetrode_directory_paths = dict()
                # for directory in tetrode_directories:
                path = os.path.join(save_path, directory)

                if not os.path.isdir(path):
                    os.mkdir(path)
                tetrode_directory_paths[directory] = path

                # # CREATING EXCEL SHEET FOR STATISTIC SCORES
                # wb = animal_workbooks[animal_id]

                if settings_dict['saveMethod'] == 'one_per_session':
                    wb = animal_workbooks[animal_id][session_key]
                    # sum_sheet = wb['Sheet']
                    # sum_sheet.title = 'Summary'
                    current_statistics_sheet = wb.create_sheet(title=directory)
                    # Copy headers into excel file
                    #current_statistics_sheet.range('A' + str(1)).value = 'Cell'
                    current_statistics_sheet['A' + str(1)] = 'Session'
                    current_statistics_sheet['B' + str(1)] = 'Tetrode'
                    current_statistics_sheet['C' + str(1)] = 'Cell ID'
                    for header in headers:
                        #current_statistics_sheet.range(get_column_letter(i+2) + str(1)).value = value
                        current_statistics_sheet[headers_dict[header] + str(1)] = header
                elif settings_dict['saveMethod'] == 'one_per_animal_tetrode':
                    if directory not in wb.sheetnames:
                        current_statistics_sheet = wb.create_sheet(title=directory)
                        # Copy headers into excel file
                        #current_statistics_sheet.range('A' + str(1)).value = 'Cell'
                        current_statistics_sheet['A' + str(1)] = 'Session'
                        current_statistics_sheet['B' + str(1)] = 'Tetrode'
                        current_statistics_sheet['C' + str(1)] = 'Cell ID'
                        for header in headers:
                            #current_statistics_sheet.range(get_column_letter(i+2) + str(1)).value = value
                            current_statistics_sheet[headers_dict[header] + str(1)] = header
                    else:
                        current_statistics_sheet = wb[str(directory)]
                elif settings_dict['saveMethod'] == 'one_for_parent':
                    current_statistics_sheet = wb['Summary']

            if tasks['spike_analysis']:
                clust = session.get_cell_data()['cell_ensemble']
                # print(np.array(clust.get_waveforms()[0]).shape)
                # print(np.array(clust.get_waveforms()[1]).shape)
                # print(np.asarray(clust.waveforms[0]).T.shape)
                # print(np.asarray(clust.waveforms[0]).shape)
                clust.get_waveforms()
                # print(tet_file)
                # print(np.array(clust.waveforms).shape, np.array(clust.waveform_ids).shape)
                # print(len(clust.waveforms), len(clust.waveforms[0]), len(clust.waveforms[0][0]))
                # print(len(clust.waveform_ids), len(clust.waveform_ids[0]), len(clust.waveform_ids[0][0]))
                for i in range(len(clust.waveforms)):
                    ch = np.asarray(clust.waveforms[i])
                    ch_lbl = np.asarray(clust.waveform_ids[i])
                    if i == 0:
                        data_concat = ch
                        data_concat_lbl = ch_lbl
                    else:
                        data_concat = np.vstack((data_concat, ch))
                        data_concat_lbl = np.hstack((data_concat_lbl, ch_lbl))
                data_concat = data_concat.reshape((4, -1, ch.shape[2]))
                # ch1, ch2, ch3, ch4 = np.asarray(clust.waveforms[0]), np.asarray(clust.waveforms[1]), np.asarray(clust.waveforms[2]), np.asarray(clust.waveforms[3])
                # ch1_lbl, ch2_lbl, ch3_lbl, ch4_lbl = np.asarray(clust.waveform_ids[0]), np.asarray(clust.waveform_ids[1]), np.asarray(clust.waveform_ids[2]), np.asarray(clust.waveform_ids[3])
                # data_concat = np.vstack((ch1, ch2, ch3, ch4)).reshape((4, -1, ch1.shape[2]))
                # data_concat_lbl = np.hstack((ch1_lbl, ch2_lbl, ch3_lbl, ch4_lbl))
                FD = create_features(data_concat)
                # print(ch1.shape, data_concat.shape, FD.shape, data_concat_lbl.shape)

                

            for cell in session.get_cell_data()['cell_ensemble'].cells:

                if settings_dict['saveMethod'] == 'one_for_parent':
                    excel_cell_index = per_animal_tracker
                elif settings_dict['saveMethod'] == 'one_per_animal_tetrode':
                    excel_cell_index = per_animal_tetrode_tracker
                elif settings_dict['saveMethod'] == 'one_per_session':
                    excel_cell_index = per_session_tracker

                current_statistics_sheet['C' + str(excel_cell_index+1)] = cell.cluster.cluster_label
                current_statistics_sheet['B' + str(excel_cell_index+1)] = tet_file[-1]
                # animal_tet_count[animal_id]
                current_statistics_sheet['A' + str(excel_cell_index+1)] = signature

                if all_cells == True or ((all_cells==False) & (cell.cluster.cluster_label >= settings_dict['start_cell'] & cell.cluster.cluster_label <= settings_dict['end_cell'])):
                    print('animal: ' + str(animalID) + ', session: ' + str(k) + ', cell_cluster_label: ' + str(cell.cluster.cluster_label))
                    # print(cell.event_times[:10])
                    # print('SpatialSPikeTrain Class')
                    spatial_spike_train = session.make_class(SpatialSpikeTrain2D, {'cell': cell, 'position': pos_obj, 'speed_bounds': (settings_dict['speed_lowerbound'], settings_dict['speed_upperbound'])})

                    pos_x, pos_y, pos_t, arena_size = spatial_spike_train.x, spatial_spike_train.y, spatial_spike_train.t, spatial_spike_train.arena_size
                    v = _speed2D(spatial_spike_train.x, spatial_spike_train.y, spatial_spike_train.t)
                    # stop()

                    # # print('HafftingOccupancyMap')
                    # occ_obj = HaftingOccupancyMap(spatial_spike_train)
                    # occ_map, _, _ = occ_obj.get_occupancy_map()

                    # # print('HaftingSpikeMap')
                    # spike_obj = HaftingSpikeMap(spatial_spike_train)
                    # spike_map = spike_obj.get_spike_map()

                    # print('HaftingRateMap')
                    # rate_obj = HaftingRateMap(spatial_spike_train)
                    rate_obj = spatial_spike_train.get_map('rate')
                    rate_map, rate_map_raw = rate_obj.get_rate_map()
                    # if settings['normalizeRate']:
                    #     rate_map, _ = rate_obj.get_rate_map()
                    # else:
                    #     _, rate_map = rate_obj.get_rate_map()

                    # spikex, spikey, spiket = spatial_spike_train.get_spike_positions()

                    cell_stats = {}

                    # print('Map Stats')

                    # UNDO COMMENT
                    # autocorr_map = autocorrelation(spatial_spike_train)

                    occ_obj = spatial_spike_train.get_map('occupancy')
                    occ_map, _, _ = occ_obj.get_occupancy_map()
                    spike_obj = spatial_spike_train.get_map('spike')
                    spike_map, _ = spike_obj.get_spike_map()

                    spikex, spikey, spiket = spike_obj.spike_x, spike_obj.spike_y, spike_obj.new_spike_times

                    if tasks['spike_analysis']:
                        idx = np.where(data_concat_lbl == cell.cluster.cluster_label)[0]
                        ClusterSpikes = idx
                        L, ratio_L, df = L_ratio(FD, ClusterSpikes)
                        iso_dist = isolation_distance(FD, ClusterSpikes)
                        bursting, avg_spikes_per_burst = find_burst(cell)
                        ISI_dict = histogram_ISI(cell)
                        spike_count = len(cell.event_times)
                    
                    if tasks['spike_width']:
                        # n_spikes, spike_times, waveforms = session.session_data.data['spike_cluster'].get_single_spike_cluster_instance(unit)
                        n_spikes = len(cell.event_times)
                        waveforms = cell.signal
                        wf_avg = np.array(waveforms).mean(axis=1)
                        max_vals = list(map(lambda x: max(x), [val for val in wf_avg]))
                        principal_channel_index = np.argmax(max_vals)
                        principal_waveform = wf_avg[principal_channel_index]
                        peak_index = np.argmax(principal_waveform)
                        trough_list = list(filter(lambda x: x > peak_index, troughs(principal_waveform)))
                        sample_rate = session.session_data.data['spike_cluster'].waveform_sample_rate
                        if len(trough_list) > 0:
                            trough_index = trough_list[0]
                            spike_width = (trough_index - peak_index) / sample_rate
                        else:
                            trough_index = len(principal_waveform) - 1
                            spike_width = int((trough_index - peak_index)/2) / sample_rate
                        if spike_width < 0:
                            warnings.warn(f'Negative spike width for unit {cell.cluster.cluster_label} in session {signature}.\n\nThe mean waveform is:\n{principal_waveform}\n\nThe peak index is {peak_index} and the trough index is {trough_index}. The spike width is {spike_width}.\n\nThe sample_rate is {sample_rate}.')
                        duration = session.session_data.data['spike_cluster'].duration
                        firing_rate = n_spikes/duration
                        cell_stats['firing_rate'] = firing_rate
                        cell_stats['spike_width'] = spike_width

                    # # print('Check Disk')
                    # # if 'disk_arena' in tasks and tasks['disk_arena'] == False:
                    # fp = session.session_metadata.file_paths['cut']
                    # possible_names = ['round', 'cylinder', 'circle']
                    # isDisk = False
                    # for name in possible_names:
                    #     if name in fp.lower():
                    #         # isDisk = True
                    #         tasks['disk_arena'] = True
                    
                    path = session.session_metadata.file_paths['tet']
                    fname = path.split('/')[-1].split('.')[0]
                    cylinder = check_cylinder(fname, settings['disk_arena'])
                    if cylinder:
                        tasks['disk_arena'] = True
                        rate_map = flat_disk_mask(rate_map)
                        occ_map = flat_disk_mask(occ_map)
                        rate_map_raw = flat_disk_mask(rate_map_raw)
                    else:
                        tasks['disk_arena'] = False

                    autocorr_map = autocorrelation(rate_map, use_map_directly=True, smoothing_factor=settings['smoothing_factor'], arena_size=spatial_spike_train.arena_size)
                    ratemap_stats_dict  = rate_map_stats(None, ratemap=rate_map, occmap=occ_map, override=True)


                    # print(fp, tasks['disk_arena'])

                    cell_stats['rate_map_smooth'] = rate_map
                    cell_stats['occupancy_map'] = occ_map
                    cell_stats['rate_map_raw'] = rate_map_raw
                    cell_stats['autocorrelation_map'] = autocorr_map
                    cell_stats['spatial_spike_train'] = spatial_spike_train

                    # print('Binary')
                    if tasks['binary_map']:
                        # binmap = binary_map(spatial_spike_train)
                        binmap = binary_map(rate_map, use_map_directly=True, smoothing_factor=settings['smoothing_factor'])
                        # if tasks['disk_arena']:
                        #     binmap = disk_mask(binmap)
                        # binmap_im = Image.fromarray(np.uint8(binmap*255))
                        cell_stats['binary_map'] = binmap
                        # cell_stats['binary_map_im'] = binmap_im

                    # print('Autocorr Img') 
                    if tasks['autocorrelation_map']:
                        cell_stats['autocorr_map'] = autocorr_map
                        # if tasks['disk_arena']:
                        #     autocorr_map = disk_mask(autocorr_map)
                        # autocorr_map_im = Image.fromarray(np.uint8(cm.jet(autocorr_map)*255))
                        # cell_stats['autocorr_map_im'] = autocorr_map_im

                    if tasks['sparsity']:
                        cell_stats['sparsity'] = ratemap_stats_dict['sparsity']

                    if tasks['selectivity']:
                        cell_stats['selectivity'] = ratemap_stats_dict['selectivity']

                    if tasks['information']:
                        cell_stats['spatial_information_content'] = ratemap_stats_dict['spatial_information_content']

                    # print('Coherence')
                    if tasks['coherence']:
                        coherence = rate_map_coherence(rate_map_raw, smoothing_factor=settings['smoothing_factor'])
                        cell_stats['coherence'] = coherence

                    # print('Speed Score')
                    if tasks['speed_score']:
                        # if tasks['disk_arena']:
                        s_score = speed_score(spatial_spike_train)
                        cell_stats['speed_score'] = s_score

                    # print('Spatial Tuning Curve')
                    if tasks['hd_score'] or tasks['tuning_curve']:
                        tuned_data, spike_angles, angular_occupancy, bin_array = spatial_tuning_curve(spatial_spike_train)
                        cell_stats['tuned_data'] = tuned_data
                        cell_stats['tuned_data_angles'] = spike_angles
                        cell_stats['angular_occupancy'] = angular_occupancy
                        cell_stats['angular_occupancy_bins'] = bin_array

                    # print('HD score')
                    if tasks['hd_score']:
                        hd_hist = hd_score(spatial_spike_train)
                        cell_stats['hd_hist'] = hd_hist

                    # print('Grid score')
                    if tasks['grid_score']:
                        true_grid_score = grid_score(spatial_spike_train)
                        cell_stats['grid_score'] = true_grid_score

                    # print('Border score')
                    if tasks['border_score'] and not tasks['disk_arena']:
                        b_score = border_score(spatial_spike_train)
                        cell_stats['b_score_top'] = b_score[0]
                        cell_stats['b_score_bottom'] = b_score[1]
                        cell_stats['b_score_left'] = b_score[2]
                        cell_stats['b_score_right'] = b_score[3]
                    elif tasks['border_score'] and tasks['disk_arena']:
                        print('Cannot compute border score on disk arena')

                    # print('Field sizes')
                    if tasks['field_sizes']:
                        image, n_labels, labels, centroids, field_sizes = map_blobs(spatial_spike_train)
                        cell_stats['field_size_data'] = {'image': image, 'n_labels': n_labels, 'labels': labels, 'centroids': centroids, 'field_sizes': field_sizes}

                    cell.stats_dict['cell_stats'] = cell_stats

                    if settings_dict['saveData'] == True:

                        if plotTasks['rate_map']:
                            # if tasks['disk_arena']:
                            #     rate_map = disk_mask(rate_map)
                            colored_ratemap = Image.fromarray(np.uint8(cm.jet(rate_map)*255))
                            colored_ratemap.save(tetrode_directory_paths[directory] + '/ratemap_cell_' + str(cell.cluster.cluster_label) + '.png')
                        
                        if plotTasks['occupancy_map']:
                            # if tasks['disk_arena']:
                            #     occ_map = disk_mask(occ_map)
                            colored_occupancy_map = Image.fromarray(np.uint8(cm.jet(occ_map)*255))
                            colored_occupancy_map.save(root_directory_paths['Occupancy_Map'] + '/pospdf_cell' + str(cell.cluster.cluster_label) + '.png')

                        # Binary ratemap
                        if plotTasks['binary_map']:
                            # if tasks['disk_arena']:
                            #     binmap = disk_mask(binmap)
                            im = Image.fromarray(np.uint8(binmap*255))
                            im.save(tetrode_directory_paths[directory] + '/Binary_Map_Cell_' + str(cell.cluster.cluster_label) + '.png')

                        # Sparsity, Selectivity and Shannon
                        if tasks['sparsity']:
                            #current_statistics_sheet.range(headers_dict['Sparsity'] + str(excel_cell_index+1)).value = ratemap_stats_dict['sparsity']
                            current_statistics_sheet[headers_dict['sparsity'] + str(excel_cell_index+1)] = ratemap_stats_dict['sparsity']


                        if tasks['selectivity']:
                            #current_statistics_sheet.range(headers_dict['Selectivity'] + str(excel_cell_index+1)).value = ratemap_stats_dict['selectivity']
                            current_statistics_sheet[headers_dict['selectivity'] + str(excel_cell_index+1)] = ratemap_stats_dict['selectivity']


                        if tasks['information']:
                            #current_statistics_sheet.range(headers_dict['Information'] + str(excel_cell_index+1)).value = ratemap_stats_dict['spatial_information_content']
                            current_statistics_sheet[headers_dict['information'] + str(excel_cell_index+1)] = ratemap_stats_dict['spatial_information_content']

                        if tasks['spike_width']:
                            current_statistics_sheet[headers_dict['spike_width'] + str(excel_cell_index+1)] = spike_width
                            current_statistics_sheet[headers_dict['firing_rate'] + str(excel_cell_index+1)] = firing_rate

                        if tasks['spike_analysis']:
                            current_statistics_sheet[headers_dict['bursting'] + str(excel_cell_index+1)] = bursting
                            current_statistics_sheet[headers_dict['Avg. Spikes/Burst'] + str(excel_cell_index+1)] = avg_spikes_per_burst
                            current_statistics_sheet[headers_dict['L_ratio'] + str(excel_cell_index+1)] = ratio_L
                            current_statistics_sheet[headers_dict['iso_dist'] + str(excel_cell_index+1)] = iso_dist
                            current_statistics_sheet[headers_dict['ISI_min'] + str(excel_cell_index+1)] = ISI_dict['min']
                            current_statistics_sheet[headers_dict['ISI_max'] + str(excel_cell_index+1)] = ISI_dict['max']
                            current_statistics_sheet[headers_dict['ISI_median'] + str(excel_cell_index+1)] = ISI_dict['median']
                            current_statistics_sheet[headers_dict['ISI_mean'] + str(excel_cell_index+1)] = ISI_dict['mean']
                            current_statistics_sheet[headers_dict['ISI_std'] + str(excel_cell_index+1)] = ISI_dict['std']
                            current_statistics_sheet[headers_dict['ISI_cv'] + str(excel_cell_index+1)] = ISI_dict['cv']
                            current_statistics_sheet[headers_dict['spike_count'] + str(excel_cell_index+1)] = spike_count
                            

                        # autocorrelation map
                        if plotTasks['autocorr_map']:
                            im = Image.fromarray(np.uint8(cm.jet(autocorr_map)*255))
                            im.save(tetrode_directory_paths[directory] + '/autocorr_cell_' + str(cell.cluster.cluster_label) + '.png')

                        if tasks['border_score'] and not tasks['disk_arena']:
                            current_statistics_sheet[headers_dict['border_score_top'] + str(excel_cell_index+1)] = b_score[0]
                            #current_statistics_sheet.range(headers_dict['Border_Score_Bottom'] + str(excel_cell_index+1)).value = b_score[1]
                            current_statistics_sheet[headers_dict['border_score_bottom'] + str(excel_cell_index+1)] = b_score[1]
                            #current_statistics_sheet.range(headers_dict['Border_Score_Left'] + str(excel_cell_index+1)).value = b_score[2]
                            current_statistics_sheet[headers_dict['border_score_left'] + str(excel_cell_index+1)] = b_score[2]
                            #current_statistics_sheet.range(headers_dict['Border_Score_Right'] + str(excel_cell_index+1)).value = b_score[3]
                            current_statistics_sheet[headers_dict['border_score_right'] + str(excel_cell_index+1)] = b_score[3]

                        # Coherence
                        if tasks['coherence']:
                            current_statistics_sheet[headers_dict['coherence'] + str(excel_cell_index+1)] = coherence

                        # Speed score
                        if tasks['speed_score']:
                            current_statistics_sheet[headers_dict['speed_score'] + str(excel_cell_index+1)] = s_score[0]['2016']

                        # Head direction scores
                        if tasks['hd_score']:
                            current_statistics_sheet[headers_dict['hd_score'] + str(excel_cell_index+1)] = get_hd_score_for_cluster(hd_hist)

                        # Grid score
                        if tasks['grid_score']:
                            current_statistics_sheet[headers_dict['grid_score'] + str(excel_cell_index+1)] = true_grid_score
                            shuffle_check = True

                        # Field sizes
                        if tasks['field_sizes']:
                            # image, n_labels, labels, centroids, field_sizes = get_map_blobs(rate_map_smooth)
                            # prev = None
                            # prevZ = False
                            # extension_list = ['AA','AB','AC','AD','AE']
                            first_fields_index = column_index_from_string(headers_dict['field_sizes']) + 1

                            for j, field_size in enumerate(cell_stats['field_size_data']['field_sizes']):
                                # if not prevZ:
                                #     if prev == 'Z':
                                #         new_j = 0 
                                #         prevZ = True
                                #         column_letter = extension_list[new_j]
                                #     else:
                                #         column_letter = chr( ord( headers_dict['field_sizes']) + j+1 )
                                # else:
                                col_check_letter = get_column_letter(first_fields_index + j)
                                # print(current_statistics_sheet[col_check_letter + "1"], 'Field_' + str(j+1))
                                if current_statistics_sheet[col_check_letter + "1"] != 'Field_' + str(j+1):
                                    # column_letter = get_column_letter(current_statistics_sheet.max_column + 1)
                                    current_statistics_sheet[col_check_letter + "1"] = 'Field_' + str(j+1)

                                current_statistics_sheet[col_check_letter + str(excel_cell_index+1)] = field_size
                                # print(column_letter, j, headers_dict['field_sizes'])                    
                                # prev = column_letter

                        # Firing rate vs. time, for a window of 400 millieconds
                        if plotTasks['Firing_Rate_vs_Time_Plots']:
                            firing_rate, firing_time = firing_rate_vs_time(spatial_spike_train.spike_times, pos_t, 400)
                            fig = plt.figure()
                            figure = plt.gcf()
                            figure.set_size_inches(4, 4)
                            plt.title('Firing Rate vs. Time')
                            plt.xlabel('Time (seconds)')
                            plt.ylabel('Firing Rate (Hertz)')
                            # print
                            plt.plot(pos_t, firing_rate, linewidth=0.25)
                            plt.savefig(tetrode_directory_paths[directory] + '/FRvT_cell_' + str(cell.cluster.cluster_label) + '.png', dpi=300, bbox_inches = 'tight')
                            plt.close(fig)

                        # Firing rate vs. speed
                        if plotTasks['Firing_Rate_vs_Speed_Plots']:
                            fig = plt.figure()
                            figure = plt.gcf()
                            figure.set_size_inches(4, 4)
                            plt.title('Firing Rate vs. Speed')
                            plt.xlabel('Speed (cm/s)')
                            plt.ylabel('Firing Rate (Hertz)')
                            plt.scatter(v, firing_rate, s = 0.5)
                            plt.savefig(tetrode_directory_paths[directory] + '/FRvS_cell_' + str(cell.cluster.cluster_label) + '.png', dpi=300, bbox_inches = 'tight')
                            plt.close(fig)

                        # Plotting tuning curves (polar plots for directional firing) per cell
                        if plotTasks['Tuning_Curve_Plots']:
                            tuned_data = cell_stats['tuned_data'] 
                            spike_angles = cell_stats['tuned_data_angles']
                            angular_occupancy = cell_stats['angular_occupancy']
                            bin_array = cell_stats['angular_occupancy_bins']
                            last_and_first_averaged = (tuned_data[0]+tuned_data[-1]) / 2
                            tuned_data[0] = last_and_first_averaged
                            tuned_data[-1] = last_and_first_averaged

                            fig, ax = plt.subplots(subplot_kw={'projection': 'polar'})
                            ax.plot(bin_array, tuned_data, linewidth=6)
                            ax.set_xticks(np.arange(0,2.0*np.pi,np.pi/2.0))
                            ax.set_yticks(np.linspace(  0, max(tuned_data), 2)[-1]  )
                            # fig = plt.figure()
                            # figure = plt.gcf()
                            # figure.set_size_inches(4, 4)
                            plt.title("Polar plot")
                            # plt.polar(bin_array, tuned_data, linewidth=3)

                            plt.box(on=None)
                            plt.savefig(tetrode_directory_paths[directory] + '/tuning_curve_cell_' + str(cell.cluster.cluster_label) + '.png', dpi=300, bbox_inches = 'tight')
                            plt.close(fig)

                        # Spikes over position map
                        if plotTasks['Spikes_Over_Position_Map']:
                            fig = plt.figure()
                            figure = plt.gcf()
                            figure.set_size_inches(4, 4)
                            plt.plot(pos_x,pos_y, linewidth=0.2)
                            plt.scatter(spikex,spikey, c='r', s=5, zorder=3)
                            plt.title("Spikes over position")
                            plt.xlabel("x coordinates")
                            plt.ylabel("y coordinates")
                            plt.gca().set_aspect('equal', adjustable='box')
                            plt.savefig(tetrode_directory_paths[directory] + '/spikes_over_position_cell_' + str(cell.cluster.cluster_label) + '.png', dpi=300, bbox_inches = 'tight')
                            plt.close(fig)

                        if plotTasks['Waveforms_Across_Channels']:
                            fig = WaveformTemplateFig()

                            for i in range(4):
                                ch = cell.signal[:,i,:]
                                idx = np.random.choice(len(ch), size=200)
                                waves = ch[idx, :]
                                avg_wave = np.mean(ch, axis=0)

                                fig.waveform_channel_plot(waves, avg_wave, str(i+1), fig.ax[str(i+1)])

                            cell_id = cell.cluster.cluster_label
                            title = str(animal_id) + '_tetrode_' + str(tet_id) + '_' + str(session_key) + '_unit_' + str(cell_id)

                            fig.f.suptitle(title, ha='center', fontweight='bold', fontsize='large')
                            plt.savefig(tetrode_directory_paths[directory] + '/waveforms_across_channels' + str(cell.cluster.cluster_label) + '.png', dpi=300, bbox_inches = 'tight')
                            plt.close(fig.f)


                        # Auto-resize columns to width of header text
                        #current_statistics_sheet.autofit(axis="columns")
                        ColumnDimension(current_statistics_sheet, bestFit=True)

                        print("Cell " + str(excel_cell_index) + " is complete")
                        per_session_tracker += 1
                        per_animal_tracker += 1
                        per_animal_tetrode_tracker += 1
                #     # -------------------------------------------------------- #

                c += 1

            k += 1

            if settings_dict['saveMethod'] == 'one_per_session':
                if animal_tet_count[animal_id] == animal_max_tet_count[animal_id]:
                    list(map(lambda x: _save_wb(animal_workbooks[animal_id][x], save_path), animal_workbooks[animal_id]))

        if settings_dict['saveMethod'] == 'one_per_animal_tetrode':
            if animal_tet_count[animal_id] == animal_max_tet_count[animal_id]:
                _save_wb(wb, root_path, animal_id=animal_id)

    if settings_dict['saveMethod'] == 'one_for_parent':
        _save_wb(wb, root_path, sum_sheet_count=sum_sheet_count)
        # wb._sheets = sorted(wb._sheets, key=lambda x: x.title)
        # print(root_path)
        # pth = root_path + '/summary_sheet'  + '.xlsx'
        # wb.save(pth)
        # wb.close()

        # xls = pd.ExcelFile(pth)
        # df_sum = pd.read_excel(xls, 'Summary')
        # df_sum = df_sum.sort_values(['Session', 'Tetrode', 'Cell'])
        # writer = pd.ExcelWriter(pth, engine="xlsxwriter")
        # df_sum.to_excel(writer, sheet_name="Summary", index=False)
        # writer.save()

# def run_spk_analysis(animal_tets):
#     ch1, ch2, ch3, ch4 = animal_tets
#     data_concat = np.vstack((ch1, ch2, ch3, ch4)).reshape((4, -1, ch1.shape[1]))
    
#     # cluster_batch = session.get_spike_data()['spike_cluster']
#     FD = create_features(data_concat)
#     L_ratio(cluster_batch)
#     iso_dist = isolation_distance(cluster_batch)
# animal_sessions_tets_events


def _save_wb(wb, root_path, animal_id=None, sum_sheet_count=None):
    wb._sheets = sorted(wb._sheets, key=lambda x: x.title)
    if animal_id is None:
        if sum_sheet_count is None:
            pth = root_path + '/summary_sheet'  + '.xlsx'
        else:
            pth = root_path + '/summary_sheet_'  + str(sum_sheet_count) + '.xlsx'
    else:
        pth = root_path + '/summary_sheet_' + str(animal_id)  + '.xlsx'
    print(root_path)
    wb.save(pth)
    wb.close()

    # xls = pd.ExcelFile(pth)
    # df_sum = pd.read_excel(xls, 'Summary')
    # writer = pd.ExcelWriter(pth, engine="xlsxwriter")
    # dfs = []
    # for sheet in xls.sheet_names:
    #     if sheet != 'Summary' and sheet != 'summary':
    #         xls = pd.ExcelFile(pth)
    #         df = pd.read_excel(xls, sheet)
    #         dfs.append(df) 

    #         df = df.sort_values(['Session', 'Tetrode', 'Cell ID'])
    #         df.to_excel(writer, sheet_name=sheet, index=True)
    #         # writer.save()

    xls = pd.read_excel(pth, sheet_name=None)
    df_sum = xls.pop('Summary')
    dfs = [df.sort_values(['Session', 'Tetrode', 'Cell ID']) for df in xls.values()]
    with pd.ExcelWriter(pth, engine='xlsxwriter') as writer:
        df_sum.to_excel(writer, sheet_name='Summary', index=False)
        for sheet, df in zip(xls.keys(), dfs):
            df.to_excel(writer, sheet_name=sheet, index=False)


    
        if len(dfs) > 0:
            df_sum = pd.concat(dfs, axis=0).sort_values(['Session', 'Tetrode', 'Cell ID'])
        else:
            df_sum = df_sum.sort_values(['Session', 'Tetrode', 'Cell ID'])
        df_sum.to_excel(writer, sheet_name="Summary", index=True)
    # writer.save()
    print('Saved ' + str(pth))

def get_hd_score_for_cluster(hd_hist):
    angles = np.linspace(-179, 180, 360)
    angles_rad = angles*np.pi/180
    dy = np.sin(angles_rad)
    dx = np.cos(angles_rad)

    totx = sum(dx * hd_hist)/sum(hd_hist)
    toty = sum(dy * hd_hist)/sum(hd_hist)
    r = np.sqrt(totx*totx + toty*toty)
    return r

class WaveformTemplateFig():
    def __init__(self):
        self.f = plt.figure(figsize=(12, 6))
        # mpl.rc('font', **{'size': 20})


        self.gs = {
            'all': gridspec.GridSpec(1, 4, left=0.05, right=0.95, bottom=0.1, top=0.85, figure=self.f),
        }

        self.ax = {
            '1': self.f.add_subplot(self.gs['all'][:, :1]),
            '2': self.f.add_subplot(self.gs['all'][:, 1:2]),
            '3': self.f.add_subplot(self.gs['all'][:, 2:3]),
            '4': self.f.add_subplot(self.gs['all'][:, 3:]),
        }

    def waveform_channel_plot(self, waveforms, avg_waveform, channel, ax):

        ax.plot(waveforms.T, color='grey')

        ax.plot(avg_waveform, c='k', lw=2)

        ax.set_title('Channel ' + str(int(channel)))


if __name__ == '__main__':

    ######################################################## EDIT BELOW HERE ########################################################

    # MAKE SURE "field_sizes" IS THE LAST ELEMENT IN "csv_header_keys"
    csv_header = {}
    csv_header_keys = ['spike_width', 'spike_count', 'firing_rate', 'Avg. Spikes/Burst', 'bursting', 'iso_dist', 'L_ratio', 'ISI_min', 'ISI_max', 'ISI_mean', 'ISI_median', 'ISI_cv', 'ISI_std',
                       'sparsity', 'selectivity', 'information', 'coherence', 'speed_score', 'hd_score', 'grid_score', 'border_score', 'field_sizes']
    for key in csv_header_keys:
        csv_header[key] = True

    tasks = {}
    task_keys = ['spike_width', 'spike_analysis', 'binary_map', 'autocorrelation_map', 'sparsity', 'selectivity', 'information', 'coherence', 'speed_score', 'hd_score', 'tuning_curve', 'grid_score', 'border_score', 'field_sizes', 'disk_arena']
    for key in task_keys:
        tasks[key] = True

    plotTasks = {}
    plot_task_keys = ['Waveforms_Across_Channels', 'Spikes_Over_Position_Map', 'Tuning_Curve_Plots', 'Firing_Rate_vs_Speed_Plots', 'Firing_Rate_vs_Time_Plots','autocorr_map', 'binary_map','rate_map', 'occupancy_map']
    for key in plot_task_keys:
        plotTasks[key] = True

    animal = {'animal_id': '001', 'species': 'mouse', 'sex': 'F', 'age': 1, 'weight': 1, 'genotype': 'type', 'animal_notes': 'notes'}
    devices = {'axona_led_tracker': True, 'implant': True}
    implant = {'implant_id': '001', 'implant_type': 'tetrode', 'implant_geometry': 'square', 'wire_length': 25, 'wire_length_units': 'um', 'implant_units': 'uV'}

    session_settings = {'channel_count': 4, 'animal': animal, 'devices': devices, 'implant': implant}

    """ FOR YOU TO EDIT """
    settings = {'ppm': 485, 'session':  session_settings, 'smoothing_factor': 3, 'useMatchedCut': True}
    """ FOR YOU TO EDIT """

    settings['disk_arena'] = True # -->

    settings['tasks'] = tasks # --> change tasks array to change tasks are run
    settings['plotTasks'] = plotTasks # --> change plot tasks array to change asks taht are plotted
    settings['header'] = csv_header # --> change csv_header header to change tasks that are saved to csv

    """ FOR YOU TO EDIT """
    settings['naming_type'] = 'LEC'
    settings['arena_size'] = None
    settings['speed_lowerbound'] = 0
    settings['speed_upperbound'] = 99
    settings['end_cell'] = None
    settings['start_cell'] = None
    settings['saveData'] = True
    settings['saveMethod'] = 'one_for_parent'
    # possible saves are:
    # 1 csv per session (all tetrode and indiv): 'one_per_session' --> 5 sheets (summary of all 4 tetrodes, tet1, tet2, tet3, tet4)
    # 1 csv per animal per tetrode (all sessions): 'one_per_animal_tetrode' --> 4 sheets one per tet 
    # 1 csv per animal (all tetrodes & sessions): 'one_for_parent' --> 1 sheet
    """ FOR YOU TO EDIT """


    start_time = time.time()
    root = tk.Tk()
    root.withdraw()
    data_dir = filedialog.askdirectory(parent=root,title='Please select a data directory.')

    ########################################################################################################################

    """ OPTION 1 """
    """ RUNS EVERYTHING UNDER PARENT FOLDER (all subfolders loaded first) """
    # study = make_study(data_dir,settings_dict=settings)
    # study.make_animals()
    # batch_map(study, settings, data_dir)

    """ OPTION 2 """
    """ RUNS EACH SUBFOLDER ONE AT A TIME """
    subdirs = np.sort([ f.path for f in os.scandir(data_dir) if f.is_dir() ])
    count = 1
    for subdir in subdirs:
        try:
            study = make_study(subdir,settings_dict=settings)
            study.make_animals()
            batch_map(study, settings, subdir, sum_sheet_count=count)
            count += 1
        except Exception:
            print(traceback.format_exc())
            print('DID NOT WORK FOR DIRECTORY ' + str(subdir))
import os
import sys

from library import spatial

PROJECT_PATH = os.getcwd()
sys.path.append(PROJECT_PATH)
 

import numpy as np
from library.shuffle_spikes import shuffle_spikes
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils.cell import get_column_letter
# from library.spatial_spike_train import SpatialSpikeTrain2D
from library.maps import binary_map
from library.hafting_spatial_maps import HaftingRateMap, SpatialSpikeTrain2D


# def border_score(binary_map: np.ndarray, rate_map: np.ndarray) -> tuple:
def border_score(spatial_spike_train: SpatialSpikeTrain2D, **kwargs) -> tuple:

    '''
        Computes 4 scores which each reflect selectivity of neurons firing at arena edges,
        namely top, bottom, left, and right border scores for each arena side.
        Score range: -1 to 1, with 1 being the highest for border selectivity.

        More details about this scoring criteria can be found under the following publication at the border score section:
        "Grid and Nongrid Cells in Medial Entorhinal Cortex Represent Spatial
            Location and Environmental Features with Complementary Coding Schemes"

        Params:
            binary_map (np.ndarray):
                Map of binarized firing fields in ratemap
            rate_map (np.ndarray):
                Array encoding neuron spike events in 2D space based on where
                the subject walked during experiment.

        Returns:
            tuple: top_bscore, bottom_bscore, left_bscore, right_bscore
    '''

    if 'use_objects_directly' in kwargs:
        if kwargs['use_objects_directly'] == True:
            rate_map = kwargs['rate_map']
            bin_map = kwargs['bin_map']
    else:
        if 'smoothing_factor' in kwargs:
            smoothing_factor = kwargs['smoothing_factor']
        else:
            smoothing_factor = spatial_spike_train.session_metadata.session_object.smoothing_factor

        rate_map, _ = spatial_spike_train.get_map('rate').get_rate_map(smoothing_factor)

        if spatial_spike_train.get_map('binary') is None:
            bin_map = binary_map(spatial_spike_train)
            spatial_spike_train.add_map_to_stats('binary', bin_map)
        bin_map = spatial_spike_train.get_map('binary')

        # bin_map = binary_map(spatial_spike_train)
    


    # If for whatever reason the supplied binary map does not match rate map dimensions, throw error.
    if bin_map.shape != rate_map.shape:
        raise Exception("The binary map and rate map must have the same dimensions")

    shortest_side = min(bin_map.shape[0], bin_map.shape[1]) / 2

    # Initializing top, bottom, left, right proportion coverage of map
    top_coverage = 0
    bottom_coverage = 0
    left_coverage = 0
    right_coverage = 0

    # Initializing distances
    top_distance_sum = 0
    bottom_distance_sum = 0
    left_distance_sum = 0
    right_distance_sum = 0

    # Compute coverage on each side
    top_coverage = sum(bin_map[0]) / len(bin_map[0])
    bottom_coverage = sum(bin_map[len(bin_map)-1]) / len(bin_map[0])
    left_coverage = sum(bin_map[:,0]) / len(bin_map[:,0])
    right_coverage = sum(bin_map[:,bin_map.shape[1]-1]) / len(bin_map[:,0])

    indices = np.argwhere(bin_map > 0)

    # Compute distances of all field pixels to each edge
    for index in indices:
        top_distance_sum += (index[0]) *  rate_map[index[0], index[1]]
        bottom_distance_sum += (len(bin_map[0]) - index[0]) * rate_map[index[0], index[1]]
        left_distance_sum += index[1] * rate_map[index[0], index[1]]
        right_distance_sum += (len(bin_map[:,0]) - index[1]) * rate_map[index[0], index[1]]

    avg_top_dist    = (top_distance_sum / len(indices))     / shortest_side
    avg_bottom_dist = (bottom_distance_sum / len(indices))  / shortest_side
    avg_left_dist   = (left_distance_sum / len(indices))    / shortest_side
    avg_right_dist  = (right_distance_sum / len(indices))   / shortest_side

    # Compute the border score
    top_bscore = (top_coverage - avg_top_dist) / (top_coverage + avg_top_dist)
    bottom_bscore = (bottom_coverage - avg_bottom_dist) / (bottom_coverage + avg_bottom_dist)
    left_bscore = (left_coverage - avg_left_dist) / (left_coverage + avg_left_dist)
    right_bscore = (right_coverage - avg_right_dist) / (right_coverage + avg_right_dist)

    return top_bscore, bottom_bscore, left_bscore, right_bscore

#public
def border_score_shuffle(self, occupancy_map: np.ndarray, arena_size: tuple, ts: np.ndarray,
                         pos_x: np.ndarray, pos_y: np.ndarray, pos_t: np.ndarray, kernlen: int, std: int, **kwargs) -> list:

    '''
        Shuffles position and spike data prior to computing border scores. Shuffling
        allows us to determine the probability if a given score is random, or
        demonstrates a meaningful association in the data.

        Params:
            occupancy_map (np.ndarray):
                A 2D numpy array enconding subjects position over entire experiment.
            arena_size (tuple):
                Dimensions of arena
            ts (np.ndarray):
                Spike time stamp array
            pos_x, pos_y, pos_t (np.ndarray):
                Arrays of x,y  coordinate positions as well as timestamps of movement respectively
            kenrnlen, std (int):
                kernel size and standard deviation of kernel for convolutional smoothing.

        **kwargs:
            xsheet: xlwings excel sheet

        Returns:
            list: border_scores
            --------
            border_scores: List of 100 border scores (1 set of scores [top,bottom,left,right] per shuffle)
    '''

    # If an excel sheet was passed in, set reference
    s = kwargs.get('xsheet', None)

    row_index = kwargs.get('row_index',None)
    cell_number = kwargs.get('cell_number',None)
    column_index = 2

    border_scores = []

    # Shuffle the spike data
    shuffled_spikes = shuffle_spikes(self, ts, pos_x, pos_y, pos_t)

    # For each shuffled set of spiek data, compute border score
    for element in shuffled_spikes:
            rate_map, _ = rate_map(pos_x, pos_y, pos_t, arena_size, element[0], element[1], kernlen, std)
            binary_map = binary_map(rate_map)
            b_score = border_score(binary_map, rate_map)
            border_scores.append(b_score)

            # Copy shuffled border scores into excel sheet
            if s != None and row_index != None and cell_number != None:
                s[get_column_letter(row_index)   + str(1)] = 'C_' + str(cell_number) + '_BS_Top'
                s[get_column_letter(row_index+1) + str(1)] = 'C_' + str(cell_number) + '_BS_Bottom'
                s[get_column_letter(row_index+2) + str(1)] = 'C_' + str(cell_number) + '_BS_Left'
                s[get_column_letter(row_index+3) + str(1)] = 'C_' + str(cell_number) + '_BS_Right'

                s[get_column_letter(row_index)   + str(column_index)] = b_score[0]
                s[get_column_letter(row_index+1) + str(column_index)] = b_score[1]
                s[get_column_letter(row_index+2) + str(column_index)] = b_score[2]
                s[get_column_letter(row_index+3) + str(column_index)] = b_score[3]

                column_index += 1
                # Resize excel columns per iteration
                #s.autofit(axis="columns")
                ColumnDimension(s, bestFit=True)

    return border_scores

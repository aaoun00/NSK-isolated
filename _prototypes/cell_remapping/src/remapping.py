import os, sys
import copy

# data manipulation
import numpy as np
import pandas as pd

# file saving/loading/naming
from openpyxl import load_workbook
import re

# stats
from scipy import stats, ndimage

# settings based
from skimage.measure import block_reduce

# unused/relevant code commented out
from scipy.spatial.distance import cdist
from pyemd import emd
import itertools
import matplotlib.pyplot as plt
import cv2
import xlsxwriter

PROJECT_PATH = os.getcwd()
sys.path.append(PROJECT_PATH)

# from neuroscikit 
from library.hafting_spatial_maps import SpatialSpikeTrain2D
from _prototypes.cell_remapping.src.rate_map_plots import plot_obj_remapping, plot_regular_remapping, plot_fields_remapping, plot_shuffled_regular_remapping, plot_matched_sesssion_waveforms
from _prototypes.cell_remapping.src.wasserstein_distance import compute_temporal_emd, single_point_wasserstein, pot_sliced_wasserstein, compute_centroid_remapping, _get_ratemap_bucket_midpoints
from _prototypes.cell_remapping.src.masks import binary_mask, make_object_ratemap, flat_disk_mask, generate_grid, _sample_grid, apply_disk_mask
from library.maps import map_blobs
from _prototypes.cell_remapping.src.processing import aggregate_map_blobs, collect_shuffled_ratemaps, make_obj_map_dict, get_rate_map
from _prototypes.cell_remapping.src.utils import scale_points, _check_rotate_evening, _get_spk_pts, _get_valid_weight_bins,_get_valid_midpoints, _get_ratemap_bucket_midpoints, check_disk_arena, _fill_centroid_dict, _copy_labels, _downsample, _get_spk_pts, read_data_from_fname, check_cylinder, check_object_location, _get_valid_weight_bins, _get_valid_midpoints, _check_object_coords, _check_rotate_evening
from _prototypes.cell_remapping.src.stats import get_vector_from_map, wasserstein_quantile, compute_modified_zscore, get_ref_change_stats, get_ref_ratio_stats, compute_cumulative_blob_stats, get_max_matched_cell_count, wasserstein_quantile, get_reference_dist_stats, get_rate_stats

# unused/commented out
from scripts.batch_map.batch_map import batch_map 
from library.shuffle_spikes import shuffle_spikes

# SETTINGS FILE
from _prototypes.cell_remapping.src.settings import obj_output, centroid_output, tasks, session_comp_categories, regular_output, context_output, variations, temporal_output, context_temporal_output


                            
def compute_remapping(study, settings, data_dir):

    # aside
    # collects stacked spike counts for each group and session 
    # stack_dict = {}
    # stack_dict['control'] = {}
    # stack_dict['app_ki'] = {}
    # stack_dict['control']['session_1'] = []
    # stack_dict['control']['session_2'] = []
    # stack_dict['control']['session_3'] = []
    # stack_dict['control']['session_4'] = []
    # stack_dict['app_ki']['session_1'] = []
    # stack_dict['app_ki']['session_2'] = []
    # stack_dict['app_ki']['session_3'] = []
    # stack_dict['app_ki']['session_4'] = []

    isStart = True
    context_paths = {}
    context_temporal_paths = {}

    ratemap_size = settings['ratemap_dims'][0]

    animal_cell_info, animal_cell_ratemaps = aggregate_map_blobs(study, settings)

    animal_ref_dist = collect_shuffled_ratemaps(animal_cell_ratemaps, settings)

    for animal in study.animals:

        animal_id = animal.animal_id.split('_')[0]

        # need to add separate option for circular shuffling
        """ UNCOMMENT TO RUN CIRCULAR SHUFFLING REF DIST"""
        """ NEED TO GO LINE BY LINE AND set ref_wass_dist to come from list(map( of shuffled samples)) """
        # max_centroid_count, blobs_dict, shuffled_ratemap_dict, shuffled_sample_dict = _aggregate_cell_info(animal, settings)
        # max_centroid_count, blobs_dict, _, _ = _aggregate_cell_info(animal, settings)
       
        max_centroid_count, blobs_dict = animal_cell_info[animal.animal_id]


        max_matched_cell_count = get_max_matched_cell_count(animal)

        # for every existing cell id across all sessions
        for k in range(int(max_matched_cell_count)):

            centroid_dict = copy.deepcopy(centroid_output)
            regular_dict = copy.deepcopy(regular_output)
            context_dict = copy.deepcopy(context_output)
            obj_dict = copy.deepcopy(obj_output)
            temporal_dict = copy.deepcopy(temporal_output)
            context_temporal_dict = copy.deepcopy(context_temporal_output)

            cell_label = k + 1
            print('Cell ' + str(cell_label))

            prev = None
            prev_spike_times = None
            curr_spike_times = None
            prev_id = None
            curr_shuffled = None
            cell_session_appearances = []
            
            if settings['ses_limit'] is None:
                ses_limit = len(list(animal.sessions.keys()))
            else:
                if settings['ses_limit'] >= len(list(animal.sessions.keys())):
                    ses_limit = len(list(animal.sessions.keys()))
                else:
                    ses_limit = settings['ses_limit']

            # for every session
            # for i in range(len(list(animal.sessions.keys()))):
            for i in range(ses_limit):
                # seskey = 'session_' + str(i+1)
                seskey = list(animal.sessions.keys())[i]
                print(seskey)
                ses = animal.sessions[seskey]
                path = ses.session_metadata.file_paths['tet']
                fname = path.split('/')[-1].split('.')[0]


                cylinder = check_cylinder(fname, settings['disk_arena'])
        
                stim, depth, name, date = read_data_from_fname(fname, settings['naming_type'], settings['type'])

                object_location = check_object_location(stim, settings['hasObject'])

                ensemble = ses.get_cell_data()['cell_ensemble']
                print(cell_label, ensemble.get_cell_label_dict())
                # Check if cell id we're iterating through is present in the ensemble of this sessions
                if cell_label in ensemble.get_cell_label_dict():
                    
                    cell = ensemble.get_cell_by_id(cell_label)
                    cell_session_appearances.append(cell)
                    spatial_spike_train = cell.stats_dict['spatial_spike_train'] 

                    rate_map_obj = spatial_spike_train.get_map('rate')

                    rate_map = get_rate_map(rate_map_obj, settings['ratemap_dims'], settings['normalizeRate'])

                    curr, curr_ratemap, disk_ids = apply_disk_mask(rate_map, settings, cylinder)
                    
                    curr_cell = cell
                    curr_spatial_spike_train = spatial_spike_train
                    curr_key = seskey
                    curr_path = ses.session_metadata.file_paths['tet'].split('/')[-1].split('.')[0]
                    curr_id = str(animal.animal_id) + '_' + str(seskey) + '_' + str(cell.cluster.cluster_label)

                    """ Saves stacked matrix of spike counts """
                    # # REMOVE
                    # app = ['1-13','1-14','1a-27', '1-30', '1-35', '1a-37']
                    # control = ['1-20','1-24','1-25', '1-28', '1-34', '1a23-S30-31', '1a-40']

                    # aiduse = str(animal.animal_id)
                    # aiduse = aiduse.split('_tet')[0].replace('_','-')
                    # if aiduse in control:
                    #     control_or_app = 'control'
                    # else:
                    #     assert aiduse in app, 'Animal id not in app ' + aiduse
                    #     control_or_app = 'app_ki'
                    # cts, _ = np.histogram(cell.event_times, bins=np.arange(0,1205,1))
                    # binned_times = cts / np.mean(cts)
                    # if len(stack_dict[control_or_app][seskey]) == 0:
                    #     stack_dict[control_or_app][seskey] = binned_times
                    # else:
                    #     stack_dict[control_or_app][seskey] = np.vstack((stack_dict[control_or_app][seskey], binned_times))
                    # # REMOVE

                    y, x = curr.shape
                    h, w = rate_map_obj.arena_size
                    # print(h, w)
                    bin_area = h/y * w/x

                    # If doing object remapping (i.e. synthetic ratemap with all firing in one bin)
                    if settings['hasObject']:
                        
                        # get relevant centroid infomarition (for field reestricted EMD)
                        _, _, labels, centroids, field_sizes = blobs_dict[curr_id]

                        assert len(np.unique(labels)[1:]) == len(centroids) == len(field_sizes), 'Mismatch in number of labels, centroids and field sizes'

                        labels_copy, c_count = _copy_labels(labels, curr)
              
                        cumulative_coverage, cumulative_area, cumulative_rate = compute_cumulative_blob_stats(labels_copy, curr)

                        height_bucket_midpoints, width_bucket_midpoints = _get_ratemap_bucket_midpoints(rate_map_obj.arena_size, y, x)

                        curr_labels = labels

                        resampled_positions = generate_grid(rate_map_obj.arena_size[0], rate_map_obj.arena_size[1], 
                                                                settings['spacing'], is_hexagonal=settings['hexagonal'], is_cylinder=cylinder)

                        print('Resampled positions: ', len(resampled_positions), len(resampled_positions[0]))

                        obj_map_dict = make_obj_map_dict(variations, settings, cylinder, disk_ids)
                        
                        for obj_score in settings['object_scores']:
                            unq_labels = np.unique(labels)[1:]
                            for lid in range(len(unq_labels)):
                                label_id = unq_labels[lid]
                                true_object_pos = None
                                true_object_ratemap = None

                                # # Possible object locations (change to get from settings)
                                # variations = [0,90,180,270,'no']
                                # compute object remapping for every object position, actual object location is store alongside wass for each object ratemap
                                resampled_wass = None
                                for var in variations:
                                    obj_wass_key = 'obj_wass_' + str(var)
                                    obj_vector_key = 'obj_vec_' + str(var)
                                    obj_quantile_key = 'obj_q_' + str(var)

                                    object_ratemap, object_pos, disk_ids = obj_map_dict[var]

                                    if var == object_location:
                                        true_object_pos = object_pos
                                        true_object_ratemap = object_ratemap

                                    obj_x, obj_y = _check_object_coords(object_pos, height_bucket_midpoints, width_bucket_midpoints)

                                    y, x = curr.shape
                                    # height_bucket_midpoints, width_bucket_midpoints = _get_ratemap_bucket_midpoints(curr_spatial_spike_train.arena_size, y, x)

                                    # lid 0 we do whole/spike density once and don't repeat forl ater lid
                                    # other lid we only repeat field and binary for new fieldlabel
                                    # EMD on norm/unnorm ratemap + object map for OBJECT remapping
                                    if obj_score == 'whole' and lid == 0:
                                        
                                        obj_wass = single_point_wasserstein(object_pos, curr_ratemap, rate_map_obj.arena_size, ids=disk_ids)
                                
                                        # compute EMD on resamples
                                        if resampled_wass is None:
                                            resampled_wass = list(map(lambda x: single_point_wasserstein(x, curr_ratemap, rate_map_obj.arena_size, ids=disk_ids, use_pos_directly=True), resampled_positions))
                                        quantile = (resampled_wass < obj_wass).mean()
                        
                                        mag, angle, pt1, pt2 = get_vector_from_map(curr, rate_map_obj.arena_size, y, x, obj_y, obj_x, 'whole')

                                    elif obj_score == 'field':

                                        # field ids is ids of binary field/map blolb
                                        # TAKE ONLY MAIN FIELD --> already sorted by size
                                        row, col = np.where(curr_labels == label_id)
                                        field_ids = np.array([row, col]).T

                                        if cylinder:
                                            # print('IT IS A CYLINDER, TAKING ONLY IDS IN FIELD AND IN DISK')
                                            field_disk_ids = np.array([x for x in field_ids if x in disk_ids])
                                        else:
                                            field_disk_ids = field_ids
                                        
                                        obj_wass = single_point_wasserstein(object_pos, curr_ratemap, rate_map_obj.arena_size, ids=field_disk_ids)

                                        # compute EMD on resamples
                                        if resampled_wass is None:
                                            resampled_wass = list(map(lambda x: single_point_wasserstein(x, curr_ratemap, rate_map_obj.arena_size, ids=field_disk_ids, use_pos_directly=True), resampled_positions))
                                        quantile = (resampled_wass < obj_wass).mean()

                                        mag, angle, pt1, pt2 = get_vector_from_map(centroids, rate_map_obj.arena_size, y, x, obj_y, obj_x, 'field')

                                    elif obj_score == 'spike_density' and lid == 0:

                                        curr_spike_pos_x, curr_spike_pos_y = curr_spatial_spike_train.spike_x, curr_spatial_spike_train.spike_y
                                        # curr_spike_pos_x *= -1
                                        # curr_spike_pos_x += np.abs(np.min(curr_spike_pos_x))
                                        # curr_spike_pos_y += np.abs(np.min(curr_spike_pos_y))

                                        if np.min(curr_spike_pos_x) < 0:
                                            curr_spike_pos_x += np.abs(np.min(curr_spike_pos_x))
                                        if np.min(curr_spike_pos_y) < 0:
                                            curr_spike_pos_y += np.abs(np.min(curr_spike_pos_y))
                                        assert np.min(curr_spike_pos_x) >= 0 and np.min(curr_spike_pos_y) >= 0, 'Negative spike positions'
                                
                                        curr_pts = np.array([curr_spike_pos_y, curr_spike_pos_x]).T

                                        obj_wass = single_point_wasserstein(object_pos, curr_ratemap, rate_map_obj.arena_size, density=True, density_map=curr_pts, use_pos_directly=False)

                                        # compute EMD on resamples
                                        if resampled_wass is None:
                                            resampled_wass = list(map(lambda x: single_point_wasserstein(x, curr_ratemap, rate_map_obj.arena_size, density=True, density_map=curr_pts, use_pos_directly=True), resampled_positions))
                                        quantile = (resampled_wass < obj_wass).mean()

                                        mag, angle, pt1, pt2 = get_vector_from_map([curr_spike_pos_y, curr_spike_pos_x], rate_map_obj.arena_size, y, x, obj_y, obj_x, 'spike_density')

                                    elif obj_score == 'binary':

                                        curr_masked, field_disk_ids = binary_mask(curr_labels, label_id, disk_ids, cylinder)

                                        obj_wass = single_point_wasserstein(object_pos, curr_masked, rate_map_obj.arena_size, ids=field_disk_ids)

                                        # compute EMD on resamples
                                        if resampled_wass is None:
                                            resampled_wass = list(map(lambda x: single_point_wasserstein(x, curr_masked, rate_map_obj.arena_size, ids=field_disk_ids, use_pos_directly=True), resampled_positions))
                                        quantile = (resampled_wass < obj_wass).mean()

                                        mag, angle, pt1, pt2 = get_vector_from_map(curr_masked, rate_map_obj.arena_size, y, x, obj_y, obj_x, 'binary')

                                    elif obj_score == 'centroid':

                                        # TAKE ONLY MAIN FIELD --> already sorted by size
                                        main_centroid = centroids[lid]

                                        # euclidean distance between point
                                        obj_wass = np.linalg.norm(np.array((obj_y, obj_x)) - np.array((main_centroid[0],main_centroid[1])))
                                        
                                        if resampled_wass is None:
                                            resampled_wass = list(map(lambda x: np.linalg.norm(np.array((obj_y, obj_x)) - np.array((x[0],x[1]))), resampled_positions))
                                        quantile = (resampled_wass < obj_wass).mean()

                                        mag, angle, pt1, pt2 = get_vector_from_map(main_centroid,rate_map_obj.arena_size, y, x, obj_y, obj_x, 'centroid')

                                    if lid != 0 and (obj_score == 'whole' or obj_score == 'spike_density') == True:
                                        pass 
                                    else:
                                        # obj_dict[obj_wass_key].append([obj_wass, obj_field_wass, obj_bin_wass, c_wass])
                                        obj_dict[obj_wass_key].append(obj_wass)
                                        obj_dict[obj_quantile_key].append(quantile)
                                        obj_dict[obj_vector_key].append([pt1, pt2, mag, angle])

                                # if first centroid label, we can save whole map annd spike density scores, if second or later label, we don't want to resave
                                # the whole map and spike density scores
                                if lid != 0 and (obj_score == 'whole' or obj_score == 'spike_density') == True:
                                    pass 
                                else:
                                    val_r, val_c = np.where(labels == label_id)
                                    field_coverage = len(val_r)/len(np.where(~np.isnan(curr))[0])
                                    field_area = len(val_r)
                                    field_rate = np.sum(curr[val_r, val_c])
                                    total_rate = np.sum(curr)
                                    field_peak_rate = np.max(curr[val_r, val_c])

                                    if obj_score == 'whole' or obj_score == 'spike_density':
                                        obj_dict['field_id'].append('all')
                                    else:
                                        obj_dict['field_id'].append(label_id)

                                    obj_dict['score'].append(obj_score)
                                    obj_dict['field_peak_rate'].append(field_peak_rate)
                                    obj_dict['total_rate'].append(total_rate)
                                    obj_dict['field_coverage'].append(field_coverage)
                                    obj_dict['field_area'].append(field_area)
                                    obj_dict['field_rate'].append(field_rate)
                                    obj_dict['cumulative_coverage'].append(cumulative_coverage)
                                    obj_dict['cumulative_area'].append(cumulative_area)
                                    obj_dict['cumulative_rate'].append(cumulative_rate)
                                    obj_dict['field_count'].append(c_count)
                                    obj_dict['bin_area'].append(bin_area[0])
                                    obj_dict['object_location'].append(object_location)
                                    obj_dict['obj_pos'].append((true_object_pos['x'], true_object_pos['y']))
                                    obj_dict['signature'].append(curr_path)
                                    obj_dict['spike_count'].append(len(curr_spatial_spike_train.spike_times))
                                    obj_dict['name'].append(name)
                                    obj_dict['date'].append(date)
                                    obj_dict['depth'].append(depth)
                                    obj_dict['unit_id'].append(cell_label)
                                    obj_dict['tetrode'].append(animal.animal_id.split('tet')[-1])
                                    obj_dict['session_id'].append(seskey)
                                    obj_dict['arena_size'].append(curr_spatial_spike_train.arena_size)
                                    obj_dict['cylinder'].append(cylinder)
                                    obj_dict['ratemap_dims'].append(curr.shape)
                                    # obj_dict['grid_sample_threshold'].append(settings['grid_sample_threshold'])
                                    obj_dict['spacing'].append(settings['spacing'])
                                    obj_dict['hexagonal'].append(settings['hexagonal'])
                                    obj_dict['sample_size'].append(len(resampled_positions))

                                    if settings['downsample']:
                                        obj_dict['downsample_factor'].append(settings['downsample_factor'])
                                    else:
                                        obj_dict['downsample_factor'].append(1)

                        if settings['plotObject']:
                            plot_obj_remapping(true_object_ratemap, curr, labels, centroids, obj_dict, data_dir)

                    if prev is not None:
                        ses_1 = prev_key.split('_')[1]
                        ses_2 = seskey.split('_')[1]
                        print('search here')
                        print(prev_key, seskey)
                        ses_comp = str(ses_1) + '_' + str(ses_2)
        
                    # If prev ratemap is not None (= we are at session2 or later, session1 has no prev session to compare)
                    if prev is not None and settings['runRegular']:
                        
                        prev_pts = _get_spk_pts(prev_spatial_spike_train)
                        curr_pts = _get_spk_pts(curr_spatial_spike_train)

                        curr_ratemap, curr_pts = _check_rotate_evening(curr_path, curr_pts, curr_ratemap, settings['rotate_evening'], settings['rotate_angle'])

                        y, x = prev_ratemap.shape

                        source_weights, row_prev, col_prev = _get_valid_weight_bins(prev_ratemap)
                        target_weights, row_curr, col_curr = _get_valid_weight_bins(curr_ratemap)

                        prev_height_bucket_midpoints, prev_width_bucket_midpoints, coord_buckets_prev = _get_valid_midpoints(prev_spatial_spike_train.arena_size, y, x, row_prev, col_prev)
                        curr_height_bucket_midpoints, curr_width_bucket_midpoints, coord_buckets_curr = _get_valid_midpoints(curr_spatial_spike_train.arena_size, y, x, row_curr, col_curr)
                                        
                        source_weights = source_weights / np.sum(source_weights)
                        target_weights = target_weights / np.sum(target_weights)

                        if settings['normalizePos']:
                            prev_height_bucket_midpoints = (prev_height_bucket_midpoints - np.min(prev_height_bucket_midpoints)) / (np.max(prev_height_bucket_midpoints) - np.min(prev_height_bucket_midpoints))
                            prev_width_bucket_midpoints = (prev_width_bucket_midpoints - np.min(prev_width_bucket_midpoints)) / (np.max(prev_width_bucket_midpoints) - np.min(prev_width_bucket_midpoints))
                            curr_height_bucket_midpoints = (curr_height_bucket_midpoints - np.min(curr_height_bucket_midpoints)) / (np.max(curr_height_bucket_midpoints) - np.min(curr_height_bucket_midpoints))
                            curr_width_bucket_midpoints = (curr_width_bucket_midpoints - np.min(curr_width_bucket_midpoints)) / (np.max(curr_width_bucket_midpoints) - np.min(curr_width_bucket_midpoints))
                        

                        if 'spike_density' in settings['rate_scores']:

                            if settings['normalizePos']:
                                curr_pts = scale_points(curr_pts)
                                prev_pts = scale_points(prev_pts)

                            spike_dens_wass = pot_sliced_wasserstein(prev_pts, curr_pts, n_projections=settings['n_projections'])      
                            

                            null_spike_dens_wass = animal_ref_dist[animal_id][ses_comp]['ref_spike_density']
                            null_spike_dens_wass_mean, null_spike_dens_wass_std, spike_dens_z_score, spike_dens_mod_z_score, median, mad = get_reference_dist_stats(spike_dens_wass, null_spike_dens_wass)
                            sd_quantile = wasserstein_quantile(spike_dens_wass, null_spike_dens_wass)
                        
                        if 'whole' in settings['rate_scores']:
                            wass = pot_sliced_wasserstein(coord_buckets_prev, coord_buckets_curr, source_weights, target_weights, n_projections=settings['n_projections'])

                            # line below for circular shuffling
                            # ref_wass_dist = list(map(lambda x, y: pot_sliced_wasserstein(coord_buckets_prev, coord_buckets_curr, x/np.sum(x), y/np.sum(y), n_projections=settings['n_shuffle_projections']), prev_shuffled, curr_shuffled))
                            print('search3')
                            print(animal_id, ses_comp)
                            print(animal_ref_dist[animal_id].keys())

                            ref_wass_dist = animal_ref_dist[animal_id][ses_comp]['ref_whole']

                            ref_wass_mean, ref_wass_std, z_score, mod_z_score, median, mad = get_reference_dist_stats(wass, ref_wass_dist)
                            
                            # print('doing modified z score')
                            mod_z_score, median, mad = compute_modified_zscore(wass, ref_wass_dist)

                            quantile = wasserstein_quantile(wass, ref_wass_dist)
                            
                            plower = quantile
                            phigher = 1 - quantile
                            ptwotail = (1- quantile if quantile > 0.5 else quantile)*2
                            
                        regular_dict['signature'].append([prev_path, curr_path])
                        regular_dict['spike_count'].append([len(prev_pts), len(curr_pts)])
                        regular_dict['name'].append(name)
                        regular_dict['date'].append(date)
                        regular_dict['depth'].append(depth)
                        regular_dict['unit_id'].append(cell_label)
                        regular_dict['tetrode'].append(animal.animal_id.split('tet')[-1])
                        regular_dict['session_ids'].append([prev_key, curr_key])
                        
                        if 'whole' in settings['rate_scores']:
                            regular_dict['plower'].append(plower)
                            regular_dict['phigher'].append(phigher)
                            regular_dict['ptwotail'].append(ptwotail)
                            regular_dict['quantile'].append(quantile)
                            regular_dict['whole_wass'].append(wass)
                            regular_dict['z_score'].append(z_score)
                            regular_dict['mod_z_score'].append(mod_z_score)
                            regular_dict['base_mean'].append(ref_wass_mean)
                            regular_dict['base_std'].append(ref_wass_std)
                            regular_dict['median'].append(median)
                            regular_dict['mad'].append(mad)
                        
                        if 'spike_density' in settings['rate_scores']:
                            regular_dict['sd_wass'].append(spike_dens_wass)
                            regular_dict['sd_z_score'].append(spike_dens_z_score)
                            regular_dict['sd_base_mean'].append(null_spike_dens_wass_mean)
                            regular_dict['sd_base_std'].append(null_spike_dens_wass_std)
                            regular_dict['sd_median'].append(median)
                            regular_dict['sd_mad'].append(mad)
                            regular_dict['sd_quantile'].append(sd_quantile)

                        pre_post = prev_spatial_spike_train.new_spike_times
                        curr_post = curr_spatial_spike_train.new_spike_times
                        prev_fr_rate, curr_fr_rate, fr_rate_ratio, fr_rate_change = get_rate_stats(prev_pts, pre_post, curr_pts, curr_post)
                        
                        ref_rate_ratio_dist = animal_ref_dist[animal_id][ses_comp]['ref_rate_ratio']                       
                        fr_ratio_mean, fr_ratio_std, fr_ratio_z = get_ref_ratio_stats(ref_rate_ratio_dist, fr_rate_ratio)
                        fr_ratio_quantile = wasserstein_quantile(fr_rate_ratio, ref_rate_ratio_dist)

                        ref_rate_change_dist = animal_ref_dist[animal_id][ses_comp]['ref_rate_change']
                        fr_change_mean, fr_change_std, fr_change_z = get_ref_change_stats(ref_rate_change_dist, fr_rate_change)
                        fr_change_quantile = wasserstein_quantile(fr_rate_change, ref_rate_change_dist)

                        regular_dict['fr'].append([prev_fr_rate, curr_fr_rate])
                        regular_dict['fr_ratio'].append(fr_rate_ratio)
                        regular_dict['fr_change'].append(fr_rate_change)
                        regular_dict['fr_ratio_z'].append(fr_ratio_z)
                        regular_dict['fr_change_z'].append(fr_change_z)
                        regular_dict['fr_ratio_q'].append(fr_ratio_quantile)
                        regular_dict['fr_change_q'].append(fr_change_quantile)
                        regular_dict['fr_ratio_mean'].append(fr_ratio_mean)
                        regular_dict['fr_ratio_std'].append(fr_ratio_std)
                        regular_dict['fr_change_mean'].append(fr_change_mean)
                        regular_dict['fr_change_std'].append(fr_change_std)
                        regular_dict['n_repeats'].append(len(ref_rate_ratio_dist))
                        regular_dict['arena_size'].append([prev_spatial_spike_train.arena_size, curr_spatial_spike_train.arena_size])
                        regular_dict['cylinder'].append(cylinder)

                        assert prev.shape == curr.shape
                        regular_dict['ratemap_dims'].append(curr.shape)

                        if settings['downsample']:
                            regular_dict['downsample_factor'].append(settings['downsample_factor'])
                        else:
                            regular_dict['downsample_factor'].append(1)

                        if settings['plotRegular']:
                            plot_regular_remapping(prev, curr, regular_dict, data_dir)

                        if settings['plotShuffled']:
                            prev_shuffled_sample = None
                            curr_shuffled_sample = None
                            prev_shuffled = animal_ref_dist[animal_id][ses_comp]['ref_weights'][0]
                            curr_shuffled = animal_ref_dist[animal_id][ses_comp]['ref_weights'][1]
                            plot_shuffled_regular_remapping(prev_shuffled, curr_shuffled, ref_wass_dist, prev_shuffled_sample, curr_shuffled_sample, regular_dict, data_dir)

                        if settings['runFields']:

                            image_prev, n_labels_prev, source_labels, source_centroids, field_sizes_prev = blobs_dict[prev_id]

                            image_curr, n_labels_curr, target_labels, target_centroids, field_sizes_curr = blobs_dict[curr_id]

                            if len(np.unique(target_labels)) > 1 and len(np.unique(source_labels)) > 1:

                                """
                                cumulative_dict has field/centroid/binary wass, cumulative = all fields used
                                'field_wass' is EMD on ALL fields for norm/unnorm RATE remapping
                                'centroid_wass' is EMD on ALL field centroids for norm/unnorm LOCATION remapping (i.e. field centre points averaged + EMD calculated)
                                'binary_wass' is EMD on ALL fields (binary) for norm/unnorm LOCATION remapping (i.e. unweighted such that each pt contributes equally within field)

                                permute_dict has field/centroid/binary wass, permute = all combinations of single fields for given session pair
                                'field_wass' is EMD on SINGLE fields for norm/unnorm RATE remapping
                                'centroid_wass' is EMD on SINGLE field centroids for norm/unnorm LOCATION remapping (i.e. EMD calculated directly between diff centroid pairs across sessions)
                                'binary_wass' is EMD on SINGLE fields (binary) for norm/unnorm LOCATION remapping (i.e. unweighted such that each pt contributes equally within field)
                                """
                                permute_dict, cumulative_dict = compute_centroid_remapping(target_labels, source_labels, curr_spatial_spike_train, prev_spatial_spike_train, target_centroids, source_centroids, settings)

                                y, x = curr.shape
                                h, w = rate_map_obj.arena_size
                                bin_area = h/y * w/x
                                field_count = [len(np.unique(source_labels)) - 1,len(np.unique(target_labels)) - 1]

                                for centroid_score in settings['centroid_scores']:

                                    score_key = centroid_score + '_wass'

                                    centroid_dict['score'].append(centroid_score)
                                    centroid_dict['name'].append(name)
                                    centroid_dict['date'].append(date)
                                    centroid_dict['depth'].append(depth)
                                    centroid_dict['bin_area'].append(bin_area[0])
                                    centroid_dict['field_count'].append(field_count)
                                    centroid_dict['unit_id'].append(cell_label)
                                    centroid_dict['tetrode'].append(animal.animal_id.split('tet')[-1])
                                    centroid_dict['session_ids'].append([prev_key, curr_key])
                                    centroid_dict['signature'].append([prev_path, curr_path])
                                    centroid_dict['arena_size'].append(curr_spatial_spike_train.arena_size)
                                    centroid_dict['cylinder'].append(cylinder)
                                    centroid_dict['ratemap_dims'].append(curr.shape)
                                    centroid_dict['cumulative_wass'].append(cumulative_dict[score_key])

                                    copy_labels, _ = _copy_labels(source_labels)

                                    cumulative_source_coverage, cumulative_source_area, cumulative_source_rate = compute_cumulative_blob_stats(copy_labels, prev)                  
                        
                                    copy_labels, _ = _copy_labels(target_labels)

                                    cumulative_target_coverage, cumulative_target_area, cumulative_target_rate = compute_cumulative_blob_stats(copy_labels, curr)
      
                                    centroid_dict['cumulative_coverage'].append([cumulative_source_coverage, cumulative_target_coverage])
                                    centroid_dict['cumulative_area'].append([cumulative_source_area, cumulative_target_area])
                                    centroid_dict['cumulative_rate'].append([cumulative_source_rate, cumulative_target_rate])

                                    wass_to_add = permute_dict[score_key]

                                    centroid_dict = _fill_centroid_dict(centroid_dict, max_centroid_count, wass_to_add, permute_dict['pairs'], permute_dict['coords'], prev, source_labels, field_sizes_prev, curr, target_labels, field_sizes_curr)

                                if settings['plotFields']:
                                    if cylinder:
                                        target_labels = flat_disk_mask(target_labels)
                                        source_labels = flat_disk_mask(source_labels)
                                    plot_fields_remapping(source_labels, target_labels, prev_spatial_spike_train, curr_spatial_spike_train, source_centroids, target_centroids, centroid_dict, data_dir, settings, cylinder=cylinder)

                    if prev is not None and settings['runTemporal']:

                        if prev_spike_times is None:
                            prev_spike_times = prev_spatial_spike_train.spike_times

                        curr_spike_times = curr_spatial_spike_train.spike_times

                        num_shuffles = settings['n_temporal_shuffles']
             
                        if settings['normalizeTime']:
                            prev_spike_times = (prev_spike_times - np.min(prev_spike_times)) / (np.max(prev_spike_times) - np.min(prev_spike_times))
                            curr_spike_times = (curr_spike_times - np.min(curr_spike_times)) / (np.max(curr_spike_times) - np.min(curr_spike_times))

                        observed_emd = compute_temporal_emd(prev_spike_times, curr_spike_times, settings['temporal_bin_size'], settings['end_time'])
                        # print("OBSERVED EMD " + str(observed_emd))

                        ref_emd_dist = animal_ref_dist[animal_id][ses_comp]['ref_temporal']
                        ref_emd_mean, ref_emd_std, z_score, mod_z_score, median, mad = get_reference_dist_stats(observed_emd, ref_emd_dist)
                        quantile = wasserstein_quantile(observed_emd, ref_emd_dist)

                        prev_duration = prev_spatial_spike_train.session_metadata.session_object.get_spike_data()['spike_cluster'].duration
                        curr_duration = curr_spatial_spike_train.session_metadata.session_object.get_spike_data()['spike_cluster'].duration
                        prev_fr_rate, curr_fr_rate, fr_rate_ratio, fr_rate_change = get_rate_stats(prev_spike_times, None, curr_spike_times, None, prev_duration=prev_duration, curr_duration=curr_duration)

                        ref_rate_ratio_dist = animal_ref_dist[animal_id][ses_comp]['ref_rate_ratio']
                        fr_ratio_mean, fr_ratio_std, fr_ratio_z = get_ref_ratio_stats(ref_rate_ratio_dist, fr_rate_ratio)
                        fr_ratio_quantile = wasserstein_quantile(fr_rate_ratio, ref_rate_ratio_dist)

                        ref_rate_change_dist = animal_ref_dist[animal_id][ses_comp]['ref_rate_change']
                        fr_change_mean, fr_change_std, fr_change_z = get_ref_change_stats(ref_rate_change_dist, fr_rate_change)
                        fr_change_quantile = wasserstein_quantile(fr_rate_change, ref_rate_change_dist)
                                                
                        temporal_dict['signature'].append([prev_path, curr_path])
                        temporal_dict['spike_count'].append([len(prev_spike_times), len(curr_spike_times)])
                        temporal_dict['depth'].append(depth)
                        temporal_dict['name'].append(name)
                        temporal_dict['date'].append(date)
                        temporal_dict['tetrode'].append(animal.animal_id.split('tet')[-1])
                        temporal_dict['unit_id'].append(cell_label)
                        temporal_dict['session_ids'].append([prev_key, curr_key])
                        temporal_dict['emd'].append(observed_emd)
                        temporal_dict['emd_z'].append(z_score)
                        temporal_dict['emd_quantile'].append(quantile)
                        temporal_dict['emd_mean'].append(ref_emd_mean)
                        temporal_dict['emd_std'].append(ref_emd_std)
                        temporal_dict['fr'].append([prev_fr_rate, curr_fr_rate])
                        temporal_dict['fr_ratio'].append(fr_rate_ratio)
                        temporal_dict['fr_change'].append(fr_rate_change)
                        temporal_dict['fr_ratio_z'].append(fr_ratio_z)
                        temporal_dict['fr_ratio_q'].append(fr_ratio_quantile)
                        temporal_dict['fr_change_z'].append(fr_change_z)
                        temporal_dict['fr_change_q'].append(fr_change_quantile)
                        temporal_dict['fr_ratio_mean'].append(fr_ratio_mean)
                        temporal_dict['fr_ratio_std'].append(fr_ratio_std)
                        temporal_dict['fr_change_mean'].append(fr_change_mean)
                        temporal_dict['fr_change_std'].append(fr_change_std)
                        temporal_dict['n_repeats'].append(len(ref_rate_ratio_dist))
                        temporal_dict['arena_size'].append([prev_spatial_spike_train.arena_size, curr_spatial_spike_train.arena_size])
                        temporal_dict['temporal_bin_size'].append(settings['temporal_bin_size'])

                    
                    prev = curr
                    prev_spike_times = curr_spike_times
                    prev_ratemap = curr_ratemap
                    prev_spatial_spike_train = curr_spatial_spike_train
                    prev_id = curr_id
                    prev_cell = curr_cell
                    prev_key = curr_key
                    prev_path = curr_path
                    prev_shuffled = curr_shuffled
                                
            # If there are context specific or otherwise specific groups to compare, can set those ids in settings
            # Will perform session to session remapping segregated by groups definned in settings
            if settings['runUniqueGroups'] or settings['runUniqueOnlyTemporal']:
                # for category group (group of session ids)
                for categ in session_comp_categories:
                    categories = session_comp_categories[categ]
                    prev_key = None
                    prev = None 
                    curr_spike_times = None
                    prev_spike_times = None
                    prev_id = None
                    prev_spatial = None
                    prev_cell = None
                    prev_shuffled = None

                    # For session in that category
                    for ses in categories:
                        seskey = 'session_' + str(ses)
                        ses = animal.sessions[seskey]
                        path = ses.session_metadata.file_paths['tet']
                        fname = path.split('/')[-1].split('.')[0]
                        curr_id = str(animal.animal_id) + '_' + str(seskey) + '_' + str(cell.cluster.cluster_label)
                        
                        if settings['disk_arena']: 
                            cylinder = True
                        else:
                            cylinder, _ = check_disk_arena(fname)

                        ensemble = ses.get_cell_data()['cell_ensemble']

                        if cell_label in ensemble.get_cell_label_dict():
                            cell = ensemble.get_cell_by_id(cell_label)

                            # spatial_spike_train = cell.stats_dict['cell_stats']['spatial_spike_train']
                            spatial_spike_train = cell.stats_dict['spatial_spike_train'] 

                            rate_map_obj = spatial_spike_train.get_map('rate')

                            if settings['normalizeRate']:
                                rate_map, _ = rate_map_obj.get_rate_map(new_size = settings['ratemap_dims'][0])
                            else:
                                _, rate_map = rate_map_obj.get_rate_map(new_size = settings['ratemap_dims'][0])

                            assert rate_map.shape == (settings['ratemap_dims'][0], settings['ratemap_dims'][1]), 'Wrong ratemap shape {} vs settings shape {}'.format(rate_map.shape, (settings['ratemap_dims'][0], settings['ratemap_dims'][1]))
                            
                            # Disk mask ratemap
                            if cylinder:
                                curr = flat_disk_mask(rate_map)
                                if settings['downsample']:
                                    curr_ratemap = _downsample(rate_map, settings['downsample_factor'])
                                    curr_ratemap = flat_disk_mask(curr_ratemap)
                                else:
                                    curr_ratemap = curr
                                row, col = np.where(~np.isnan(curr_ratemap))
                                disk_ids = np.array([row, col]).T
                            else:
                                curr = rate_map
                                if settings['downsample']:
                                    curr_ratemap = _downsample(rate_map, settings['downsample_factor']) 
                                else:
                                    curr_ratemap = curr
                                disk_ids = None

                            curr_spatial = spatial_spike_train
                            curr_cell = cell
                            curr_key = seskey
                            curr_ratemap = rate_map
                            curr_path = ses.session_metadata.file_paths['tet'].split('/')[-1].split('.')[0]

                            if prev is not None:
                                ses_1 = prev_key.split('_')[1]
                                ses_2 = seskey.split('_')[1]
                                # ses_comp = '_'.join([ses_1, ses_2])
                                ses_comp = str(ses_1) + '_' + str(ses_2)
                                
                                if settings['runUniqueGroups']:
                                    # if prev_shuffled is None:  
                                    #     prev_shuffled = shuffled_ratemap_dict[prev_id]
                                    # curr_shuffled = shuffled_ratemap_dict[curr_id]
                                    # get x and y pts for spikes in pair of sessions (prev and curr) for a given comparison

                                    # prev_spike_pos_x, prev_spike_pos_y, prev_spike_pos_t = prev_spatial.get_spike_positions()
                                    prev_spike_pos_x, prev_spike_pos_y, prev_spike_pos_t = prev_spatial.spike_x, prev_spatial.spike_y, prev_spatial.new_spike_times
                                    prev_pts = np.array([prev_spike_pos_x, prev_spike_pos_y]).T

                                    # curr_spike_pos_x, curr_spike_pos_y, curr_spike_pos_t = curr_spatial.get_spike_positions()
                                    curr_spike_pos_x, curr_spike_pos_y, curr_spike_pos_t = curr_spatial.spike_x, curr_spatial.spike_y, curr_spatial.new_spike_times
                                    curr_pts = np.array([curr_spike_pos_x, curr_spike_pos_y]).T

                                    if settings['rotate_evening']:
                                        if 'evening' or 'rotated' in curr_path.lower():
                                            prev_ratemap = ndimage.rotate(prev_ratemap, settings['rotate_angle'])
                                            curr_ratemap = ndimage.rotate(curr_ratemap, settings['rotate_angle'])

                                            if settings['rotate_angle'] == 90:
                                                prev_pts = np.array([prev_pts[:,1], -prev_pts[:,0]]).T
                                                curr_pts = np.array([curr_pts[:,1], -curr_pts[:,0]]).T
                                                print('rotating 90 degrees for {}'.format(curr_path))
                                            else:
                                                raise ValueError('Rotation angle not supported {}'.format(settings['rotate_angle']))

                                    y, x = prev_ratemap.shape
                                    # find indices of not nan 
                                    row_prev, col_prev = np.where(~np.isnan(prev_ratemap))
                                    row_curr, col_curr = np.where(~np.isnan(curr_ratemap))

                                    # # for first map
                                    # if prev_shuffled is None:  
                                    #     prev_shuffled_samples = list(map(lambda x: _single_shuffled_sample(prev_spatial_spike_train, settings), np.arange(settings['n_repeats'])))
                                    #     if cylinder:
                                    #         prev_shuffled_samples = list(map(lambda x: flat_disk_mask(x), prev_shuffled_samples))
                                    #     prev_shuffled = list(map(lambda sample: np.array(list(map(lambda x, y: sample[x,y], row_prev, col_prev))), prev_shuffled_samples))

                                    # curr_shuffled_samples = list(map(lambda x: _single_shuffled_sample(curr_spatial_spike_train, settings), np.arange(settings['n_repeats'])))
                                    # if cylinder:
                                    #     curr_shuffled_samples = list(map(lambda x: flat_disk_mask(x), curr_shuffled_samples))   
                                    # curr_shuffled = list(map(lambda sample: np.array(list(map(lambda x, y: sample[x,y], row_curr, col_curr))), curr_shuffled_samples))

                                    # assert row_prev.all() == row_curr.all() and col_prev.all() == col_curr.all(), 'Nans in different places'

                                    prev_height_bucket_midpoints, prev_width_bucket_midpoints = _get_ratemap_bucket_midpoints(prev_spatial_spike_train.arena_size, y, x)
                                    curr_height_bucket_midpoints, curr_width_bucket_midpoints = _get_ratemap_bucket_midpoints(curr_spatial_spike_train.arena_size, y, x)
                                    
                                    prev_height_bucket_midpoints = prev_height_bucket_midpoints[row_prev]
                                    prev_width_bucket_midpoints = prev_width_bucket_midpoints[col_prev]
                                    curr_height_bucket_midpoints = curr_height_bucket_midpoints[row_curr]
                                    curr_width_bucket_midpoints = curr_width_bucket_midpoints[col_curr]
                                    source_weights = np.array(list(map(lambda x, y: prev_ratemap[x,y], row_prev, col_prev)))
                                    target_weights = np.array(list(map(lambda x, y: curr_ratemap[x,y], row_curr, col_curr)))
                                    source_weights = source_weights / np.sum(source_weights)
                                    target_weights = target_weights / np.sum(target_weights)
                                    coord_buckets_curr = np.array(list(map(lambda x, y: [curr_height_bucket_midpoints[x],curr_width_bucket_midpoints[y]], row_curr, col_curr)))
                                    coord_buckets_prev = np.array(list(map(lambda x, y: [prev_height_bucket_midpoints[x],prev_width_bucket_midpoints[y]], row_prev, col_prev)))

                                    if settings['normalizePos']:
                                        curr_pts = scale_points(curr_pts)
                                        prev_pts = scale_points(prev_pts)
                                    spike_dens_wass = pot_sliced_wasserstein(prev_pts, curr_pts, n_projections=settings['n_projections'])
                                        # elif rate_score == 'whole':
                                            # This is EMD on whole map for normalized/unnormalized rate remapping
                                    wass = pot_sliced_wasserstein(coord_buckets_prev, coord_buckets_curr, source_weights, target_weights, n_projections=settings['n_projections'])
                                    # ref_wass_dist = list(map(lambda x, y: pot_sliced_wasserstein(coord_buckets_prev, coord_buckets_curr, x/np.sum(x), y/np.sum(y), n_projections=settings['n_shuffle_projections']), prev_shuffled, curr_shuffled))
                                    ref_wass_dist = animal_ref_dist[animal_id][ses_comp]['ref_whole']
                                    ref_wass_mean = np.mean(ref_wass_dist)
                                    ref_wass_std = np.std(ref_wass_dist)
                                    z_score = (wass - ref_wass_mean) / (ref_wass_std)
                                    mod_z_score, median, mad = compute_modified_zscore(wass, ref_wass_dist)
                                    ref_spike_dens_wass_dist = animal_ref_dist[animal_id][ses_comp]['ref_spike_density']
                                    ref_spike_dens_wass_mean = np.mean(ref_spike_dens_wass_dist)
                                    ref_spike_dens_wass_std = np.std(ref_spike_dens_wass_dist)
                                    spike_dens_z_score = (spike_dens_wass - ref_spike_dens_wass_mean) / (ref_spike_dens_wass_std)

                                    # assert len(ref_wass_dist) == settings['n_repeats'], 'n_repeats does not match length of ref_wass_dist'
                                    
                                    quantile = wasserstein_quantile(wass, ref_wass_dist)
                                    plower = quantile
                                    phigher = 1 - quantile
                                    ptwotail = (1- quantile if quantile > 0.5 else quantile)*2

                                    sd_quantile = wasserstein_quantile(spike_dens_wass, ref_spike_dens_wass_dist)
                                    _, sd_median, sd_mad = compute_modified_zscore(spike_dens_wass, ref_spike_dens_wass_dist)
                            
                                    # pvalue = stats.t.cdf(z_score, len(ref_wass_dist)-1)
                                    # mod_pvalue = stats.t.cdf(mod_z_score, len(ref_wass_dist)-1)
                                    # pvalue for wass, 2 sided
                                    # pvalue = 2 * stats.t.cdf(-np.abs(t_score), len(ref_wass_dist)-1)
                                    # pvalue = 2 * stats.norm.cdf(-np.abs(t_score))

                                    # prev_shapiro_coeff, prev_shapiro_pval = stats.shapiro(prev_shuffled)
                                    # curr_shapiro_coeff, curr_shapiro_pval = stats.shapiro(curr_shuffled)

                                    context_dict[categ]['signature'].append([prev_path, curr_path])
                                    context_dict[categ]['spike_count'].append([len(prev_pts), len(curr_pts)])
                                    context_dict[categ]['name'].append(name)
                                    context_dict[categ]['date'].append(date)
                                    context_dict[categ]['depth'].append(depth)
                                    context_dict[categ]['unit_id'].append(cell_label)
                                    context_dict[categ]['tetrode'].append(animal.animal_id.split('tet')[-1])
                                    context_dict[categ]['session_ids'].append([prev_key, curr_key])
                                    context_dict[categ]['whole_wass'].append(wass)
                                    # context_dict[categ]['sd_wass'].append(spike_dens_wass)
                                    context_dict[categ]['z_score'].append(z_score)
                                    context_dict[categ]['mod_z_score'].append(mod_z_score)
                                    context_dict[categ]['plower'].append(plower)
                                    context_dict[categ]['phigher'].append(phigher)
                                    context_dict[categ]['ptwotail'].append(ptwotail)
                                    context_dict[categ]['quantile'].append(quantile)
                                    # context_dict[categ]['p_value'].append(pvalue)
                                    # context_dict[categ]['mod_p_value'].append(mod_pvalue)
                                    # context_dict[categ]['shapiro_pval'].append([prev_shapiro_pval, curr_shapiro_pval])
                                    # context_dict[categ]['shapiro_coeff'].append([prev_shapiro_coeff, curr_shapiro_coeff])
                                    context_dict[categ]['base_mean'].append(ref_wass_mean)
                                    context_dict[categ]['base_std'].append(ref_wass_std)
                                    context_dict[categ]['median'].append(median)
                                    context_dict[categ]['mad'].append(mad)
                                    context_dict[categ]['sd_wass'].append(spike_dens_wass)
                                    context_dict[categ]['sd_z_score'].append(spike_dens_z_score)
                                    context_dict[categ]['sd_quantile'].append(sd_quantile)
                                    context_dict[categ]['sd_base_mean'].append(ref_spike_dens_wass_mean)
                                    context_dict[categ]['sd_base_std'].append(ref_spike_dens_wass_std)
                                    context_dict[categ]['sd_median'].append(sd_median)
                                    context_dict[categ]['sd_mad'].append(sd_mad)

                                    curr_fr_rate = len(curr_pts) / (curr_spike_pos_t[-1] - curr_spike_pos_t[0])
                                    prev_fr_rate = len(prev_pts) / (prev_spike_pos_t[-1] - prev_spike_pos_t[0])
                                    fr_rate_ratio = curr_fr_rate / prev_fr_rate
                                    fr_rate_change = curr_fr_rate - prev_fr_rate

                                    ref_rate_ratio_dist = animal_ref_dist[animal_id][ses_comp]['ref_rate_ratio']
                                    fr_ratio_mean = np.mean(ref_rate_ratio_dist)
                                    fr_ratio_std = np.std(ref_rate_ratio_dist)
                                    fr_ratio_z = (fr_rate_ratio - fr_ratio_mean) / (fr_ratio_std)
                                    fr_ratio_quantile = wasserstein_quantile(fr_rate_ratio, ref_rate_ratio_dist)

                                    ref_rate_change_dist = animal_ref_dist[animal_id][ses_comp]['ref_rate_change']
                                    fr_change_mean = np.mean(ref_rate_change_dist)
                                    fr_change_std = np.std(ref_rate_change_dist)
                                    fr_change_z = (fr_rate_change - fr_change_mean) / (fr_change_std)
                                    fr_change_quantile = wasserstein_quantile(fr_rate_change, ref_rate_change_dist)

                                    context_dict[categ]['fr'].append([prev_fr_rate, curr_fr_rate])
                                    context_dict[categ]['fr_ratio'].append(fr_rate_ratio)
                                    context_dict[categ]['fr_change'].append(fr_rate_change)
                                    context_dict[categ]['fr_ratio_z'].append(fr_ratio_z)
                                    context_dict[categ]['fr_change_z'].append(fr_change_z)
                                    context_dict[categ]['fr_ratio_q'].append(fr_ratio_quantile)
                                    context_dict[categ]['fr_change_q'].append(fr_change_quantile)
                                    # mean and std
                                    context_dict[categ]['fr_ratio_mean'].append(fr_ratio_mean)
                                    context_dict[categ]['fr_ratio_std'].append(fr_ratio_std)
                                    context_dict[categ]['fr_change_mean'].append(fr_change_mean)
                                    context_dict[categ]['fr_change_std'].append(fr_change_std)

                                    context_dict[categ]['n_repeats'].append(len(ref_rate_ratio_dist))
                                    context_dict[categ]['arena_size'].append([prev_spatial_spike_train.arena_size, curr_spatial_spike_train.arena_size])
                                    context_dict[categ]['cylinder'].append(cylinder)
                                    assert prev.shape == curr.shape
                                    context_dict[categ]['ratemap_dims'].append(curr.shape)
                                    if settings['downsample']:
                                        context_dict[categ]['downsample_factor'].append(settings['downsample_factor'])
                                    else:
                                        context_dict[categ]['downsample_factor'].append(1)
                                
                                if settings['runUniqueOnlyTemporal']:
                                    if prev_spike_times is None:
                                        prev_spike_times = prev_spatial.spike_times

                                    curr_spike_times = curr_spatial.spike_times

                                    num_shuffles = settings['n_temporal_shuffles']

                
                                    # print('computing shuffled temporal emd')
                                    observed_emd = compute_temporal_emd(prev_spike_times, curr_spike_times, settings['temporal_bin_size'], settings['end_time'])
                                    ref_emd_dist = animal_ref_dist[animal_id][ses_comp]['ref_temporal']
                                    emd_mean = np.mean(ref_emd_dist)
                                    emd_std = np.std(ref_emd_dist)
                                    emd_z = (observed_emd - emd_mean) / (emd_std)
                                    emd_quantile = wasserstein_quantile(observed_emd, ref_emd_dist)

                                    # print('doing modified z score')
                                    prev_duration = prev_spatial.session_metadata.session_object.get_spike_data()['spike_cluster'].duration
                                    curr_duration = curr_spatial.session_metadata.session_object.get_spike_data()['spike_cluster'].duration

                                    curr_fr_rate = len(curr_spike_times) / prev_duration
                                    prev_fr_rate = len(prev_spike_times) / curr_duration
                                    fr_rate_ratio = curr_fr_rate / prev_fr_rate
                                    fr_rate_change = curr_fr_rate - prev_fr_rate

                                    ref_rate_ratio_dist = animal_ref_dist[animal_id][ses_comp]['ref_rate_ratio']
                                    fr_ratio_mean = np.mean(ref_rate_ratio_dist)
                                    fr_ratio_std = np.std(ref_rate_ratio_dist)
                                    fr_ratio_z = (fr_rate_ratio - fr_ratio_mean) / (fr_ratio_std)
                                    fr_ratio_quantile = wasserstein_quantile(fr_rate_ratio, ref_rate_ratio_dist)

                                    ref_rate_change_dist = animal_ref_dist[animal_id][ses_comp]['ref_rate_change']
                                    fr_change_mean = np.mean(ref_rate_change_dist)
                                    fr_change_std = np.std(ref_rate_change_dist)
                                    fr_change_z = (fr_rate_change - fr_change_mean) / (fr_change_std)
                                    fr_change_quantile = wasserstein_quantile(fr_rate_change, ref_rate_change_dist)

                                    context_temporal_dict[categ]['fr'].append([prev_fr_rate, curr_fr_rate])
                                    context_temporal_dict[categ]['fr_ratio'].append(fr_rate_ratio)
                                    context_temporal_dict[categ]['fr_change'].append(fr_rate_change)
                                    context_temporal_dict[categ]['fr_ratio_z'].append(fr_ratio_z)
                                    context_temporal_dict[categ]['fr_change_z'].append(fr_change_z)
                                    context_temporal_dict[categ]['fr_ratio_q'].append(fr_ratio_quantile)
                                    context_temporal_dict[categ]['fr_change_q'].append(fr_change_quantile)
                                    # mean and std
                                    context_temporal_dict[categ]['fr_ratio_mean'].append(fr_ratio_mean)
                                    context_temporal_dict[categ]['fr_ratio_std'].append(fr_ratio_std)
                                    context_temporal_dict[categ]['fr_change_mean'].append(fr_change_mean)
                                    context_temporal_dict[categ]['fr_change_std'].append(fr_change_std)
                                    context_temporal_dict[categ]['n_repeats'].append(len(ref_rate_ratio_dist))

                                    context_temporal_dict[categ]['signature'].append([prev_path, curr_path])
                                    context_temporal_dict[categ]['depth'].append(depth)
                                    context_temporal_dict[categ]['name'].append(name)
                                    context_temporal_dict[categ]['date'].append(date)
                                    context_temporal_dict[categ]['tetrode'].append(animal.animal_id.split('tet')[-1])
                                    context_temporal_dict[categ]['unit_id'].append(cell_label)
                                    context_temporal_dict[categ]['session_ids'].append([prev_key, curr_key])
                                    context_temporal_dict[categ]['emd'].append(observed_emd)
                                    context_temporal_dict[categ]['emd_z'].append(emd_z)
                                    context_temporal_dict[categ]['emd_quantile'].append(emd_quantile)
                                    context_temporal_dict[categ]['spike_count'].append([len(prev_spike_times), len(curr_spike_times)])
                                    # mean and std 
                                    context_temporal_dict[categ]['emd_mean'].append(emd_mean)
                                    context_temporal_dict[categ]['emd_std'].append(emd_std)
                                    context_temporal_dict[categ]['arena_size'].append([prev_spatial.arena_size, curr_spatial.arena_size])


                            prev = curr
                            prev_spike_times = curr_spike_times 
                            prev_cell = curr_cell
                            prev_spatial = curr_spatial
                            prev_key = curr_key
                            prev_path = curr_path
                            prev_shuffled = curr_shuffled
                            prev_ratemap = curr_ratemap

                            

            if settings['plotMatchedWaveforms'] and settings['runRegular']:     
                plot_matched_sesssion_waveforms(cell_session_appearances, settings, regular_dict, data_dir)

            to_save = {'regular': regular_dict, 'object': obj_dict, 'centroid': centroid_dict, 
                       'context': context_dict, 'temporal': temporal_dict, 'context_temporal': context_temporal_dict}
            if settings['runRegular']:
                df = pd.DataFrame(to_save['regular'])
                if isStart:
                    path = data_dir + '/remapping_output/regular_remapping.xlsx'
                    if not os.path.isfile(path):
                        regular_path_to_use = path
                    else:
                        counter = 2
                        while os.path.isfile(data_dir + '/remapping_output/regular_remapping_' + str(counter) + '.xlsx'):
                            counter += 1
                        regular_path_to_use = data_dir + '/remapping_output/regular_remapping_' + str(counter) + '.xlsx'
                    writer = pd.ExcelWriter(regular_path_to_use, engine='openpyxl')
                    df.to_excel(writer, sheet_name='Summary')
                    writer.save()
                    writer.close()

                else:
                    book = load_workbook(regular_path_to_use)
                    writer = pd.ExcelWriter(regular_path_to_use, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Summary', header=False, startrow=writer.sheets['Summary'].max_row)
                    writer.close()
                    book.save(regular_path_to_use)

            if settings['hasObject']:
                df = pd.DataFrame(to_save['object'])
                if isStart:
                    path = data_dir + '/remapping_output/obj_remapping.xlsx'
                    if not os.path.isfile(path):
                        obj_path_to_use = path
                    else:
                        counter = 2
                        while os.path.isfile(data_dir + '/remapping_output/obj_remapping_' + str(counter) + '.xlsx'):
                            counter += 1
                        obj_path_to_use = data_dir + '/remapping_output/obj_remapping_' + str(counter) + '.xlsx'
                    writer = pd.ExcelWriter(obj_path_to_use, engine='openpyxl')
                    df.to_excel(writer, sheet_name='Summary')
                    writer.save()
                    writer.close()
                else:
                    book = load_workbook(obj_path_to_use)
                    writer = pd.ExcelWriter(obj_path_to_use, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Summary', header=False, startrow=writer.sheets['Summary'].max_row)
                    writer.close()
                    book.save(obj_path_to_use)
            
            if settings['runFields']:
                df = pd.DataFrame(to_save['centroid'])
                if isStart:
                    path = data_dir + '/remapping_output/centroid_remapping.xlsx'
                    if not os.path.isfile(path):
                        centroid_path_to_use = path
                    else:
                        counter = 2
                        while os.path.isfile(data_dir + '/remapping_output/centroid_remapping_' + str(counter) + '.xlsx'):
                            counter += 1
                        centroid_path_to_use = data_dir + '/remapping_output/centroid_remapping_' + str(counter) + '.xlsx'
                    writer = pd.ExcelWriter(centroid_path_to_use, engine='openpyxl')
                    df.to_excel(writer, sheet_name='Summary')
                    writer.save()
                    writer.close()
                else:
                    book = load_workbook(centroid_path_to_use)
                    writer = pd.ExcelWriter(centroid_path_to_use, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Summary', header=False, startrow=writer.sheets['Summary'].max_row)
                    # writer.save()
                    writer.close()
                    book.save(centroid_path_to_use)
            if settings['runUniqueGroups']:
        
                for context in to_save['context']:
                    df = pd.DataFrame(to_save['context'][context])
                    if isStart:
                        path = data_dir + '/remapping_output/context_' + context + '_remapping.xlsx'
                        if not os.path.isfile(path):
                            context_path_to_use = path
                        else:
                            counter = 2
                            while os.path.isfile(data_dir + '/remapping_output/context_' + context + '_' + str(counter) + '.xlsx'):
                                counter += 1
                            context_path_to_use = data_dir + '/remapping_output/context_' + context + '_' + str(counter) + '.xlsx'
                        writer = pd.ExcelWriter(context_path_to_use, engine='openpyxl')
                        df.to_excel(writer, sheet_name='Summary')
                        writer.save()
                        writer.close()
                        context_paths[context] = context_path_to_use
                    else:
                        context_path_to_use = context_paths[context]
                        book = load_workbook(context_path_to_use)
                        writer = pd.ExcelWriter(context_path_to_use, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df.to_excel(writer, sheet_name='Summary', header=False, startrow=writer.sheets['Summary'].max_row)
                        # writer.save()
                        writer.close()
                        book.save(context_path_to_use)
            if settings['runTemporal']:
                df = pd.DataFrame(to_save['temporal'])
                if isStart:
                    path = data_dir + '/remapping_output/temporal_remapping.xlsx'
                    if not os.path.isfile(path):
                        temporal_path_to_use = path
                    else:
                        counter = 2
                        while os.path.isfile(data_dir + '/remapping_output/temporal_remapping_' + str(counter) + '.xlsx'):
                            counter += 1
                        temporal_path_to_use = data_dir + '/remapping_output/temporal_remapping_' + str(counter) + '.xlsx'
                    writer = pd.ExcelWriter(temporal_path_to_use, engine='openpyxl')
                    df.to_excel(writer, sheet_name='Summary')
                    writer.save()
                    writer.close()
                else:
                    book = load_workbook(temporal_path_to_use)
                    writer = pd.ExcelWriter(temporal_path_to_use, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Summary', header=False, startrow=writer.sheets['Summary'].max_row)
                    # writer.save()
                    writer.close()
                    book.save(temporal_path_to_use)
            if settings['runUniqueOnlyTemporal']:
                for context in to_save['context_temporal']:
                    df = pd.DataFrame(to_save['context_temporal'][context])
                    if isStart:
                        path = data_dir + '/remapping_output/context_' + context + '_temporal_remapping.xlsx'
                        if not os.path.isfile(path):
                            context_temporal_path_to_use = path
                        else:
                            counter = 2
                            while os.path.isfile(data_dir + '/remapping_output/context_' + context + '_temporal_remapping_' + str(counter) + '.xlsx'):
                                counter += 1
                            context_temporal_path_to_use = data_dir + '/remapping_output/context_' + context + '_temporal_remapping_' + str(counter) + '.xlsx'
                        writer = pd.ExcelWriter(context_temporal_path_to_use, engine='openpyxl')
                        df.to_excel(writer, sheet_name='Summary')
                        writer.save()
                        writer.close()
                        context_temporal_paths[context] = context_temporal_path_to_use
                    else:
                        context_temporal_path_to_use = context_temporal_paths[context]
                        book = load_workbook(context_temporal_path_to_use)
                        writer = pd.ExcelWriter(context_temporal_path_to_use, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df.to_excel(writer, sheet_name='Summary', header=False, startrow=writer.sheets['Summary'].max_row)
                        writer.close()
                        book.save(context_temporal_path_to_use)
            
            isStart = False

    # For smae commented out block as at start of function
    # np.save(data_dir + '/control_ses1.npy', stack_dict['control']['session_1'])
    # np.save(data_dir + '/control_ses2.npy', stack_dict['control']['session_2'])
    # np.save(data_dir + '/control_ses3.npy', stack_dict['control']['session_3'])
    # np.save(data_dir + '/control_ses3.npy', stack_dict['control']['session_4'])

    # np.save(data_dir + '/app_ses1.npy', stack_dict['app_ki']['session_1'])
    # np.save(data_dir + '/app_ses2.npy', stack_dict['app_ki']['session_2'])
    # np.save(data_dir + '/app_ses3.npy', stack_dict['app_ki']['session_3'])
    # np.save(data_dir + '/app_ses3.npy', stack_dict['app_ki']['session_4'])

    # return {'regular': regular_dict, 'object': obj_dict, 'centroid': centroid_dict, 'context': dict}




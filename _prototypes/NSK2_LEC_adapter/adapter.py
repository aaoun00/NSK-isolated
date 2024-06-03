import os, sys
import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import time
import traceback
import re
import xarray as xr
from datetime import datetime
from openpyxl import load_workbook, Workbook


PROJECT_PATH = os.getcwd()
sys.path.append(PROJECT_PATH)
print(PROJECT_PATH)

from library.study_space import SpatialSpikeTrain2D
from x_io.rw.axona.batch_read import make_study
from scripts.batch_map.LEC_naming import LEC_naming_format, extract_name_lec
from _prototypes.cell_remapping.src.stats import get_max_matched_cell_count
from _prototypes.cell_remapping.src.utils import check_disk_arena, read_data_from_fname
from library.maps.map_utils import disk_mask

def datetime_to_microseconds(timestamp: datetime) -> float:
    """Converts a datetime object to microseconds, with microsecond precition.
    Arguments:
        timestamp {datetime} -- The timestamp to convert.
        Returns:
            int -- The timestamp in microseconds (full precision).
    """
    if timestamp is None:
        return None
    return float(timestamp.timestamp() * 1000000)


def flat_disk_mask(rate_map):
    masked_rate_map = disk_mask(rate_map)
    masked_rate_map.data[masked_rate_map.mask] = np.nan
    return  masked_rate_map.data


def main(study, settings_dict, output_path):

    custom_lec_prefix = "Study1_LEC_Andrew"

    if not os.path.isdir(output_path):
        os.mkdir(output_path)
            
    for animal in study.animals:

        tetrode = str(int(animal.animal_id.split("_tet")[-1])) # INT
        
        for seskey in animal.sessions:
            ses = animal.sessions[seskey]
            path = ses.session_metadata.file_paths["tet"]
            fname = path.split("/")[-1].split(".")[0]

            stim, depth, name, date = read_data_from_fname(fname, settings_dict["naming_type"], settings_dict["type"])

            aid = str(name) # STR
            if stim != "NO":
                stim = int(stim) # INT
            else:
                assert stim == "NO", "Stimulus type is not NO or object"
                stim = "NO" # STR
            date = int(date) # INT
            depth = int(depth) # INT

            # check if animal record csv file exists
            animal_record_path = os.path.join(output_path, "animal_record.xlsx")

            if os.path.isfile(animal_record_path):

                animal_record_wb = load_workbook(animal_record_path)
                animal_record_ws = animal_record_wb.active

                animalRecordExists = True
            else:
                animalRecordExists = False


            if settings_dict["hasObject"]:
                stimtype = "object"
            else:
                raise ValueError("Need to agree on stimytype naming if not object")

            pos_obj = ses.get_position_data()["position"]
            spike_cluster = ses.get_spike_data()["spike_cluster"]

            # max_spike_time = spike_cluster.duration # last spike time
            cluster_labels = spike_cluster.cluster_labels
            cluster_labels = np.array(cluster_labels, dtype=np.float32).squeeze() # FLOAT
            cluster_event_times = spike_cluster.event_times # FLOAT
            cluster_event_times = np.array(cluster_event_times, dtype=np.float32).squeeze()
            cluster_waveforms = spike_cluster.waveforms # dict

            # get idx of cluster labels that are not noise which is 0
            non_noise_idx = np.where(cluster_labels != 0)[0]
            cluster_labels = cluster_labels[non_noise_idx]
            cluster_event_times = cluster_event_times[non_noise_idx]
            spike_param_dict = spike_cluster.spikeparam
            timebase = float(spike_param_dict["timebase"]) # FLOAT
            bytes_per_sample = float(spike_param_dict["bytes_per_sample"]) # FLOAT
            samples_per_spike = float(spike_param_dict["samples_per_spike"]) # FLOAT
            bytes_per_timestamp = float(spike_param_dict["bytes_per_timestamp"]) # FLOAT
            duration = float(spike_param_dict["duration"]) # FLOAT
            dateandtime = spike_param_dict["datetime"] # datetime object (NOT SAVED)
            trial_time = str(dateandtime.strftime("%H:%M:%S")) # STR
            date_with_dash = str(dateandtime.strftime("%Y-%m-%d")) # STR

            trial_time_without_colon = re.sub(":", "", trial_time)
            ses_name = str(aid) + "_" + str(tetrode) + '_' + str(date) + "_" + str(trial_time_without_colon) # with tet id
            ses_pos_name = str(aid) + '_' + str(date) + "_" + str(trial_time_without_colon) # without tet id

            t, x, y, arena_height, arena_width = pos_obj.t, pos_obj.x, pos_obj.y, pos_obj.arena_height, pos_obj.arena_width
            position_pairs = np.array([x, y, t]).T.squeeze()
            arena_size = [float(arena_height), float(arena_width)]

            # make position pairs of shape (n_spikes, n_probes, n_features)
            n_pos_bins = len(x)
            n_probes = 1
            n_features = 2
            spike_positions = np.zeros((n_pos_bins, n_probes, n_features))
            spike_positions[:, 0, 0] = x.squeeze()
            spike_positions[:, 0, 1] = y.squeeze()

            session_data_ref_dict = {"schema_ref": "session", "data_name": ses_pos_name}
            animal_data_ref_dict = {"schema_ref": "animal", "data_name": aid}
            probe_data_ref_dict = {"schema_ref": "probe", "data_name": str(tetrode)}
            session_data_ref = str(session_data_ref_dict)
            animal_data_ref = str(animal_data_ref_dict)
            probe_data_ref = str(probe_data_ref_dict)

            position_features_dataarray = xr.DataArray(data=position_pairs,
                                                    dims=("sample", "xyt"),
                                                    coords={"sample": np.array(np.arange(len(position_pairs)),dtype=np.float32).squeeze(),
                                                            "xyt": np.array(np.arange(len(position_pairs[0])),dtype=np.float32).squeeze()})
                                                       
            position_features_dataarray.attrs["schema_ref"] = "animal_position"
            position_features_dataarray.attrs["data_name"] = ses_pos_name
            position_features_dataarray.attrs["has_file"] = "true"
            position_features_dataarray.attrs["data_dimensions"] = ["sample", "xyt"]
            position_features_dataarray.attrs["dimension_of_measure"] = "[space]"
            position_features_dataarray.attrs["session_data_ref"] = session_data_ref
            position_features_dataarray.attrs["animal_data_ref"] = animal_data_ref
            position_features_dataarray.attrs["probe_data_ref"] = probe_data_ref
            position_features_dataarray.attrs["recording_length"] = str(duration)
            position_features_dataarray.attrs["unit_of_measure"] = "cm"

            spike_times_dataarray = xr.DataArray(
                data=np.asarray(cluster_event_times).reshape((-1,1)),
                dims=("spike_idx", "1"),
                coords={"spike_idx": np.array(np.arange(len(cluster_event_times)),dtype=np.float32).squeeze()},
            )

            spike_times_dataarray.attrs["schema_ref"] = "spike_times"
            spike_times_dataarray.attrs["data_name"] = ses_name
            spike_times_dataarray.attrs["has_file"] = "true"
            spike_times_dataarray.attrs["data_dimensions"] = ["spike_idx", "1"]
            spike_times_dataarray.attrs["dimension_of_measure"] = "[time]"
            spike_times_dataarray.attrs["session_data_ref"] = session_data_ref
            spike_times_dataarray.attrs["animal_data_ref"] = animal_data_ref
            spike_times_dataarray.attrs["probe_data_ref"] = probe_data_ref
            spike_times_dataarray.attrs["spike_count"] = str(len(cluster_event_times))

            spike_labels_dataarray = xr.DataArray(
                data=np.asarray(cluster_labels).reshape((-1,1)),
                dims=("spike_idx", "1"),
                coords={"spike_idx": np.arange(len(cluster_labels))},
            )

            spike_labels_dataarray.attrs["schema_ref"] = "spike_labels"
            spike_labels_dataarray.attrs["data_name"] = ses_name # aid_date_depth_tetrode_LEC_study1
            spike_labels_dataarray.attrs["has_file"] = "true"
            spike_labels_dataarray.attrs["data_dimensions"] = ["spike_idx", "1"]
            spike_labels_dataarray.attrs["dimension_of_measure"] = "[nominal]"
            spike_labels_dataarray.attrs["session_data_ref"] = session_data_ref
            spike_labels_dataarray.attrs["animal_data_ref"] = animal_data_ref
            spike_labels_dataarray.attrs["probe_data_ref"] = probe_data_ref

            assert len(cluster_event_times) == len(cluster_labels), "Spike times and labels do not match"

            # convert dict to array
            ch1 = cluster_waveforms["channel_1"]
            ch2 = cluster_waveforms["channel_2"]
            ch3 = cluster_waveforms["channel_3"]
            ch4 = cluster_waveforms["channel_4"]

            cluster_waveforms = np.array([ch1, ch2, ch3, ch4], dtype=np.float32)
            cluster_waveforms = cluster_waveforms.transpose(1, 0, 2) # (time, channel, sample)
            # take only non_noise_idx
            cluster_waveforms = cluster_waveforms[non_noise_idx] # (time, channel, sample)

            spike_waveforms_dataarray = xr.DataArray(cluster_waveforms, 
                                          dims=("spike_idx", "channel", "sample"), coords={
                                                                            "spike_idx": np.array(np.arange(len(cluster_event_times)),dtype=np.float32).squeeze(),
                                                                            "channel": np.array(np.arange(4),dtype=np.float32).squeeze(),
                                                                            "sample": np.array(np.arange(len(ch1[0])),dtype=np.float32).squeeze()})
             
            assert len(cluster_waveforms) == len(cluster_event_times), "Number of spikes does not match"

            spike_waveforms_dataarray.attrs["schema_ref"] = "spike_waveforms"
            spike_waveforms_dataarray.attrs["data_name"] = ses_name
            spike_waveforms_dataarray.attrs["has_file"] = "true"
            spike_waveforms_dataarray.attrs["data_dimensions"] = ["spike_idx", "channel", "sample"]
            spike_waveforms_dataarray.attrs["unit_of_measure"] = "microvolts"
            spike_waveforms_dataarray.attrs["dimension_of_measure"] = "[charge]"
            spike_waveforms_dataarray.attrs["session_data_ref"] = session_data_ref
            spike_waveforms_dataarray.attrs["animal_data_ref"] = animal_data_ref
            spike_waveforms_dataarray.attrs["probe_data_ref"] = probe_data_ref

            duration_unit = "second"
            srecord = {"schema_ref": "session", "data_name": str(ses_pos_name),
                       "animal_id": str(aid), "session_date": str(date_with_dash), "session_time": str(trial_time), 
                       "tetrode_depth": str(depth), 
                       "stimulus_id": str(stim), "duration": str(duration), 
                       "duration_unit": str(duration_unit),"stimulus_type": str(stimtype)}

            # # session record
            session_record_path = os.path.join(output_path, "session" +"_record.xlsx")
            if os.path.isfile(session_record_path):
                session_record_wb = load_workbook(session_record_path)
                session_record_ws = session_record_wb.active
                sessionRecordExists = True
            else:
                sessionRecordExists = False

            if sessionRecordExists:
                session_duplicate_found = False
                for row in session_record_ws.iter_rows(values_only=True):
                    # make every column a string
                    row = [str(i) for i in row] 
                    if (row[0] == "session" and 
                        row[1] == str(ses_pos_name) and 
                        row[2] == str(aid) and 
                        row[3] == str(date_with_dash) and
                        row[4] == str(trial_time) and
                        row[5] == str(depth) and 
                        row[6] == str(stim) and 
                        row[7] == str(duration) and 
                        row[8] == str(duration_unit) and
                        row[9] == str(stimtype)):
                        session_duplicate_found = True
                        break

                if session_duplicate_found:
                    pass
                else:
                    session_record_ws.append([srecord["schema_ref"], srecord["data_name"], srecord["animal_id"], srecord["session_date"], srecord["session_time"], srecord["tetrode_depth"], srecord["stimulus_id"], srecord["duration"], srecord["duration_unit"], srecord["stimulus_type"]])
            else:
                session_record_wb = Workbook()
                session_record_ws = session_record_wb.active
                # add column headers
                session_record_ws.append(["schema_ref", "data_name", "animal_id", "session_date", "session_time", "tetrode_depth", "stimulus_id", "duration", "duration_unit", "stimulus_type"])
                # add data
                session_record_ws.append([srecord["schema_ref"], srecord["data_name"], srecord["animal_id"], srecord["session_date"], srecord["session_time"], srecord["tetrode_depth"], srecord["stimulus_id"], srecord["duration"], srecord["duration_unit"], srecord["stimulus_type"]])


            if "ANT" in aid:
                strain = "EC-APP/Tau"
                genotype = "Homozygous"
                species = "mouse"
                age = "unknown"
                age_lower_bound = "8"
                age_upper_bound = "10"
                age_unit = "months"
            elif "NON" in aid:
                genotype = "Hemizygous"
                strain = "Neuropsin-tta"
                species = "mouse"
                age = "unknown"
                age_lower_bound = "8"
                age_upper_bound = "10"
                age_unit = "months"
            elif "B6" in aid:
                strain = "C57BL/6J"
                genotype = "Hemizygous"
                species = "mouse"
                age = "unknown"
                age_lower_bound = "8"
                age_upper_bound = "10"
                age_unit = "months"

            arecord = {"schema_ref": "animal", "data_name": str(aid), "genotype": genotype, "animal_strain": strain, 
                       "animal_species": species, "age": age, "age_unit": str(age_unit), "age_lower_bound": str(age_lower_bound), 
                       "age_upper_bound": age_upper_bound}

            if animalRecordExists:
                animal_duplicate_found = False
                for row in animal_record_ws.iter_rows(values_only=True):
                    if (row[0] == "animal" and 
                        row[1] == str(aid) and 
                        row[2] == genotype and 
                        row[3] == strain and 
                        row[4] == species and 
                        row[5] == age and 
                        row[6] == age_unit and 
                        row[7] == age_lower_bound and 
                        row[8] == age_upper_bound):
                        animal_duplicate_found = True
                        break

                if animal_duplicate_found:
                    pass
                else:
                    # add values to dataframe
                    animal_record_ws.append([arecord["schema_ref"], arecord["data_name"], arecord["genotype"], arecord["animal_strain"], arecord["animal_species"], arecord["age"], arecord["age_unit"], arecord["age_lower_bound"], arecord["age_upper_bound"]])
            else:
                # define workbook and worksheet
                animal_record_wb = Workbook()
                animal_record_ws = animal_record_wb.active
                # add column headers
                animal_record_ws.append(["schema_ref", "data_name", "genotype", "animal_strain", "animal_species", "age", "age_unit", "age_lower_bound", "age_upper_bound"])
                # add data
                animal_record_ws.append([arecord["schema_ref"], arecord["data_name"], arecord["genotype"], arecord["animal_strain"], arecord["animal_species"], arecord["age"], arecord["age_unit"], arecord["age_lower_bound"], arecord["age_upper_bound"]])

            # save as netcdf
            pos_obj_path = os.path.join(output_path, ses_pos_name + "_position.nc")
            spike_labels_path = os.path.join(output_path, ses_name +"_cluster_labels.nc")
            spike_waveforms_path = os.path.join(output_path, ses_name +"_waveforms.nc")
            spike_times_path = os.path.join(output_path, ses_name +"_spike_times.nc")

            spike_waveforms_dataarray.to_netcdf(spike_waveforms_path, engine="netcdf4")
            position_features_dataarray = position_features_dataarray.to_dataset(name="spike_features")
            position_features_dataarray.to_netcdf(pos_obj_path, engine="netcdf4")
            spike_labels_dataarray.to_netcdf(spike_labels_path, engine="netcdf4")
            spike_times_dataarray.to_netcdf(spike_times_path, engine="netcdf4")

            # save workbooks
            session_record_wb.save(session_record_path)
            animal_record_wb.save(animal_record_path)

            # close workbooks
            session_record_wb.close()
            animal_record_wb.close()

            print("Saved " + "session" +" to " + output_path)

if __name__ == "__main__":

    STUDY_SETTINGS = {

        "ppm": 485,  # EDIT HERE

        "smoothing_factor": 3, # EDIT HERE

        "useMatchedCut": True,  # EDIT HERE
    }

    # Switch devices to True/False based on what is used in the acquisition (to be extended for more devices in future)
    device_settings = {"axona_led_tracker": True, "implant": True} 
    # Make sure implant metadata is correct, change if not, AT THE MINIMUM leave implant_type: tetrode
    implant_settings = {"implant_type": "tetrode", "implant_geometry": "square", "wire_length": 25, "wire_length_units": "um", "implant_units": "uV"}
    # WE ASSUME DEVICE AND IMPLANT SETTINGS ARE CONSISTENCE ACROSS SESSIONS
    # Set channel count + add device/implant settings
    SESSION_SETTINGS = {
        "channel_count": 4, # EDIT HERE, default is 4, you can change to other number but code will check how many tetrode files are present and set that to channel copunt regardless
        "devices": device_settings, # EDIT HERE
        "implant": implant_settings, # EDIT HERE
    }
    STUDY_SETTINGS["session"] = SESSION_SETTINGS
    settings_dict = STUDY_SETTINGS

    settings_dict["speed_lowerbound"] = 0 
    settings_dict["speed_upperbound"] = 99
    settings_dict["ratemap_dims"] = (32,32)
    settings_dict["disk_arena"] = True
    settings_dict["naming_type"] = "LEC"
    settings_dict["type"] = "object"
    settings_dict["arena_size"] = None
    settings_dict["ses_limit"] = None
    settings_dict["hasObject"] = True

    start_time = time.time()
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(parent=root,title="Please select a data directory.")
    output_folder_path = filedialog.askdirectory(parent=root,title="Please select an output folder.")

    subdirs = np.sort([ f.path for f in os.scandir(folder_path) if f.is_dir() ])
    # sys.stdout = open(r"C:\Users\aaoun\OneDrive - cumc.columbia.edu\Desktop\HussainiLab\neuroscikit\_prototypes\cell_remapping\testlogRadha.txt", "w")
    for subdir in subdirs:
        try:
            study = make_study(subdir,settings_dict=settings_dict)
            study.make_animals()

            main(study, settings_dict, output_folder_path)
     
        except Exception:
            print(traceback.format_exc())
            print("DID NOT WORK FOR DIRECTORY " + str(subdir))

    custom_lec_prefix = "Study1_LEC"
    custom_arena_prefix = "Arena1_LEC"
    study_record = pd.DataFrame(columns=["schema_ref", "data_name", "study_description"])  
    study_record_dict = {"schema_ref":"study","data_name": custom_lec_prefix, "study_description": "LEC recordings of object exploration"}
    study_record_path = os.path.join(output_folder_path, "study_record.xlsx")
    study_record = pd.DataFrame(study_record_dict, index=[0])
    study_record.to_excel(study_record_path, index=False)
    arena_record = pd.DataFrame(columns=["schema_ref", "data_name", "arena_description", "unit_of_measure", "arena_shape", "diameter"])
    arena_record_dict = {"schema_ref":"arena","data_name": custom_arena_prefix, "arena_description": "Cylinder arena for LEC recordings",
                    "unit_of_measure": "m", "arena_shape": "cylinder", "diameter": ".4"}
    arena_record_path = os.path.join(output_folder_path, "arena_record.xlsx")
    arena_record = pd.DataFrame(arena_record_dict, index=[0])
    arena_record.to_excel(arena_record_path, index=False)

    print("COMPLETED ALL FOLDERS")
    print("Total run time: " + str(time.time() - start_time))
